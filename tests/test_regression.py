import csv
import tempfile
import unittest
from pathlib import Path

import openpyxl
import xlwt

from src.nodes.merge_nodes import WorkbookCreateNode, WorkbookAppendNode, SheetCopyNode, WorkbookSaveNode


class TestCsvXlsToXlsxRegression(unittest.TestCase):
    def setUp(self):
        self.temp_dir = tempfile.TemporaryDirectory()
        self.base = Path(self.temp_dir.name)

        self.csv_path = self.base / "sample.csv"
        self.xls_path = self.base / "sample.xls"
        self.xlsx_merged_path = self.base / "sample_merged.xlsx"

        self.out_csv_xlsx = self.base / "out_from_csv.xlsx"
        self.out_xls_xlsx = self.base / "out_from_xls.xlsx"
        self.out_sheet_copy_xlsx = self.base / "out_sheet_copy.xlsx"
        self.out_merged_xlsx = self.base / "out_merged_copy.xlsx"

        self._build_csv()
        self._build_xls()
        self._build_xlsx_with_merged_cells()

    def tearDown(self):
        self.temp_dir.cleanup()

    def _build_csv(self):
        with open(self.csv_path, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.writer(f)
            writer.writerow(["Name", "Amount", "Qty"])
            writer.writerow(["A", "100", "2"])
            writer.writerow(["B", "250.5", "3"])
            writer.writerow(["C", "0", "10"])

    def _build_xls(self):
        wb = xlwt.Workbook()

        sh1 = wb.add_sheet("Data")
        sh1.write(0, 0, "Item")
        sh1.write(0, 1, "Value")
        sh1.write(1, 0, "X")
        sh1.write(1, 1, 123)
        sh1.write(2, 0, "Y")
        sh1.write(2, 1, 456)

        sh2 = wb.add_sheet("More")
        sh2.write(0, 0, "Code")
        sh2.write(0, 1, "Desc")
        sh2.write(1, 0, "C1")
        sh2.write(1, 1, "ok")

        wb.save(str(self.xls_path))

    def _build_xlsx_with_merged_cells(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "MergeSrc"

        ws["A1"] = "Merged Header"
        ws.merge_cells("A1:C1")

        ws["A2"] = "Item"
        ws["B2"] = "Amount"
        ws["C2"] = "Qty"

        ws["A3"] = "Group-1"
        ws.merge_cells("A3:A4")
        ws["B3"] = 100
        ws["C3"] = 2
        ws["B4"] = 250
        ws["C4"] = 3

        wb.save(self.xlsx_merged_path)

    def test_csv_append_then_save_xlsx(self):
        creator = WorkbookCreateNode("c1")
        workbook = creator.execute({})["workbook"]

        appender = WorkbookAppendNode("a1")
        appender.set_param("source_type", "file")
        appender.set_param("file_path", str(self.csv_path))
        appender.set_param("sheet_mode", "first")
        appender.set_param("target_name", "CSV_DATA")
        appender.set_param("csv_delimiter", "comma")
        appender.set_param("csv_encoding", "utf-8-sig")
        appender.set_param("header_row", 0)
        workbook = appender.execute({"workbook": workbook})["workbook"]

        saver = WorkbookSaveNode("s1")
        saver.set_param("output_file", str(self.out_csv_xlsx))
        saver.execute({"workbook": workbook})

        out_wb = openpyxl.load_workbook(self.out_csv_xlsx, data_only=False)
        self.assertIn("CSV_DATA", out_wb.sheetnames)
        ws = out_wb["CSV_DATA"]
        self.assertGreaterEqual(ws.max_row, 4)
        self.assertIsInstance(ws["B2"].value, (int, float))

    def test_xls_append_all_sheets_then_save_xlsx(self):
        creator = WorkbookCreateNode("c2")
        workbook = creator.execute({})["workbook"]

        appender = WorkbookAppendNode("a2")
        appender.set_param("source_type", "file")
        appender.set_param("file_path", str(self.xls_path))
        appender.set_param("sheet_mode", "all")
        appender.set_param("target_name", "")
        workbook = appender.execute({"workbook": workbook})["workbook"]

        saver = WorkbookSaveNode("s2")
        saver.set_param("output_file", str(self.out_xls_xlsx))
        saver.execute({"workbook": workbook})

        out_wb = openpyxl.load_workbook(self.out_xls_xlsx, data_only=False)
        self.assertIn("Data", out_wb.sheetnames)
        self.assertIn("More", out_wb.sheetnames)
        self.assertEqual(out_wb["Data"]["A2"].value, "X")
        self.assertEqual(out_wb["Data"]["B3"].value, 456)

    def test_sheet_copy_csv_to_target_then_save_xlsx(self):
        creator = WorkbookCreateNode("c3")
        workbook = creator.execute({})["workbook"]

        copy_node = SheetCopyNode("sc1")
        copy_node.set_param("file_path", str(self.csv_path))
        copy_node.set_param("sheet_name", "")
        copy_node.set_param("target_sheet", "TARGET")
        copy_node.set_param("copy_mode", "whole")
        copy_node.set_param("write_mode", "overwrite")
        copy_node.set_param("column_mapping", "")
        copy_node.set_param("header_row", 0)
        copy_node.set_param("filter_query", "")
        copy_node.set_param("remove_duplicates", False)
        copy_node.set_param("strip_whitespace", True)
        copy_node.set_param("preserve_formatting", True)
        copy_node.set_param("csv_delimiter", "comma")
        copy_node.set_param("csv_encoding", "utf-8-sig")

        workbook = copy_node.execute({"workbook": workbook})["workbook"]

        saver = WorkbookSaveNode("s3")
        saver.set_param("output_file", str(self.out_sheet_copy_xlsx))
        saver.execute({"workbook": workbook})

        out_wb = openpyxl.load_workbook(self.out_sheet_copy_xlsx, data_only=False)
        self.assertIn("TARGET", out_wb.sheetnames)
        ws = out_wb["TARGET"]
        self.assertGreaterEqual(ws.max_row, 4)
        self.assertIsInstance(ws["B2"].value, (int, float))

    def test_sheet_copy_preserves_merged_cells_from_xlsx(self):
        creator = WorkbookCreateNode("c4")
        workbook = creator.execute({})["workbook"]

        copy_node = SheetCopyNode("sc2")
        copy_node.set_param("file_path", str(self.xlsx_merged_path))
        copy_node.set_param("sheet_name", "MergeSrc")
        copy_node.set_param("target_sheet", "MERGED_OUT")
        copy_node.set_param("copy_mode", "whole")
        copy_node.set_param("write_mode", "overwrite")
        copy_node.set_param("column_mapping", "")
        copy_node.set_param("header_row", 0)
        copy_node.set_param("filter_query", "")
        copy_node.set_param("remove_duplicates", False)
        copy_node.set_param("strip_whitespace", False)
        copy_node.set_param("preserve_formatting", True)
        copy_node.set_param("quick_mode", False)
        copy_node.set_param("csv_delimiter", "auto")
        copy_node.set_param("csv_encoding", "auto")

        workbook = copy_node.execute({"workbook": workbook})["workbook"]

        saver = WorkbookSaveNode("s4")
        saver.set_param("output_file", str(self.out_merged_xlsx))
        saver.execute({"workbook": workbook})

        out_wb = openpyxl.load_workbook(self.out_merged_xlsx, data_only=False)
        self.assertIn("MERGED_OUT", out_wb.sheetnames)
        ws = out_wb["MERGED_OUT"]

        merged_ranges = {str(rng) for rng in ws.merged_cells.ranges}
        self.assertIn("A1:C1", merged_ranges)
        self.assertIn("A3:A4", merged_ranges)


if __name__ == "__main__":
    unittest.main(verbosity=2)
