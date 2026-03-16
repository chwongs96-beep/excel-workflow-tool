import csv
import tempfile
import unittest
from pathlib import Path

import openpyxl
import xlwt

from src.nodes.excel_nodes import ConvertToXlsxNode
from src.nodes.merge_nodes import WorkbookCreateNode, SheetCopyNode, WorkbookSaveNode


class TestConvertToXlsxNode(unittest.TestCase):
    def setUp(self):
        self.temp_dir = tempfile.TemporaryDirectory()
        self.base = Path(self.temp_dir.name)
        self.csv_path = self.base / "source.csv"
        self.xls_path = self.base / "source.xls"
        self.csv_out = self.base / "converted_csv.xlsx"
        self.xls_out = self.base / "converted_xls.xlsx"
        self._build_csv()
        self._build_xls()

    def tearDown(self):
        self.temp_dir.cleanup()

    def _build_csv(self):
        with open(self.csv_path, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.writer(f)
            writer.writerow(["Name", "Amount", "Qty", "Code"])
            writer.writerow(["A", "100", "2", "00123"])
            writer.writerow(["B", "250.5", "3", "00045"])

    def _build_xls(self):
        wb = xlwt.Workbook()
        ws = wb.add_sheet("Data")
        ws.write(0, 0, "Item")
        ws.write(0, 1, "Value")
        ws.write(0, 2, "Ref")
        ws.write(1, 0, "X")
        ws.write(1, 1, 123)
        ws.write(1, 2, "001")
        ws.write(2, 0, "Y")
        ws.write(2, 1, 456.5)
        ws.write(2, 2, "002")
        wb.save(str(self.xls_path))

    def test_convert_csv_writes_numeric_cells(self):
        node = ConvertToXlsxNode("convert_csv")
        node.set_param("file_path", str(self.csv_path))
        node.set_param("output_file", str(self.csv_out))
        node.set_param("sheet_name", "Sheet1")
        node.set_param("header_row", 0)
        node.set_param("csv_delimiter", "comma")
        node.set_param("csv_encoding", "utf-8-sig")

        result = node.execute({})
        wb = openpyxl.load_workbook(result["file_path"], data_only=False)
        ws = wb["Sheet1"]

        self.assertIsInstance(ws["B2"].value, (int, float))
        self.assertIsInstance(ws["C2"].value, (int, float))
        self.assertEqual(ws["D2"].value, "00123")

    def test_convert_xls_writes_numeric_cells(self):
        node = ConvertToXlsxNode("convert_xls")
        node.set_param("file_path", str(self.xls_path))
        node.set_param("output_file", str(self.xls_out))
        node.set_param("sheet_name", "Sheet1")
        node.set_param("header_row", 0)

        result = node.execute({})
        wb = openpyxl.load_workbook(result["file_path"], data_only=False)
        ws = wb["Data"]

        self.assertIsInstance(ws["B2"].value, (int, float))
        self.assertIsInstance(ws["B3"].value, (int, float))
        self.assertEqual(ws["C2"].value, "001")

    def test_convert_csv_then_copy_to_target_xlsx(self):
        convert_node = ConvertToXlsxNode("convert_chain")
        convert_node.set_param("file_path", str(self.csv_path))
        convert_node.set_param("output_file", str(self.csv_out))
        convert_node.set_param("sheet_name", "Data")
        convert_node.set_param("header_row", 0)
        convert_node.set_param("csv_delimiter", "comma")
        convert_node.set_param("csv_encoding", "utf-8-sig")
        convert_result = convert_node.execute({})

        creator = WorkbookCreateNode("create_chain")
        workbook = creator.execute({})["workbook"]

        copy_node = SheetCopyNode("copy_chain")
        copy_node.set_param("file_path", convert_result["file_path"])
        copy_node.set_param("sheet_name", "Data")
        copy_node.set_param("target_sheet", "OUT")
        copy_node.set_param("copy_mode", "whole")
        copy_node.set_param("quick_mode", False)
        copy_node.set_param("write_mode", "overwrite")
        copy_node.set_param("column_mapping", "")
        copy_node.set_param("header_row", 0)
        copy_node.set_param("filter_query", "")
        copy_node.set_param("remove_duplicates", False)
        copy_node.set_param("strip_whitespace", False)
        copy_node.set_param("preserve_formatting", True)
        workbook = copy_node.execute({"workbook": workbook})["workbook"]

        target_xlsx = self.base / "target.xlsx"
        saver = WorkbookSaveNode("save_chain")
        saver.set_param("output_file", str(target_xlsx))
        save_result = saver.execute({"workbook": workbook})

        wb = openpyxl.load_workbook(save_result["file_path"], data_only=False)
        ws = wb["OUT"]

        self.assertTrue(Path(convert_result["file_path"]).exists())
        self.assertTrue(Path(save_result["file_path"]).exists())
        self.assertIn("OUT", wb.sheetnames)
        self.assertIsInstance(ws["B2"].value, (int, float))
        self.assertIsInstance(ws["C2"].value, (int, float))
        self.assertEqual(ws["D2"].value, "00123")


if __name__ == "__main__":
    unittest.main(verbosity=2)
