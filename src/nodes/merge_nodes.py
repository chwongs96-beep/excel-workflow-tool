import shutil
import warnings
from copy import copy as copy_obj
from pathlib import Path
from typing import Any, Dict, List, Union

import openpyxl
import openpyxl.styles
import pandas as pd
from openpyxl.utils import get_column_letter

from .base_node import BaseNode
from .node_registry import register_node
from .utils import (
    clean_unnamed_columns,
    clear_dataframe_excel_range,
    convert_text_to_numeric,
    detect_used_range,
    detect_used_range_from_dataframe,
    get_excel_file,
    make_unique_sheet_name,
    normalize_dataframe_for_excel,
    normalize_excel_value,
    parse_a1_range,
    read_csv_with_options,
    read_excel_with_engine,
    sanitize_sheet_name,
    sanitize_value,
)

# Suppress warnings from pandas and openpyxl
warnings.filterwarnings('ignore', category=UserWarning)
warnings.filterwarnings('ignore', category=FutureWarning)


class StyledSheet:
    """Wrapper to hold sheet info for style-preserved copying"""
    def __init__(self, file_path, sheet_name, df_filtered=None, header_row=0,
                 is_full_copy=False, copy_mode="whole"):
        self.file_path = file_path
        self.sheet_name = sheet_name
        self.df_filtered = df_filtered
        self.header_row = header_row
        self.is_full_copy = is_full_copy
        self.copy_mode = copy_mode


# ============================================================================
# 批量合并节点 (Batch Merge)
# ============================================================================

@register_node
class MergeExcelFilesNode(BaseNode):
    """Node to merge multiple Excel files into one"""
    
    node_type = "merge_excel_files"
    node_name = "批量合并Excel"
    node_category = "批量处理"
    node_description = "将多个Excel文件合并到一个文件中"
    node_color = "#8b5cf6"  # Violet
    
    def _setup_ports(self):
        self.add_output("file_path")
    
    def get_config_ui_schema(self) -> List[Dict[str, Any]]:
        return [
            {
                "key": "base_file",
                "label": "基础文件 (File 1)",
                "type": "file",
                "file_filter": "Excel文件 (*.xlsx *.xls)",
                "required": True
            },
            {
                "key": "files_to_merge",
                "label": "要合并的文件 (File 2, 3, 4...)",
                "type": "file_multiple",
                "file_filter": "Excel文件 (*.xlsx *.xls)",
                "required": True
            },
            {
                "key": "sheet_mode",
                "label": "工作表选择模式",
                "type": "select",
                "options": [
                    {"value": "all", "label": "所有工作表"},
                    {"value": "first", "label": "仅第一个工作表"},
                    {"value": "name", "label": "指定工作表名称"}
                ],
                "default": "all"
            },
            {
                "key": "sheet_name",
                "label": "指定工作表名称 (如果选择)",
                "type": "text",
                "default": "",
                "placeholder": "例如: Sheet1"
            },
            {
                "key": "output_file",
                "label": "输出文件路径",
                "type": "file_save",
                "file_filter": "Excel文件 (*.xlsx)",
                "required": True
            }
        ]
    
    def validate(self) -> tuple[bool, str]:
        base_file = self.get_param("base_file", "")
        if not base_file or not Path(base_file).exists():
            return False, "基础文件是必需的"
            
        files = self.get_param("files_to_merge", "")
        if not files:
            return False, "要合并的文件是必需的"
            
        output_file = self.get_param("output_file", "")
        if not output_file:
            return False, "输出文件路径是必需的"
            
        return True, ""
    
    def execute(self, input_data: Dict[str, Any]) -> Dict[str, Any]:
        base_file = self.get_param("base_file")
        files_str = self.get_param("files_to_merge", "")
        files_to_merge = [f.strip() for f in files_str.split('\n') if f.strip()]
        output_file = self.get_param("output_file")
        
        sheet_mode = self.get_param("sheet_mode", "all")
        target_sheet_name = self.get_param("sheet_name", "")
        
        def _clean_dict(dfs):
            return {k: clean_unnamed_columns(v) for k, v in dfs.items()}

        # Helper to read sheets based on mode
        def read_sheets(file_path):
            if sheet_mode == "all":
                return _clean_dict(read_excel_with_engine(file_path, sheet_name=None, header=0, dtype=object))
            elif sheet_mode == "first":
                df = clean_unnamed_columns(read_excel_with_engine(file_path, sheet_name=0, header=0, dtype=object))
                return {"Sheet1": df}
            elif sheet_mode == "name":
                if not target_sheet_name:
                    return _clean_dict(read_excel_with_engine(file_path, sheet_name=None, header=0, dtype=object))
                try:
                    df = clean_unnamed_columns(read_excel_with_engine(file_path, sheet_name=target_sheet_name, header=0, dtype=object))
                    return {target_sheet_name: df}
                except Exception:
                    print(f"Warning: Sheet '{target_sheet_name}' not found in {file_path}")
                    return {}
            return {}

        # Read base file sheets
        try:
            base_dfs = _clean_dict(read_excel_with_engine(base_file, sheet_name=None, header=0, dtype=object))
        except Exception as e:
            raise ValueError(f"读取基础文件失败: {e}")
            
        merged_sheets = {}
        
        # Add base sheets first
        for sheet_name, df in base_dfs.items():
            merged_sheets[sheet_name] = df
            
        # Process other files
        for i, file_path in enumerate(files_to_merge):
            try:
                dfs = read_sheets(file_path)
                
                for sheet_name, df in dfs.items():
                    # Create a unique sheet name
                    # If mode is 'first', we might want to name it after the file
                    if sheet_mode == "first":
                        p = Path(file_path)
                        new_sheet_name = p.stem
                    else:
                        new_sheet_name = sheet_name
                        
                    # Conflict resolution
                    if new_sheet_name in merged_sheets:
                        p = Path(file_path)
                        new_sheet_name = f"{p.stem}_{sheet_name}"
                        
                    if new_sheet_name in merged_sheets:
                        counter = 1
                        while f"{new_sheet_name}_{counter}" in merged_sheets:
                            counter += 1
                        new_sheet_name = f"{new_sheet_name}_{counter}"
                    
                    merged_sheets[new_sheet_name] = df
                    
            except Exception as e:
                print(f"Warning: Failed to read {file_path}: {e}")
        
        # Write to output file (normalize so numeric-looking strings become real numbers)
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            for sheet_name, df in merged_sheets.items():
                safe_name = sheet_name[:31]
                df = normalize_dataframe_for_excel(df)
                df.to_excel(writer, sheet_name=safe_name, index=False)
                
        return {"file_path": output_file}


# ============================================================================
# 灵活工作流节点 (Flexible Workflow)
# ============================================================================

@register_node
class WorkbookCreateNode(BaseNode):
    """Node to start a workbook workflow"""
    
    node_type = "workbook_create"
    node_name = "创建工作簿(输入)"
    node_category = "灵活合并"
    node_description = "开始一个新的工作簿，可选从现有文件加载（注意：.xls格式不支持样式保留）"
    node_color = "#22c55e"  # Green
    
    def _setup_ports(self):
        self.add_output("workbook")
    
    def get_config_ui_schema(self) -> List[Dict[str, Any]]:
        return [
            {
                "key": "base_file",
                "label": "基础文件 (可选)",
                "type": "file",
                "file_filter": "Excel文件 (*.xlsx *.xls)",
                "placeholder": "留空则创建一个空工作簿"
            }
        ]
    
    def execute(self, input_data: Dict[str, Any]) -> Dict[str, Any]:
        base_file = self.get_param("base_file", "")
        workbook = {}
        
        if base_file and Path(base_file).exists():
            try:
                # Warn if .xls file is used
                if str(base_file).lower().endswith('.xls'):
                    self.report_progress(f"⚠️ .xls格式不支持样式保留: {Path(base_file).name}")
                
                dfs = read_excel_with_engine(base_file, sheet_name=None, header=0, dtype=object)
                
                for sheet_name, df in dfs.items():
                    df = clean_unnamed_columns(df)
                    workbook[sheet_name] = StyledSheet(base_file, sheet_name, df, is_full_copy=True)
                    
            except Exception as e:
                raise ValueError(f"读取基础文件失败: {e}")
        
        return {"workbook": workbook}


@register_node
class WorkbookAppendNode(BaseNode):
    """Node to append a sheet from another file"""
    
    node_type = "workbook_append"
    node_name = "追加工作表"
    node_category = "灵活合并"
    node_description = "从另一个Excel文件读取工作表并添加到当前工作簿（注意：.xls格式不支持样式保留）"
    node_color = "#8b5cf6"  # Violet
    
    def _setup_ports(self):
        self.add_input("workbook")
        self.add_input("file_path")  # Optional input for dynamic source file
        self.add_output("workbook")
    
    def get_config_ui_schema(self) -> List[Dict[str, Any]]:
        return [
            {
                "key": "source_type",
                "label": "来源类型",
                "type": "select",
                "options": [
                    {"value": "file", "label": "指定文件"},
                    {"value": "search", "label": "搜索文件夹"}
                ],
                "default": "file"
            },
            {
                "key": "file_path",
                "label": "文件路径 (指定文件时)",
                "type": "file",
                "file_filter": "Excel/CSV文件 (*.xlsx *.xls *.csv)",
                "placeholder": "请选择要追加的Excel或CSV文件 (或连接输入节点)"
            },
            {
                "key": "folder_path",
                "label": "文件夹路径 (搜索时)",
                "type": "directory",
                "placeholder": "请选择要搜索的文件夹"
            },
            {
                "key": "keyword",
                "label": "文件名关键字 (搜索时)",
                "type": "text",
                "placeholder": "例如: 2023年报 (留空则匹配所有文件)"
            },
            {
                "key": "sheet_mode",
                "label": "选择模式",
                "type": "select",
                "options": [
                    {"value": "first", "label": "第一个工作表"},
                    {"value": "name", "label": "指定名称"},
                    {"value": "all", "label": "所有工作表"}
                ],
                "default": "first"
            },
            {
                "key": "sheet_name",
                "label": "源工作表名称 (指定名称时)",
                "type": "sheet_selector",
                "dependency": "file_path",
                "default": ""
            },
            {
                "key": "target_name",
                "label": "目标工作表名称 (可选)",
                "type": "text",
                "default": "",
                "placeholder": "留空则自动命名 (文件名/原名)"
            },
            {
                "key": "csv_delimiter",
                "label": "CSV分隔符 (仅CSV)",
                "type": "select",
                "options": [
                    {"value": "auto", "label": "自动检测"},
                    {"value": "comma", "label": "逗号 (,)"},
                    {"value": "tab", "label": "制表符 (Tab)"},
                    {"value": "semicolon", "label": "分号 (;)"},
                    {"value": "pipe", "label": "竖线 (|)"},
                    {"value": "space", "label": "空格"}
                ],
                "default": "auto"
            },
            {
                "key": "csv_encoding",
                "label": "CSV编码 (仅CSV)",
                "type": "select",
                "options": [
                    {"value": "auto", "label": "自动检测"},
                    {"value": "utf-8", "label": "UTF-8"},
                    {"value": "gbk", "label": "GBK/GB18030"},
                    {"value": "utf-8-sig", "label": "UTF-8-SIG"}
                ],
                "default": "auto"
            },
            {
                "key": "header_row",
                "label": "标题所在行 (仅CSV, 从0开始)",
                "type": "number",
                "default": 0,
                "min": 0
            }
        ]
    
    def validate(self) -> tuple[bool, str]:
        source_type = self.get_param("source_type", "file")
        # Relax validation to allow dynamic input via connection
        if source_type == "search":
            if not self.get_param("folder_path"):
                return False, "搜索文件夹是必需的"
        return True, ""
    
    def execute(self, input_data: Dict[str, Any]) -> Dict[str, Any]:
        workbook = input_data.get("workbook")
        if workbook is None:
            workbook = {} # Start fresh if no input
        else:
            workbook = workbook.copy() # Shallow copy dict
            
        # Determine file path
        source_type = self.get_param("source_type", "file")
        file_path = ""
        
        # Check input port first for dynamic file path
        if "file_path" in input_data and input_data["file_path"]:
            file_path = input_data["file_path"]
        elif source_type == "file":
            file_path = self.get_param("file_path")
        else:
            folder_path = self.get_param("folder_path")
            keyword = self.get_param("keyword", "")
            
            if not folder_path or not Path(folder_path).exists():
                raise ValueError(f"文件夹不存在: {folder_path}")
                
            p = Path(folder_path)
            # Find excel/csv files
            files = list(p.glob("*.xlsx")) + list(p.glob("*.xls")) + list(p.glob("*.csv"))
            
            # Filter by keyword
            if keyword:
                files = [f for f in files if keyword in f.name]
            
            if not files:
                raise ValueError(f"在 {folder_path} 中未找到匹配 '{keyword}' 的Excel/CSV文件")
            
            # Sort by name and take first
            files.sort(key=lambda f: f.name)
            file_path = str(files[0])
            print(f"Found file by keyword '{keyword}': {file_path}")

        sheet_mode = self.get_param("sheet_mode", "first")
        src_sheet_name = self.get_param("sheet_name", "")
        target_name = self.get_param("target_name", "")
        
        csv_delimiter = self.get_param("csv_delimiter", "auto")
        csv_encoding = self.get_param("csv_encoding", "auto")
        header_row = self.get_param("header_row", 0)
        
        try:
            is_csv = str(file_path).lower().endswith('.csv')
            is_xls = str(file_path).lower().endswith('.xls')
            
            # Warn if .xls file is used
            if is_xls:
                self.report_progress(f"⚠️ .xls格式不支持样式保留: {Path(file_path).name}")
            
            if is_csv:
                df = read_csv_with_options(
                    file_path, csv_encoding, csv_delimiter, header_row,
                    dtype=object, auto_convert_numeric=False,
                )
                df = clean_unnamed_columns(df)
                default_name = Path(file_path).stem
                
                # Determine target name
                if target_name:
                    t_name = target_name
                else:
                    t_name = default_name
                
                # Ensure unique
                base_t_name = t_name
                counter = 1
                while t_name in workbook:
                    t_name = f"{base_t_name}_{counter}"
                    counter += 1
                
                workbook[t_name] = df
                
            elif sheet_mode == "all":
                dfs = read_excel_with_engine(file_path, sheet_name=None, header=0, dtype=object)
                for name, df in dfs.items():
                    df = clean_unnamed_columns(df)
                    if target_name:
                        t_name = name
                    else:
                        t_name = name
                        
                    # Conflict resolution
                    if t_name in workbook:
                        p = Path(file_path)
                        t_name = f"{p.stem}_{name}"
                    
                    # Ensure unique
                    base_t_name = t_name
                    counter = 1
                    while t_name in workbook:
                        t_name = f"{base_t_name}_{counter}"
                        counter += 1
                        
                    workbook[t_name] = StyledSheet(file_path, name, df, is_full_copy=True)
                    
            else:
                # Single sheet
                actual_sheet_name = src_sheet_name
                if sheet_mode == "first":
                    xl = get_excel_file(file_path)
                    actual_sheet_name = xl.sheet_names[0]
                    df = clean_unnamed_columns(read_excel_with_engine(file_path, sheet_name=0, header=0, dtype=object))
                    default_name = Path(file_path).stem
                else: # name
                    if not src_sheet_name:
                        raise ValueError("未指定源工作表名称")
                    df = clean_unnamed_columns(read_excel_with_engine(file_path, sheet_name=src_sheet_name, header=0, dtype=object))
                    default_name = src_sheet_name
                    actual_sheet_name = src_sheet_name
                
                # Determine target name
                if target_name:
                    t_name = target_name
                else:
                    t_name = default_name
                
                # Ensure unique
                base_t_name = t_name
                counter = 1
                while t_name in workbook:
                    t_name = f"{base_t_name}_{counter}"
                    counter += 1
                
                # Wrap in StyledSheet
                workbook[t_name] = StyledSheet(file_path, actual_sheet_name, df, is_full_copy=True)
                
        except Exception as e:
            raise ValueError(f"读取文件失败 {file_path}: {e}")
            
        return {"workbook": workbook}


@register_node
class SheetCopyNode(BaseNode):
    """Node to copy data between sheets with advanced modes"""
    
    node_type = "sheet_copy"
    node_name = "复制/合并数据"
    node_category = "灵活合并"
    node_description = "将Excel/CSV数据复制到工作簿，支持列映射和空值检查（注意：.xls格式不支持样式保留）"
    node_color = "#f59e0b"  # Amber
    
    def _setup_ports(self):
        self.add_input("workbook")
        self.add_input("file_path")  # Optional input for dynamic source file
        self.add_output("workbook")
    
    def get_config_ui_schema(self) -> List[Dict[str, Any]]:
        return [
            {
                "key": "file_path",
                "label": "来源文件",
                "type": "file",
                "file_filter": "Excel/CSV文件 (*.xlsx *.xls *.csv)",
                "required": True,
                "placeholder": "请选择来源文件 (或连接输入节点)"
            },
            {
                "key": "sheet_name",
                "label": "来源工作表 (Excel)",
                "type": "sheet_selector",
                "dependency": "file_path",
                "default": "",
                "placeholder": "CSV文件可忽略此项"
            },
            {
                "key": "csv_delimiter",
                "label": "CSV分隔符 (仅CSV)",
                "type": "select",
                "options": [
                    {"value": "auto", "label": "自动检测"},
                    {"value": "comma", "label": "逗号 (,)"},
                    {"value": "tab", "label": "制表符 (Tab)"},
                    {"value": "semicolon", "label": "分号 (;)"},
                    {"value": "pipe", "label": "竖线 (|)"},
                    {"value": "space", "label": "空格"}
                ],
                "default": "auto"
            },
            {
                "key": "csv_encoding",
                "label": "CSV编码 (仅CSV)",
                "type": "select",
                "options": [
                    {"value": "auto", "label": "自动检测"},
                    {"value": "utf-8", "label": "UTF-8"},
                    {"value": "gbk", "label": "GBK/GB18030"},
                    {"value": "utf-8-sig", "label": "UTF-8-SIG"}
                ],
                "default": "auto"
            },
            {
                "key": "header_row",
                "label": "标题所在行 (从0开始)",
                "type": "number",
                "default": 0,
                "min": 0
            },
            {
                "key": "target_sheet",
                "label": "目标工作表名称",
                "type": "sheet_selector",
                "dependency": "__upstream__",
                "required": True,
                "placeholder": "选择或输入目标工作表名称"
            },
            {
                "key": "copy_mode",
                "label": "复制模式",
                "type": "select",
                "options": [
                    {"value": "whole", "label": "整页复制"},
                    {"value": "used_range", "label": "自动检测有值区域"},
                    {"value": "columns", "label": "指定列到列"},
                    {"value": "no_blank", "label": "自动检查值(无空白)"}
                ],
                "default": "whole"
            },
            {
                "key": "quick_mode",
                "label": "⚡ 快速模式(推荐)",
                "type": "checkbox",
                "default": True
            },
            {
                "key": "column_mapping",
                "label": "列映射 (仅指定列模式)",
                "type": "text",
                "placeholder": "格式: 源列A=目标列B; 源列C=目标列D"
            },
            {
                "key": "filter_query",
                "label": "行过滤条件 (Pandas Query)",
                "type": "text",
                "placeholder": "例如: 状态 == '完成' and 金额 > 1000"
            },
            {
                "key": "remove_duplicates",
                "label": "去除重复行",
                "type": "checkbox",
                "default": False
            },
            {
                "key": "strip_whitespace",
                "label": "去除文本首尾空格",
                "type": "checkbox",
                "default": True
            },
            {
                "key": "preserve_formatting",
                "label": "保留原始Excel格式 (较慢)",
                "type": "checkbox",
                "default": True
            },
            {
                "key": "write_mode",
                "label": "写入方式",
                "type": "select",
                "options": [
                    {"value": "overwrite", "label": "覆盖目标表"},
                    {"value": "append", "label": "追加到末尾"}
                ],
                "default": "overwrite"
            }
        ]
    
    def validate(self) -> tuple[bool, str]:
        # Relax validation to allow dynamic input via connection
        # if not self.get_param("file_path"):
        #     return False, "来源文件是必需的"
        if not self.get_param("target_sheet"):
            return False, "目标工作表名称是必需的"
        
        mode = self.get_param("copy_mode")
        quick_mode = self.get_param("quick_mode", True)
        if (not quick_mode) and mode == "columns" and not self.get_param("column_mapping"):
            return False, "指定列模式下需要填写列映射"
            
        return True, ""
    
    def execute(self, input_data: Dict[str, Any]) -> Dict[str, Any]:
        workbook = input_data.get("workbook")
        if workbook is None:
            workbook = {}
        else:
            workbook = workbook.copy()
            
        # Check input port first for dynamic file path
        if "file_path" in input_data and input_data["file_path"]:
            file_path = input_data["file_path"]
        else:
            file_path = self.get_param("file_path")
            
        src_sheet_name = self.get_param("sheet_name")
        target_sheet = self.get_param("target_sheet")
        copy_mode = self.get_param("copy_mode", "whole")
        write_mode = self.get_param("write_mode", "overwrite")
        col_mapping_str = self.get_param("column_mapping", "")
        quick_mode = self.get_param("quick_mode", True)
        
        header_row = self.get_param("header_row", 0)
        filter_query = self.get_param("filter_query", "")
        remove_duplicates = self.get_param("remove_duplicates", False)
        strip_whitespace = self.get_param("strip_whitespace", True)
        preserve_formatting = self.get_param("preserve_formatting", True)
        
        csv_delimiter = self.get_param("csv_delimiter", "auto")
        csv_encoding = self.get_param("csv_encoding", "auto")

        # Quick mode: apply fastest safe defaults with one click
        if quick_mode:
            copy_mode = "whole"
            write_mode = "overwrite"
            filter_query = ""
            remove_duplicates = False
            strip_whitespace = False
            preserve_formatting = False
            self.report_progress("⚡ 已启用快速模式: 整页复制 + 覆盖写入 + 关闭格式保留/过滤/去重")
        
        # 1. Read Source Data
        try:
            is_csv = str(file_path).lower().endswith('.csv')
            if is_csv:
                df = read_csv_with_options(
                    file_path, csv_encoding, csv_delimiter, header_row,
                    dtype=object, auto_convert_numeric=False,
                )
                df = clean_unnamed_columns(df)
                self.report_progress(f"CSV读取成功: {len(df)}行 x {len(df.columns)}列")
            else:
                if not src_sheet_name:
                    df = read_excel_with_engine(file_path, sheet_name=0, header=header_row, dtype=object)
                else:
                    df = read_excel_with_engine(file_path, sheet_name=src_sheet_name, header=header_row, dtype=object)
                df = clean_unnamed_columns(df)
                self.report_progress(f"Excel读取成功: {len(df)}行 x {len(df.columns)}列")
        except Exception as e:
            raise ValueError(f"读取来源文件失败: {e}")
            
        # 2. Pre-process Data (Cleaning & Filtering)
        try:
            # Strip whitespace from string columns
            if strip_whitespace:
                for col in df.columns:
                    if df[col].dtype == "object":
                        try:
                            df[col] = df[col].str.strip()
                        except AttributeError:
                            # Column contains non-string objects, skip
                            pass
                # Also strip column names if they are strings
                df.columns = df.columns.map(lambda x: x.strip() if isinstance(x, str) else x)
                
            # Filter rows
            if filter_query:
                try:
                    # Support simple syntax like: 状态 == '完成'
                    # Pandas query uses the dataframe's columns
                    df = df.query(filter_query)
                except Exception as e:
                    raise ValueError(f"行过滤条件错误: {e}")
            
            # Remove duplicates
            if remove_duplicates:
                before_count = len(df)
                df = df.drop_duplicates()
                after_count = len(df)
                if before_count != after_count:
                    self.report_progress(f"去重: 删除了 {before_count - after_count} 个重复行")

            # 3. Process Data based on Mode
            if copy_mode == "no_blank":
                # Remove rows where all elements are NaN
                before_count = len(df)
                df = df.dropna(how='all')
                after_count = len(df)
                if before_count != after_count:
                    self.report_progress(f"删除空行: {before_count - after_count} 行")
                
            elif copy_mode == "columns":
                # Parse mapping: "A=B; C=D" or "Name=Name"
                if not col_mapping_str or not col_mapping_str.strip():
                    raise ValueError("列映射模式需要指定列映射规则")
                    
                mappings = [m.strip() for m in col_mapping_str.split(';') if m.strip()]
                
                if not mappings:
                    raise ValueError("列映射规则为空，请检查格式")
                
                new_df = pd.DataFrame()
                mapped_count = 0
                
                for m in mappings:
                    if '=' in m:
                        src_col, tgt_col = m.split('=', 1)
                        src_col = src_col.strip()
                        tgt_col = tgt_col.strip()
                    else:
                        # If no =, assume src = tgt
                        src_col = m.strip()
                        tgt_col = m.strip()
                    
                    # Check if src_col exists (by name)
                    if src_col in df.columns:
                        new_df[tgt_col] = df[src_col].copy()
                        mapped_count += 1
                    else:
                        # Try by index if integer?
                        if src_col.isdigit() and int(src_col) < len(df.columns):
                            new_df[tgt_col] = df.iloc[:, int(src_col)].copy()
                            mapped_count += 1
                        else:
                            warning_msg = f"⚠️ 列 '{src_col}' 未找到，跳过"
                            self.report_progress(warning_msg)
                
                if mapped_count == 0:
                    raise ValueError(f"没有成功映射任何列。可用列: {list(df.columns)}")
                    
                self.report_progress(f"列映射: 成功映射 {mapped_count}/{len(mappings)} 列")
                df = new_df

            # Validate final dataframe before writing
            if df is None or df.empty:
                raise ValueError(f"处理后的数据为空，请检查过滤条件和列映射设置")
            
            self.report_progress(f"准备写入: {len(df)}行 x {len(df.columns)}列 → 工作表 '{target_sheet}'")

            # 4. Write to Target
            if preserve_formatting and not is_csv:
                # Warn if .xls file is used with style preservation
                if str(file_path).lower().endswith('.xls'):
                    self.report_progress(f"⚠️ .xls格式不支持样式保留: {Path(file_path).name}")
                
                # Check if this is a full copy (optimization)
                # used_range is also a "full copy" in the sense that we bypass
                # DataFrame iteration and copy directly via openpyxl, but with
                # a trimmed boundary.
                is_full_copy = (
                    copy_mode in ("whole", "used_range") and
                    not filter_query and 
                    not remove_duplicates and 
                    not strip_whitespace
                )
                
                def _make_styled(fp, sn, d, hr):
                    return StyledSheet(fp, sn, d, hr,
                                       is_full_copy=is_full_copy,
                                       copy_mode=copy_mode)

                # Use StyledSheet wrapper
                if target_sheet in workbook:
                    existing = workbook[target_sheet]
                    if isinstance(existing, list):
                        existing.append(_make_styled(file_path, src_sheet_name, df, header_row))
                    else:
                        if write_mode == "append":
                            workbook[target_sheet] = [existing, _make_styled(file_path, src_sheet_name, df, header_row)]
                        else:
                            workbook[target_sheet] = _make_styled(file_path, src_sheet_name, df, header_row)
                else:
                    workbook[target_sheet] = _make_styled(file_path, src_sheet_name, df, header_row)
                    
            else:
                # Standard DataFrame mode
                if target_sheet in workbook and write_mode == "append":
                    target_data = workbook[target_sheet]
                    # If target is StyledSheet, we can't easily append DataFrame to it without breaking style logic
                    # For now, if mixing, convert everything to DataFrame (lose styles)
                    if isinstance(target_data, StyledSheet):
                        target_data = target_data.df_filtered
                    elif isinstance(target_data, list) and len(target_data) > 0 and isinstance(target_data[0], StyledSheet):
                        # Concatenate all StyledSheets DFs
                        dfs = [s.df_filtered for s in target_data]
                        target_data = pd.concat(dfs, ignore_index=True)
                    
                    if isinstance(target_data, pd.DataFrame):
                        combined_df = pd.concat([target_data, df], ignore_index=True)
                        workbook[target_sheet] = combined_df
                        self.report_progress(f"追加模式: 合并后共 {len(combined_df)} 行")
                else:
                    # Overwrite or Create new
                    workbook[target_sheet] = df
                    self.report_progress(f"覆盖模式: 已写入 {len(df)} 行到工作表 '{target_sheet}'")
        
        except Exception as e:
            raise ValueError(f"处理文件失败 [{file_path}]: {e}")
            
        return {"workbook": workbook}


@register_node
class SheetClearRangeNode(BaseNode):
    """Clear cell contents in workbook sheets (DataFrame grid: row 1 = headers)."""

    node_type = "sheet_clear_range"
    node_name = "清理工作表内容"
    node_category = "灵活合并"
    node_description = (
        "清空指定工作表中的单元格内容：支持自定义 A1 区域，或按数据表自动检测有值区域。"
        "输出为普通数据表，不再保留原 Excel 样式。"
    )
    node_color = "#64748b"

    def _setup_ports(self):
        self.add_input("workbook")
        self.add_output("workbook")

    def get_config_ui_schema(self) -> List[Dict[str, Any]]:
        return [
            {
                "key": "target_scope",
                "label": "目标工作表",
                "type": "select",
                "options": [
                    {"value": "single", "label": "指定一张工作表"},
                    {"value": "all", "label": "全部工作表"},
                ],
                "default": "single",
            },
            {
                "key": "sheet_name",
                "label": "工作表名称",
                "type": "sheet_selector",
                "dependency": "__upstream__",
                "placeholder": "选择或输入要清理的工作表",
            },
            {
                "key": "clear_mode",
                "label": "清理范围",
                "type": "select",
                "options": [
                    {"value": "custom", "label": "自定义 (A1 区域)"},
                    {"value": "used_range", "label": "自动检测有值区域"},
                    {"value": "all_data", "label": "整张表数据 (保留列标题行)"},
                    {"value": "all_with_header", "label": "整张表含列标题"},
                ],
                "default": "used_range",
            },
            {
                "key": "range_a1",
                "label": "Excel 区域 (如 B2:Z99)",
                "type": "text",
                "default": "",
                "placeholder": "仅“自定义”模式需要，例如 A2:H100",
            },
        ]

    def validate(self) -> tuple[bool, str]:
        if self.get_param("target_scope", "single") == "single":
            if not (self.get_param("sheet_name") or "").strip():
                return False, "请指定要清理的工作表名称"
        mode = self.get_param("clear_mode", "used_range")
        if mode == "custom":
            if not (self.get_param("range_a1") or "").strip():
                return False, "自定义模式请填写 Excel 区域"
        return True, ""

    @staticmethod
    def _workbook_item_to_dataframe(item: Any) -> pd.DataFrame:
        # No copy for plain DataFrame: clear_dataframe_excel_range copies internally.
        if isinstance(item, pd.DataFrame):
            return item
        if isinstance(item, StyledSheet):
            df0 = item.df_filtered
            if df0 is not None and (len(df0) > 0 or len(df0.columns) > 0):
                return df0
            return read_excel_with_engine(
                item.file_path,
                sheet_name=item.sheet_name,
                header=item.header_row,
                dtype=object,
            )
        if isinstance(item, list):
            if not item:
                return pd.DataFrame()
            dfs: List[pd.DataFrame] = []
            for x in item:
                dfs.append(SheetClearRangeNode._workbook_item_to_dataframe(x))
            return pd.concat(dfs, ignore_index=True)
        raise ValueError(f"无法清理此类型的工作表数据: {type(item).__name__}")

    def execute(self, input_data: Dict[str, Any]) -> Dict[str, Any]:
        workbook = input_data.get("workbook")
        if workbook is None:
            raise ValueError("没有接收到工作簿数据")
        workbook = dict(workbook)

        target_scope = self.get_param("target_scope", "single")
        sheet_name_param = (self.get_param("sheet_name") or "").strip()
        clear_mode = self.get_param("clear_mode", "used_range")
        range_a1 = (self.get_param("range_a1") or "").strip()

        if target_scope == "single":
            if not sheet_name_param:
                raise ValueError("请指定工作表名称")
            keys = [sheet_name_param]
            for k in keys:
                if k not in workbook:
                    raise ValueError(f"工作簿中没有名为「{k}」的工作表")
        else:
            keys = list(workbook.keys())

        for key in keys:
            item = workbook[key]
            try:
                df = self._workbook_item_to_dataframe(item)
            except ValueError as e:
                raise ValueError(f"[{key}] {e}") from e

            nrows = len(df)
            ncols = len(df.columns)

            if clear_mode == "custom":
                min_r, max_r, min_c, max_c = parse_a1_range(range_a1)
                clear_headers = min_r <= 1 <= max_r
                cleared = clear_dataframe_excel_range(
                    df, min_r, max_r, min_c, max_c, clear_header_cells=clear_headers,
                )
            elif clear_mode == "used_range":
                min_r, max_r, min_c, max_c = detect_used_range_from_dataframe(df)
                if ncols == 0 or ((min_r, max_r, min_c, max_c) == (1, 1, 1, 1) and nrows == 0):
                    cleared = df.copy()
                else:
                    clear_headers = min_r <= 1
                    cleared = clear_dataframe_excel_range(
                        df, min_r, max_r, min_c, max_c, clear_header_cells=clear_headers,
                    )
            elif clear_mode == "all_data":
                if ncols == 0 or nrows == 0:
                    cleared = df.copy()
                else:
                    cleared = clear_dataframe_excel_range(
                        df, 2, 1 + nrows, 1, ncols, clear_header_cells=False,
                    )
            elif clear_mode == "all_with_header":
                if ncols == 0:
                    cleared = df.copy()
                else:
                    last_excel_row = 1 + nrows
                    cleared = clear_dataframe_excel_range(
                        df, 1, last_excel_row, 1, ncols, clear_header_cells=True,
                    )
            else:
                raise ValueError(f"未知的清理模式: {clear_mode}")

            workbook[key] = clean_unnamed_columns(cleared)
            self.report_progress(
                f"已清理「{key}」: 模式={clear_mode}, 形状={len(workbook[key])}×{len(workbook[key].columns)}",
            )

        return {"workbook": workbook}


@register_node
class WorkbookSaveNode(BaseNode):
    """Node to save the workbook"""
    
    node_type = "workbook_save"
    node_name = "保存工作簿(输出)"
    node_category = "灵活合并"
    node_description = "将工作簿保存到文件"
    node_color = "#ef4444"  # Red
    
    def _setup_ports(self):
        self.add_input("workbook")
        self.add_output("file_path")
    
    def get_config_ui_schema(self) -> List[Dict[str, Any]]:
        return [
            {
                "key": "output_file",
                "label": "保存路径",
                "type": "file_save",
                "file_filter": "Excel文件 (*.xlsx)",
                "required": True
            }
        ]
    
    def validate(self) -> tuple[bool, str]:
        if not self.get_param("output_file"):
            return False, "保存路径是必需的"
        return True, ""
    
    def execute(self, input_data: Dict[str, Any]) -> Dict[str, Any]:
        workbook = input_data.get("workbook")
        if workbook is None:
            raise ValueError("没有接收到工作簿数据")
            
        output_file = self.get_param("output_file")
        
        # Ensure file extension is .xlsx
        if not str(output_file).lower().endswith('.xlsx'):
            output_file = str(output_file) + '.xlsx'
        
        if not workbook:
            workbook = {"Sheet1": pd.DataFrame()}
            
        # Check if we have any StyledSheets
        has_styles = False
        for val in workbook.values():
            if isinstance(val, StyledSheet) or (isinstance(val, list) and len(val) > 0 and isinstance(val[0], StyledSheet)):
                has_styles = True
                break
        
        self._source_wb_cache = {} # Initialize cache
        try:
            if has_styles:
                self._save_with_styles(output_file, workbook)
            else:
                self._save_standard(output_file, workbook)
        finally:
            # Close cached openpyxl workbooks to release file handles
            for wb in self._source_wb_cache.values():
                try:
                    wb.close()
                except Exception:
                    pass
            self._source_wb_cache = {}
                
        return {"file_path": output_file}

    def _save_standard(self, output_file, workbook):
        """Save using standard Pandas (no style preservation)"""
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            used_names = set()
            for sheet_name, df in workbook.items():
                if isinstance(df, pd.DataFrame):
                    safe_name = make_unique_sheet_name(sheet_name, used_names)
                    used_names.add(safe_name)
                    df = clean_unnamed_columns(df)
                    df = normalize_dataframe_for_excel(df)
                    df.to_excel(writer, sheet_name=safe_name, index=False)

    def _save_with_styles(self, output_file, workbook):
        """Save using OpenPyXL to preserve styles"""
        
        # Optimization: Check if we can use a template file (Base File)
        # This avoids slow cell-by-cell copying for untouched sheets
        template_file = self._find_template_file(workbook)
        
        if template_file:
            try:
                # Copy template to output
                shutil.copy2(template_file, output_file)
                wb = openpyxl.load_workbook(output_file)
                
                # Track which sheets we have handled (updated or verified as existing)
                handled_sheets = set()
                used_names = set(wb.sheetnames)
                
                # Update/Add sheets
                for sheet_name, data in workbook.items():
                    sanitized = sanitize_sheet_name(sheet_name)

                    # Check if this is the original sheet from template (unmodified)
                    is_original = False
                    if isinstance(data, StyledSheet):
                        if (data.file_path == template_file and
                            sanitized == sanitize_sheet_name(data.sheet_name) and
                            data.is_full_copy and
                            sanitized in wb.sheetnames):
                            is_original = True

                    if is_original:
                        # Already in the template file — keep as-is
                        used_names.add(sanitized)
                        handled_sheets.add(sanitized)
                        continue

                    # Not original: if a sheet with this name already exists in the
                    # template, we overwrite it and must keep the same name.
                    # Calling make_unique_sheet_name here would wrongly produce
                    # e.g. Sheet1_1 because *sanitized* is already in used_names.
                    if sanitized in wb.sheetnames:
                        safe_name = sanitized
                    else:
                        safe_name = make_unique_sheet_name(sheet_name, used_names)

                    used_names.add(safe_name)
                    handled_sheets.add(safe_name)

                    # If it exists in template (but we are overwriting it), remove it first.
                    if safe_name in wb.sheetnames:
                        wb.remove(wb[safe_name])

                    # Create new sheet
                    target_ws = wb.create_sheet(title=safe_name)

                    # Write data
                    self._write_items_to_sheet(data, target_ws)
                
                # Remove sheets that are in template but not in workbook (deleted)
                for sheet in list(wb.sheetnames):
                    if sheet not in handled_sheets:
                        wb.remove(wb[sheet])
                
                wb.save(output_file)
                return
                
            except Exception as e:
                print(f"Template optimization failed, falling back to slow copy: {e}")
                # Fallback to standard creation if optimization fails
                # CRITICAL: Reset template_file to None so we don't skip 'original' sheets
                template_file = None
        
        # Create new workbook (Fallback or if no template)
        wb = openpyxl.Workbook()
        # Remove default sheet
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])
            
        for sheet_name, data in workbook.items():
            safe_name = make_unique_sheet_name(sheet_name, set(wb.sheetnames))
            target_ws = wb.create_sheet(title=safe_name)
            self._write_items_to_sheet(data, target_ws)
        
        wb.save(output_file)

    def _find_template_file(self, workbook):
        """Find a potential template file (Base File) from workbook data"""
        # We look for the most common file path among StyledSheets that are full copies
        # Or just the first one if it covers most sheets.
        # Simple heuristic: If there is at least one StyledSheet with is_full_copy=True,
        # use its file path.
        
        for data in workbook.values():
            if isinstance(data, StyledSheet) and data.is_full_copy:
                if Path(data.file_path).exists() and not str(data.file_path).lower().endswith('.xls'):
                    return data.file_path
            elif isinstance(data, list) and len(data) > 0:
                if isinstance(data[0], StyledSheet) and data[0].is_full_copy:
                    if Path(data[0].file_path).exists() and not str(data[0].file_path).lower().endswith('.xls'):
                        return data[0].file_path
        return None

    def _write_items_to_sheet(self, data, target_ws):
        """Write data items (StyledSheet or DataFrame) to target worksheet"""
        items = data if isinstance(data, list) else [data]
        current_row = 1
        
        for item in items:
            if isinstance(item, StyledSheet):
                current_row = self._copy_styled_sheet(item, target_ws, current_row)
            elif isinstance(item, pd.DataFrame):
                # Write dataframe values (no styles)
                # Optimization: Use append() for much faster writing
                
                # Write header if it's the first item
                if current_row == 1:
                    headers = [
                        sanitize_value(str(col)) if col != "" else None
                        for col in item.columns
                    ]
                    target_ws.append(headers)
                    current_row += 1
                
                # Performance optimization: Batch write for better speed
                total_rows = len(item)
                
                # For large datasets, report progress less frequently
                progress_interval = 1000 if total_rows > 5000 else 2000
                
                try:
                    # Using itertuples is faster than iterrows
                    for row_idx, row in enumerate(item.itertuples(index=False), 1):
                        if row_idx % progress_interval == 0:
                            percentage = int(row_idx/total_rows*100)
                            self.report_progress(f"⚡ 写入数据: {row_idx}/{total_rows} ({percentage}%)")
                            
                        try:
                            # Sanitize row values
                            sanitized_row = [normalize_excel_value(val) for val in row]
                            target_ws.append(sanitized_row)
                            current_row += 1
                        except Exception as e:
                            raise ValueError(f"写入数据失败，位置: 第 {row_idx} 行. 错误: {e}")
                except Exception as e:
                    if "写入数据失败" in str(e):
                        raise e
                    raise ValueError(f"处理数据失败: {e}")

    def _copy_styled_sheet(self, styled: StyledSheet, target_ws, start_row):
        """Copy data and styles from StyledSheet to target worksheet"""
        try:
            # Check for .xls file (not supported by openpyxl for style reading)
            if str(styled.file_path).lower().endswith('.xls'):
                self.report_progress(f"处理 .xls 文件: {Path(styled.file_path).name}")
                
                # Strategy: Read .xls using pandas with all data, preserve as much as possible
                # Note: .xls format is legacy binary format, full style preservation is not possible
                # We'll copy the data with basic formatting
                
                try:
                    # Read the Excel file to get all data including formulas
                    # Use openpyxl's data_only=False to try to preserve formulas (won't work for .xls though)
                    df = styled.df_filtered
                    
                    if df is None or df.empty:
                        self.report_progress(f"⚠️ .xls 文件数据为空")
                        return start_row
                    
                    # Write header if needed
                    if start_row == 1:
                        headers = [sanitize_value(str(col)) for col in df.columns]
                        target_ws.append(headers)
                        
                        # Apply basic header formatting
                        header_row = target_ws[start_row]
                        for cell in header_row:
                            cell.font = openpyxl.styles.Font(bold=True)
                            cell.fill = openpyxl.styles.PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                            cell.alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")
                        
                        start_row += 1
                    
                    # Write data rows - optimized for performance
                    total_rows = len(df)
                    progress_interval = 1000 if total_rows > 5000 else 2000
                    
                    for row_idx, row in enumerate(df.itertuples(index=False), 1):
                        if row_idx % progress_interval == 0:
                            percentage = int(row_idx/total_rows*100)
                            self.report_progress(f"⚡ 复制 .xls: {row_idx}/{total_rows} ({percentage}%)")
                        
                        sanitized_row = [normalize_excel_value(val) for val in row]
                        target_ws.append(sanitized_row)
                        start_row += 1
                    
                    # Auto-adjust column widths for better readability
                    for column in target_ws.columns:
                        max_length = 0
                        column_letter = column[0].column_letter
                        for cell in column:
                            try:
                                if cell.value:
                                    max_length = max(max_length, len(str(cell.value)))
                            except (TypeError, ValueError):
                                pass
                        adjusted_width = min(max_length + 2, 50)  # Cap at 50
                        target_ws.column_dimensions[column_letter].width = adjusted_width
                    
                    self.report_progress(f"✓ 已复制 .xls 文件数据 (共 {total_rows} 行)")
                    return start_row
                    
                except Exception as e:
                    error_msg = f"⚠️ 处理 .xls 文件失败: {e}"
                    self.report_progress(error_msg)
                    raise ValueError(error_msg)

            # Use cached workbook if available
            if styled.file_path in self._source_wb_cache:
                src_wb = self._source_wb_cache[styled.file_path]
            else:
                print(f"Loading source workbook for styles: {styled.file_path}")
                src_wb = openpyxl.load_workbook(styled.file_path, data_only=False)
                self._source_wb_cache[styled.file_path] = src_wb

            if styled.sheet_name and styled.sheet_name in src_wb.sheetnames:
                src_ws = src_wb[styled.sheet_name]
            else:
                src_ws = src_wb.active
                
            df = styled.df_filtered
            if src_ws is None:
                raise ValueError(f"无法读取来源工作表: {styled.sheet_name or 'active'}")
            if df is None:
                raise ValueError("来源数据为空，无法复制")

            header_row_idx = styled.header_row + 1 # 1-based
            
            # Check if data is sequential (unfiltered)
            # If index is RangeIndex(0, N, 1), it means no rows were dropped/reordered
            is_sequential = isinstance(df.index, pd.RangeIndex) and df.index.step == 1 and df.index.start == 0
            
            # Helper to copy a row from source to target
            def copy_src_row(s_row, t_row, min_col=1, max_col=None):
                if max_col is None:
                    max_col = src_ws.max_column
                # Copy row dimensions
                if s_row in src_ws.row_dimensions:
                    target_ws.row_dimensions[t_row] = copy_obj(src_ws.row_dimensions[s_row])
                    
                for col in range(min_col, max_col + 1):
                    src_cell = src_ws.cell(row=s_row, column=col)
                    tgt_cell = target_ws.cell(row=t_row, column=col)
                    
                    tgt_cell.value = normalize_excel_value(src_cell.value)
                    if src_cell.has_style:
                        tgt_cell.font = copy_obj(src_cell.font)
                        tgt_cell.border = copy_obj(src_cell.border)
                        tgt_cell.fill = copy_obj(src_cell.fill)
                        tgt_cell.number_format = copy_obj(src_cell.number_format)
                        tgt_cell.protection = copy_obj(src_cell.protection)
                        tgt_cell.alignment = copy_obj(src_cell.alignment)

            # =================================================================
            # STRATEGY A: Full Copy (Direct OpenPyXL Copy)
            # =================================================================
            if styled.is_full_copy and not str(styled.file_path).lower().endswith('.csv'):
                use_detected_range = (styled.copy_mode == "used_range")

                if use_detected_range:
                    ur_min_row, ur_max_row, ur_min_col, ur_max_col = detect_used_range(src_ws)
                    self.report_progress(
                        f"🔍 自动检测有值区域: 行 {ur_min_row}-{ur_max_row}, "
                        f"列 {ur_min_col}-{ur_max_col} "
                        f"(原始范围: {src_ws.max_row} 行 x {src_ws.max_column} 列)"
                    )
                    row_start = ur_min_row
                    row_end = ur_max_row
                    col_start = ur_min_col
                    col_end = ur_max_col
                else:
                    self.report_progress(f"⚡ 快速模式: 直接复制工作表 '{styled.sheet_name}'")
                    row_start = 1
                    row_end = src_ws.max_row
                    col_start = 1
                    col_end = src_ws.max_column

                # Calculate offset: source row_start -> target start_row
                offset = start_row - row_start
                total_rows = row_end - row_start + 1

                progress_interval = 1000 if total_rows > 5000 else 5000
                
                for r in range(row_start, row_end + 1):
                    copied = r - row_start + 1
                    if copied % progress_interval == 0:
                        self.report_progress(f"复制行 {copied}/{total_rows} ({int(copied/total_rows*100)}%)")
                    
                    copy_src_row(r, start_row, col_start, col_end)
                    start_row += 1
                
                # Copy column dimensions (only within range)
                for col_idx in range(col_start, col_end + 1):
                    col_letter = get_column_letter(col_idx)
                    if col_letter in src_ws.column_dimensions:
                        target_ws.column_dimensions[col_letter] = copy_obj(
                            src_ws.column_dimensions[col_letter]
                        )
                
                # Copy merged cells (only those within the copied range)
                for range_ in src_ws.merged_cells.ranges:
                    if use_detected_range:
                        if (range_.min_row < row_start or range_.max_row > row_end or
                                range_.min_col < col_start or range_.max_col > col_end):
                            continue
                    min_row = range_.min_row + offset
                    max_row = range_.max_row + offset
                    try:
                        target_ws.merge_cells(start_row=min_row, start_column=range_.min_col,
                                            end_row=max_row, end_column=range_.max_col)
                    except Exception as e:
                        print(f"Warning: Failed to merge cells {range_}: {e}")
                        
                return start_row

            # =================================================================
            # STRATEGY B: DataFrame Iteration (Filtered/CSV/Partial)
            # =================================================================
            
            # 0. Pre-Header Rows (Only for full copy - but we handled full copy above for Excel)
            # If it's CSV full copy, we don't have pre-header rows in DF usually (unless skipped)
            # So this block is mostly legacy or for edge cases where is_full_copy is True but we fell through
            if styled.is_full_copy and header_row_idx > 1 and not str(styled.file_path).lower().endswith('.csv'):
                 for r in range(1, header_row_idx):
                    copy_src_row(r, start_row)
                    start_row += 1

            # 1. Copy Header
            if start_row == 1 or (styled.is_full_copy and header_row_idx > 1):
                # If we copied pre-header rows, start_row is already advanced.
                # But usually header is copied if it's the first thing we write to this sheet (start_row==1)
                # OR if we are doing a full copy, we definitely want the header.
                
                # Copy header row from source
                for col in range(1, src_ws.max_column + 1):
                    src_cell = src_ws.cell(row=header_row_idx, column=col)
                    tgt_cell = target_ws.cell(row=start_row, column=col)
                    
                    tgt_cell.value = normalize_excel_value(src_cell.value)
                    if src_cell.has_style:
                        tgt_cell.font = copy_obj(src_cell.font)
                        tgt_cell.border = copy_obj(src_cell.border)
                        tgt_cell.fill = copy_obj(src_cell.fill)
                        tgt_cell.number_format = copy_obj(src_cell.number_format)
                        tgt_cell.protection = copy_obj(src_cell.protection)
                        tgt_cell.alignment = copy_obj(src_cell.alignment)
                
                # Copy column dimensions
                for col_letter, dim in src_ws.column_dimensions.items():
                    target_ws.column_dimensions[col_letter] = copy_obj(dim)
                
                # Copy merged cells in header area
                for range_ in src_ws.merged_cells.ranges:
                    if range_.max_row <= header_row_idx:
                        # Shift to target header row (start_row)
                        # Source header is at header_row_idx
                        # Target header is at start_row
                        offset = start_row - header_row_idx
                        
                        # We need to shift the range
                        min_row = range_.min_row + offset
                        max_row = range_.max_row + offset
                        target_ws.merge_cells(start_row=min_row, start_column=range_.min_col,
                                            end_row=max_row, end_column=range_.max_col)
                                            
                start_row += 1
            
            # 2. Copy Data Rows
            total_rows = len(df)
            last_src_row = header_row_idx # Track last processed source row
            data_target_start = start_row
            data_source_start = header_row_idx + 1
            
            col_names = list(df.columns)
            for row_num, row_tuple in enumerate(df.itertuples(index=False)):
                # Progress logging for large files
                if row_num % 500 == 0:
                    self.report_progress(f"处理行 {row_num}/{total_rows}")
                    
                try:
                    src_row_idx = header_row_idx + 1 + row_num
                    last_src_row = src_row_idx
                    
                    for col_pos, value in enumerate(row_tuple):
                        tgt_col_idx = col_pos + 1
                        
                        # Try to find source cell for style
                        src_cell = None
                        if tgt_col_idx <= src_ws.max_column:
                            src_cell = src_ws.cell(row=src_row_idx, column=tgt_col_idx)
                        
                        tgt_cell = target_ws.cell(row=start_row, column=tgt_col_idx)
                        tgt_cell.value = normalize_excel_value(value)
                        
                        if src_cell and src_cell.has_style:
                            tgt_cell.font = copy_obj(src_cell.font)
                            tgt_cell.border = copy_obj(src_cell.border)
                            tgt_cell.fill = copy_obj(src_cell.fill)
                            tgt_cell.number_format = copy_obj(src_cell.number_format)
                            tgt_cell.protection = copy_obj(src_cell.protection)
                            tgt_cell.alignment = copy_obj(src_cell.alignment)
                    
                    # Copy row dimensions
                    if src_row_idx in src_ws.row_dimensions:
                        target_ws.row_dimensions[start_row] = copy_obj(src_ws.row_dimensions[src_row_idx])
                    
                    start_row += 1
                except Exception as e:
                    raise ValueError(f"复制带格式数据失败，位置: 第 {row_num} 行 (源文件行号: {src_row_idx}). 错误: {e}")
            
            # 3. Post-Data Rows (Only for full copy)
            # This handles empty rows at the end that Pandas skipped but have formatting
            if styled.is_full_copy and last_src_row < src_ws.max_row and not str(styled.file_path).lower().endswith('.csv'):
                print(f"Copying trailing empty rows from {last_src_row + 1} to {src_ws.max_row}")
                for r in range(last_src_row + 1, src_ws.max_row + 1):
                    copy_src_row(r, start_row)
                    start_row += 1

            # 4. Copy Merged Cells in Data Area (Only if sequential/unfiltered)
            if is_sequential:
                # Source data block copied from DataFrame rows
                source_data_end = data_source_start + total_rows - 1
                # Mapping: source row -> target row by fixed offset
                offset = data_target_start - data_source_start

                for range_ in src_ws.merged_cells.ranges:
                    # Skip header/pre-header merges; handled above
                    if range_.max_row <= header_row_idx:
                        continue

                    # For non-full copy, only keep merges fully inside copied data block
                    if not styled.is_full_copy:
                        if range_.min_row < data_source_start or range_.max_row > source_data_end:
                            continue

                    min_row = range_.min_row + offset
                    max_row = range_.max_row + offset

                    try:
                        target_ws.merge_cells(
                            start_row=min_row,
                            start_column=range_.min_col,
                            end_row=max_row,
                            end_column=range_.max_col
                        )
                    except Exception as e:
                        print(f"Warning: Failed to merge cells {range_}: {e}")

            return start_row
            
        except Exception as e:
            # If style copy fails, raise error so user knows
            raise ValueError(f"复制样式失败: {e}")
