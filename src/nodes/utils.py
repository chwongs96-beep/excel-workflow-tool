"""
Shared utility functions for Excel/CSV file handling.

Centralises helpers that were previously duplicated across
excel_nodes.py and merge_nodes.py.
"""

import re

import numpy as np
import pandas as pd
from pathlib import Path
from typing import Any, List, Optional, Tuple

from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE


# ---------------------------------------------------------------------------
# Value normalisation
# ---------------------------------------------------------------------------

def sanitize_value(value: Any) -> Any:
    """Remove illegal XML characters from string values (required by openpyxl)."""
    if isinstance(value, str):
        return ILLEGAL_CHARACTERS_RE.sub('', value)
    return value


def normalize_excel_value(value: Any) -> Any:
    """Normalize a value for Excel writing.

    * Numeric-looking text → int / float
    * Leading-zero identifiers (e.g. "00123") → kept as text
    * Formulas (starts with "=") → kept as text
    * Everything else → unchanged
    """
    value = sanitize_value(value)

    if not isinstance(value, str):
        return value

    text = value.strip()
    if text == "":
        return ""

    # Keep formulas
    if text.startswith("="):
        return text

    # Keep identifiers with leading zeros (e.g. "00123")
    if re.fullmatch(r"[+-]?0\d+", text):
        return text

    # Remove thousand separators for numeric parsing
    compact = text.replace(",", "")

    # Integer
    if re.fullmatch(r"[+-]?\d+", compact):
        try:
            return int(compact)
        except (ValueError, OverflowError):
            return text

    # Decimal / scientific notation
    if re.fullmatch(r"[+-]?(\d+\.\d*|\d*\.\d+|\d+)([eE][+-]?\d+)?", compact):
        try:
            return float(compact)
        except (ValueError, OverflowError):
            return text

    return text


def clean_unnamed_columns(df: pd.DataFrame) -> pd.DataFrame:
    """Replace auto-generated placeholder column names with empty strings.

    Handles both pandas ``Unnamed: N`` and our own ``_col_N`` placeholders.
    """
    df = df.copy()
    df.columns = [
        "" if isinstance(c, str) and (
            re.match(r"^Unnamed:\s*\d+", c) or re.match(r"^_col_\d+$", c)
        ) else c
        for c in df.columns
    ]
    return df


def normalize_dataframe_for_excel(df: pd.DataFrame) -> pd.DataFrame:
    """Apply :func:`normalize_excel_value` to every cell in *df*."""
    normalized = df.copy()
    for col in normalized.columns:
        normalized[col] = normalized[col].map(normalize_excel_value)
    return normalized


# ---------------------------------------------------------------------------
# Sheet-name helpers
# ---------------------------------------------------------------------------

def sanitize_sheet_name(name: str) -> str:
    """Sanitize a sheet name so Excel accepts it (max 31 chars, no special chars)."""
    name = str(name)
    for ch in ('\\', '/', '?', '*', ':', '[', ']'):
        name = name.replace(ch, '_')
    return name[:31]


def make_unique_sheet_name(name: str, used_names=None) -> str:
    """Return an Excel-safe sheet name that is not in *used_names*."""
    if used_names is None:
        used_names = set()

    base_name = sanitize_sheet_name(name) or "Sheet"
    candidate = base_name
    counter = 1
    while candidate in used_names:
        suffix = f"_{counter}"
        candidate = f"{base_name[:31 - len(suffix)]}{suffix}"
        counter += 1
    return candidate


# ---------------------------------------------------------------------------
# File reading helpers
# ---------------------------------------------------------------------------

def read_excel_with_engine(file_path, sheet_name: Any = 0, header: int = 0, dtype=None):
    """Read an Excel file using the correct engine based on extension.

    Parameters
    ----------
    dtype : optional
        Passed to :func:`pd.read_excel`.  Use ``object`` to preserve all
        values as strings (prevents leading-zero loss, etc.).
    """
    ext = str(file_path).lower()
    engine = 'xlrd' if ext.endswith('.xls') else 'openpyxl'
    kwargs: dict[str, Any] = dict(
        sheet_name=sheet_name, header=header, engine=engine,
    )
    if dtype is not None:
        kwargs['dtype'] = dtype
    return pd.read_excel(file_path, **kwargs)


def get_xls_merged_cells(file_path, sheet_name=None) -> List[Tuple[int, int, int, int]]:
    """Return merged-cell ranges from an ``.xls`` file via *xlrd*.

    Each range is ``(row_start, row_end, col_start, col_end)`` using
    **0-indexed, exclusive-end** convention (same as xlrd).

    Returns an empty list when merged-cell info cannot be read.
    """
    try:
        import xlrd
    except ImportError:
        return []
    try:
        wb = xlrd.open_workbook(str(file_path), formatting_info=True)
    except Exception:
        try:
            wb = xlrd.open_workbook(str(file_path))
        except Exception:
            return []

    try:
        if sheet_name:
            ws = wb.sheet_by_name(sheet_name)
        else:
            ws = wb.sheet_by_index(0)
        return list(ws.merged_cells)
    except Exception:
        return []


def get_excel_file(file_path) -> pd.ExcelFile:
    """Create a :class:`pd.ExcelFile` with the correct engine for the extension."""
    ext = str(file_path).lower()
    engine = 'xlrd' if ext.endswith('.xls') else 'openpyxl'
    return pd.ExcelFile(file_path, engine=engine)


def convert_text_to_numeric(df: pd.DataFrame) -> Tuple[pd.DataFrame, List[str]]:
    """Heuristically convert text columns that are mostly numeric.

    Returns the (possibly modified) DataFrame and a list of converted column names.
    """
    converted_cols: List[str] = []
    for col in df.columns:
        if df[col].dtype != 'object':
            continue
        try:
            numeric_series = pd.to_numeric(df[col], errors='coerce')
            non_null_count = df[col].notna().sum()
            if non_null_count > 0:
                converted_count = numeric_series.notna().sum()
                if converted_count / non_null_count > 0.5:
                    df[col] = numeric_series
                    converted_cols.append(col)
        except Exception:
            pass
    return df, converted_cols


def read_csv_with_options(
    file_path,
    encoding_opt: str = 'auto',
    delimiter_opt: str = 'auto',
    header_row: int = 0,
    *,
    dtype=None,
    auto_convert_numeric: bool = True,
):
    """Read a CSV with flexible encoding/delimiter options.

    Parameters
    ----------
    dtype : optional
        Passed to :func:`pd.read_csv`.  Use ``object`` to preserve leading
        zeros; leave *None* for standard dtype inference.
    auto_convert_numeric : bool
        When *True* (and *dtype* is not ``object``), run
        :func:`convert_text_to_numeric` after reading.
    """
    # Determine delimiter / engine
    sep = None
    engine = 'python'

    if delimiter_opt == 'comma':
        sep, engine = ',', 'c'
    elif delimiter_opt == 'tab':
        sep, engine = '\t', 'c'
    elif delimiter_opt == 'semicolon':
        sep, engine = ';', 'c'
    elif delimiter_opt == 'pipe':
        sep, engine = '|', 'python'
    elif delimiter_opt == 'space':
        sep, engine = r'\s+', 'python'

    # Determine encoding list
    if encoding_opt and encoding_opt != 'auto':
        encodings = [encoding_opt]
    else:
        encodings = ['utf-8', 'gbk', 'utf-8-sig', 'gb18030', 'latin1']

    last_error: Exception | None = None
    for enc in encodings:
        try:
            # --- Phase 1: determine max field count across all rows ----------
            # Without this, pandas uses the first row's field count and drops
            # any row that has more fields (treating it as a "bad line").
            actual_sep = sep
            try:
                with open(file_path, encoding=enc) as fh:
                    if actual_sep is None:
                        import csv as _csv
                        sample = fh.read(8192)
                        dialect = _csv.Sniffer().sniff(sample)
                        actual_sep = dialect.delimiter
                        fh.seek(0)
                    max_fields = 0
                    for line in fh:
                        if line.strip():
                            n = line.count(actual_sep) + 1
                            if n > max_fields:
                                max_fields = n
            except Exception:
                max_fields = 0

            # --- Phase 2: read CSV with explicit column count ----------------
            read_kwargs: dict[str, Any] = {
                'sep': sep,
                'encoding': enc,
                'header': None,
                'engine': 'python',
                'keep_default_na': True,
                'skipinitialspace': True,
                'on_bad_lines': 'warn',
            }
            if dtype is not None:
                read_kwargs['dtype'] = dtype
            if max_fields > 0:
                read_kwargs['names'] = list(range(max_fields))

            df = pd.read_csv(file_path, **read_kwargs)

            # Promote the designated header row to column names
            hdr_idx = header_row if header_row is not None else 0
            if hdr_idx < len(df):
                new_cols = []
                for i, val in enumerate(df.iloc[hdr_idx]):
                    if pd.notna(val) and str(val).strip():
                        new_cols.append(str(val).strip())
                    else:
                        new_cols.append(f"_col_{i}")
                df.columns = new_cols
                df = df.iloc[hdr_idx + 1:].reset_index(drop=True)

            if auto_convert_numeric and dtype is None:
                df, converted_cols = convert_text_to_numeric(df)
                if converted_cols:
                    print(f"已转换文本数字为数值类型: {', '.join(converted_cols)}")

            if enc != 'utf-8':
                print(f"成功使用 {enc} 编码读取CSV: {Path(file_path).name}")

            return df
        except Exception as e:
            last_error = e
            continue

    raise ValueError(f"无法读取CSV文件 (尝试了编码: {encodings}): {last_error}")


def read_tabular_file(
    file_path: str,
    sheet_name: Any = 0,
    header_row: int = 0,
    csv_encoding: str = "auto",
    csv_delimiter: str = "auto",
):
    """Read CSV / XLS / XLSX using the correct engine and options.

    Always reads with ``dtype=object`` so leading-zero strings are preserved.
    """
    lower_path = str(file_path).lower()

    if lower_path.endswith('.csv'):
        return read_csv_with_options(
            file_path,
            encoding_opt=csv_encoding,
            delimiter_opt=csv_delimiter,
            header_row=header_row,
            dtype=object,
            auto_convert_numeric=False,
        )

    engine = 'xlrd' if lower_path.endswith('.xls') else 'openpyxl'
    return pd.read_excel(
        file_path, sheet_name=sheet_name, header=header_row,
        engine=engine, dtype=object,
    )


# ---------------------------------------------------------------------------
# Used-range detection
# ---------------------------------------------------------------------------

def detect_used_range(ws) -> Tuple[int, int, int, int]:
    """Detect the actual used range of an openpyxl worksheet.

    Scans cell values **and** merged-cell ranges to determine the tightest
    bounding rectangle that contains all non-empty content.

    Returns ``(min_row, max_row, min_col, max_col)`` using 1-based indexing
    (same convention as openpyxl).  For a completely empty sheet the return
    value is ``(1, 1, 1, 1)``.
    """
    real_max_row = 0
    real_max_col = 0
    real_min_row = ws.max_row or 1
    real_min_col = ws.max_column or 1

    # Scan from the reported max_row downward to find the true last row with
    # a value, and similarly for columns.  We iterate all cells that openpyxl
    # knows about rather than probing row-by-row from the end, because
    # ws.iter_rows is efficient and handles sparse sheets well.
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row,
                            min_col=1, max_col=ws.max_column):
        for cell in row:
            if cell.value is not None:
                r, c = cell.row, cell.column
                if r < real_min_row:
                    real_min_row = r
                if r > real_max_row:
                    real_max_row = r
                if c < real_min_col:
                    real_min_col = c
                if c > real_max_col:
                    real_max_col = c

    # Expand boundaries to include merged-cell ranges (a merged region may
    # extend beyond the last valued cell).
    for rng in ws.merged_cells.ranges:
        if rng.min_row < real_min_row:
            real_min_row = rng.min_row
        if rng.max_row > real_max_row:
            real_max_row = rng.max_row
        if rng.min_col < real_min_col:
            real_min_col = rng.min_col
        if rng.max_col > real_max_col:
            real_max_col = rng.max_col

    # Fallback for completely empty sheets
    if real_max_row == 0:
        return (1, 1, 1, 1)

    return (real_min_row, real_max_row, real_min_col, real_max_col)


# ---------------------------------------------------------------------------
# DataFrame grid ↔ Excel coordinates (row 1 = column headers)
# ---------------------------------------------------------------------------


def _df_cell_has_content(value: Any) -> bool:
    if value is None:
        return False
    if isinstance(value, float) and pd.isna(value):
        return False
    if isinstance(value, str) and value.strip() == "":
        return False
    return True


def parse_a1_range(range_str: str) -> Tuple[int, int, int, int]:
    """Parse an Excel A1 range into *(min_row, max_row, min_col, max_col)* (1-based inclusive).

    Examples: ``"B2:D10"``, ``"A1"`` (single cell).  Compatible with openpyxl.
    """
    from openpyxl.utils.cell import range_boundaries

    s = (range_str or "").strip()
    if not s:
        raise ValueError("请填写 Excel 区域，例如 A2:Z100")
    if ":" not in s:
        s = f"{s}:{s}"
    min_col, min_row, max_col, max_row = range_boundaries(s)
    return (min_row, max_row, min_col, max_col)


def detect_used_range_from_dataframe(df: pd.DataFrame) -> Tuple[int, int, int, int]:
    """Tight bounding box of non-empty cells if *df* were saved with row 1 as headers.

    Returns ``(min_row, max_row, min_col, max_col)`` in Excel 1-based coordinates.
    Row **1** is the header row (``df.columns``); row **2+** are data rows.
    """
    nrows = len(df)
    ncols = len(df.columns)
    if ncols == 0:
        return (1, 1, 1, 1)

    min_r: Optional[int] = None
    max_r: Optional[int] = None
    min_c: Optional[int] = None
    max_c: Optional[int] = None

    def touch(ex_row: int, ex_col: int) -> None:
        nonlocal min_r, max_r, min_c, max_c
        if min_r is None:
            min_r = max_r = ex_row
            min_c = max_c = ex_col
            return
        min_r = min(min_r, ex_row)
        max_r = max(max_r, ex_row)
        min_c = min(min_c, ex_col)
        max_c = max(max_c, ex_col)

    for j, col in enumerate(df.columns):
        if _df_cell_has_content(col):
            touch(1, j + 1)

    if nrows > 0 and ncols > 0:
        arr = df.to_numpy(dtype=object, copy=False)
        vm = np.vectorize(_df_cell_has_content, otypes=[bool])(arr)
        if vm.any():
            r_idx, c_idx = np.nonzero(vm)
            bmin_r = int(r_idx.min()) + 2
            bmax_r = int(r_idx.max()) + 2
            bmin_c = int(c_idx.min()) + 1
            bmax_c = int(c_idx.max()) + 1
            touch(bmin_r, bmin_c)
            touch(bmin_r, bmax_c)
            touch(bmax_r, bmin_c)
            touch(bmax_r, bmax_c)

    if min_r is None:
        return (1, 1, 1, 1)

    return (min_r, max_r, min_c, max_c)


def clear_dataframe_excel_range(
    df: pd.DataFrame,
    min_row: int,
    max_row: int,
    min_col: int,
    max_col: int,
    *,
    clear_header_cells: bool = True,
) -> pd.DataFrame:
    """Clear cell *values* inside an Excel-style rectangle (1-based inclusive).

    Row **1** is treated as column headers: when *clear_header_cells* is true,
    affected header positions are set to empty strings.  Data rows start at **2**.

    Uses block assignment on ``.iloc`` instead of per-cell ``iat`` for performance.
    """
    out = df.copy()
    nrows = len(out)
    ncols = len(out.columns)
    if ncols == 0:
        return out

    min_row = max(1, int(min_row))
    min_col = max(1, int(min_col))
    max_col = min(int(max_col), ncols)
    last_data_excel_row = 1 + nrows
    max_row = min(int(max_row), last_data_excel_row)

    if max_row < min_row or max_col < min_col:
        return out

    display_names = list(out.columns)
    if clear_header_cells and min_row <= 1 <= max_row:
        for c_ex in range(min_col, max_col + 1):
            if 1 <= c_ex <= ncols:
                display_names[c_ex - 1] = ""

    # Unique internal names so duplicate "" headers do not break assignment, and
    # we can promote int/bool columns via ``out[nm] = out[nm].astype(object)``.
    internal = [f"__sheet_clr_{i}" for i in range(ncols)]
    out.columns = internal

    r_ex_start = max(min_row, 2)
    r_ex_end = min(max_row, last_data_excel_row)
    if r_ex_start <= r_ex_end:
        i0 = r_ex_start - 2
        i1 = r_ex_end - 2
        c0 = min_col - 1
        c1 = max_col
        i0 = max(0, min(i0, nrows - 1))
        i1 = max(0, min(i1, nrows - 1))
        c0 = max(0, min(c0, ncols - 1))
        c1 = max(0, min(c1, ncols))
        if i0 <= i1 and c0 < c1:
            for ci in range(c0, c1):
                nm = internal[ci]
                if not pd.api.types.is_object_dtype(out[nm]):
                    out[nm] = out[nm].astype(object)
            out.iloc[i0 : i1 + 1, c0:c1] = None

    out.columns = pd.Index(display_names)
    return out


def _cell_looks_like_formula(value: Any) -> bool:
    return isinstance(value, str) and len(value) > 0 and value.lstrip().startswith("=")


def detect_used_range_from_xlsx_sheet(
    file_path, sheet_name: str,
) -> Optional[Tuple[int, int, int, int]]:
    """Like :func:`detect_used_range` but reads a live ``.xlsx`` sheet.

    Uses openpyxl so merged regions extend the bounds (closer to Excel's view).
    Returns ``None`` if the file is missing, not ``.xlsx``, or the sheet cannot
    be opened.
    """
    try:
        p = Path(file_path)
        if not p.is_file() or p.suffix.lower() != ".xlsx":
            return None
        import openpyxl

        wb = openpyxl.load_workbook(str(p), data_only=True)
        try:
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
            else:
                alt = sanitize_sheet_name(sheet_name)
                if alt not in wb.sheetnames:
                    return None
                ws = wb[alt]
            return detect_used_range(ws)
        finally:
            wb.close()
    except Exception:
        return None


def restore_formula_cells(original: pd.DataFrame, cleared: pd.DataFrame) -> pd.DataFrame:
    """Put back cells that look like Excel formulas (strings starting with ``=``).

    Matches cells by **position** only so cleared sheets with blanked headers
    (e.g. duplicate empty column names) still restore correctly.
    """
    if original.shape != cleared.shape:
        return cleared
    arr_o = original.to_numpy(dtype=object, copy=False)
    arr_c = cleared.to_numpy(dtype=object, copy=True)
    is_f = np.vectorize(_cell_looks_like_formula, otypes=[bool])(arr_o)
    if not is_f.any():
        return cleared
    arr_c = arr_c.copy()
    arr_c[is_f] = arr_o[is_f]
    return pd.DataFrame(arr_c, index=cleared.index, columns=cleared.columns)


def trim_dataframe_blank_edges(
    df: pd.DataFrame,
    *,
    trim_rows: bool = True,
    trim_cols: bool = False,
) -> pd.DataFrame:
    """Drop rows and/or columns that contain no *content* (see :func:`_df_cell_has_content`)."""
    if df.empty:
        return df
    out = df
    if trim_rows:
        arr = out.to_numpy(dtype=object, copy=False)
        row_ok = np.vectorize(_df_cell_has_content)(arr).any(axis=1)
        if not row_ok.all():
            out = out.loc[row_ok].reset_index(drop=True)
    if trim_cols and len(out.columns) > 0:
        arr = out.to_numpy(dtype=object, copy=False)
        col_ok = np.vectorize(_df_cell_has_content)(arr).any(axis=0)
        if not col_ok.all():
            out = out.loc[:, col_ok]
    return out
