"""
Excel-related nodes for reading, writing, and processing Excel files
"""

import pandas as pd
from pathlib import Path
from typing import Any, Dict, List
from .base_node import BaseNode
from .node_registry import register_node


@register_node
class ReadExcelNode(BaseNode):
    """Node to read Excel files"""
    
    node_type = "read_excel"
    node_name = "读取Excel"
    node_category = "输入/输出"
    node_description = "从Excel文件读取数据"
    node_color = "#22c55e"  # Green
    
    def _setup_ports(self):
        self.add_output("data")
    
    def get_config_ui_schema(self) -> List[Dict[str, Any]]:
        return [
            {
                "key": "file_path",
                "label": "文件路径",
                "type": "file",
                "file_filter": "Excel文件 (*.xlsx *.xls);;所有文件 (*.*)",
                "required": True
            },
            {
                "key": "sheet_name",
                "label": "工作表名称",
                "type": "text",
                "default": "",
                "placeholder": "留空则读取第一个工作表"
            },
            {
                "key": "header_row",
                "label": "标题行",
                "type": "number",
                "default": 0,
                "min": 0
            }
        ]
    
    def validate(self) -> tuple[bool, str]:
        file_path = self.get_param("file_path", "")
        if not file_path:
            return False, "File path is required"
        if not Path(file_path).exists():
            return False, f"File not found: {file_path}"
        return True, ""
    
    def execute(self, input_data: Dict[str, Any]) -> Dict[str, Any]:
        file_path = self.get_param("file_path")
        sheet_name = self.get_param("sheet_name", "") or 0
        header_row = self.get_param("header_row", 0)
        
        df = pd.read_excel(
            file_path,
            sheet_name=sheet_name,
            header=header_row
        )
        
        return {"data": df}


@register_node
class WriteExcelNode(BaseNode):
    """Node to write Excel files"""
    
    node_type = "write_excel"
    node_name = "写入Excel"
    node_category = "输入/输出"
    node_description = "将数据写入Excel文件"
    node_color = "#22c55e"  # Green
    
    def _setup_ports(self):
        self.add_input("data")
        self.add_output("data")  # Pass through
    
    def get_config_ui_schema(self) -> List[Dict[str, Any]]:
        return [
            {
                "key": "file_path",
                "label": "输出文件路径",
                "type": "file_save",
                "file_filter": "Excel文件 (*.xlsx);;所有文件 (*.*)",
                "required": True
            },
            {
                "key": "sheet_name",
                "label": "工作表名称",
                "type": "text",
                "default": "Sheet1"
            },
            {
                "key": "include_index",
                "label": "包含索引",
                "type": "checkbox",
                "default": False
            }
        ]
    
    def validate(self) -> tuple[bool, str]:
        file_path = self.get_param("file_path", "")
        if not file_path:
            return False, "Output file path is required"
        return True, ""
    
    def execute(self, input_data: Dict[str, Any]) -> Dict[str, Any]:
        df = input_data.get("data")
        if df is None:
            raise ValueError("No input data received")
        
        file_path = self.get_param("file_path")
        
        # Ensure file extension is .xlsx
        if not str(file_path).lower().endswith('.xlsx'):
            file_path = str(file_path) + '.xlsx'
            
        sheet_name = self.get_param("sheet_name", "Sheet1")
        include_index = self.get_param("include_index", False)
        
        df.to_excel(
            file_path,
            sheet_name=sheet_name,
            index=include_index
        )
        
        return {"data": df}


@register_node
class FilterRowsNode(BaseNode):
    """Node to filter rows based on conditions"""
    
    node_type = "filter_rows"
    node_name = "筛选行"
    node_category = "转换"
    node_description = "根据列条件筛选行"
    node_color = "#f59e0b"  # Amber
    
    def _setup_ports(self):
        self.add_input("data")
        self.add_output("data")
    
    def get_config_ui_schema(self) -> List[Dict[str, Any]]:
        return [
            {
                "key": "column",
                "label": "列名",
                "type": "text",
                "required": True
            },
            {
                "key": "operator",
                "label": "运算符",
                "type": "select",
                "options": [
                    {"value": "==", "label": "等于 (==)"},
                    {"value": "!=", "label": "不等于 (!=)"},
                    {"value": ">", "label": "大于 (>)"},
                    {"value": ">=", "label": "大于等于 (>=)"},
                    {"value": "<", "label": "小于 (<)"},
                    {"value": "<=", "label": "小于等于 (<=)"},
                    {"value": "contains", "label": "包含"},
                    {"value": "startswith", "label": "开头是"},
                    {"value": "endswith", "label": "结尾是"},
                    {"value": "isnull", "label": "为空"},
                    {"value": "notnull", "label": "不为空"}
                ],
                "default": "=="
            },
            {
                "key": "value",
                "label": "值",
                "type": "text",
                "default": ""
            }
        ]
    
    def validate(self) -> tuple[bool, str]:
        column = self.get_param("column", "")
        if not column:
            return False, "Column name is required"
        return True, ""
    
    def execute(self, input_data: Dict[str, Any]) -> Dict[str, Any]:
        df = input_data.get("data")
        if df is None:
            raise ValueError("No input data received")
        
        df = df.copy()
        column = self.get_param("column")
        operator = self.get_param("operator", "==")
        value = self.get_param("value", "")
        
        if column not in df.columns:
            raise ValueError(f"Column '{column}' not found in data")
        
        # Try to convert value to numeric if possible
        try:
            numeric_value = float(value)
            if numeric_value.is_integer():
                numeric_value = int(numeric_value)
        except (ValueError, TypeError):
            numeric_value = value
        
        if operator == "==":
            mask = df[column] == numeric_value
        elif operator == "!=":
            mask = df[column] != numeric_value
        elif operator == ">":
            mask = df[column] > numeric_value
        elif operator == ">=":
            mask = df[column] >= numeric_value
        elif operator == "<":
            mask = df[column] < numeric_value
        elif operator == "<=":
            mask = df[column] <= numeric_value
        elif operator == "contains":
            mask = df[column].astype(str).str.contains(value, case=False, na=False)
        elif operator == "startswith":
            mask = df[column].astype(str).str.startswith(value, na=False)
        elif operator == "endswith":
            mask = df[column].astype(str).str.endswith(value, na=False)
        elif operator == "isnull":
            mask = df[column].isna()
        elif operator == "notnull":
            mask = df[column].notna()
        else:
            mask = pd.Series([True] * len(df))
        
        return {"data": df[mask].reset_index(drop=True)}


@register_node
class SelectColumnsNode(BaseNode):
    """Node to select specific columns"""
    
    node_type = "select_columns"
    node_name = "选择列"
    node_category = "转换"
    node_description = "从数据中选择特定列"
    node_color = "#f59e0b"  # Amber
    
    def _setup_ports(self):
        self.add_input("data")
        self.add_output("data")
    
    def get_config_ui_schema(self) -> List[Dict[str, Any]]:
        return [
            {
                "key": "columns",
                "label": "列名 (逗号分隔)",
                "type": "text",
                "required": True,
                "placeholder": "列1, 列2, 列3"
            }
        ]
    
    def validate(self) -> tuple[bool, str]:
        columns = self.get_param("columns", "")
        if not columns:
            return False, "At least one column is required"
        return True, ""
    
    def execute(self, input_data: Dict[str, Any]) -> Dict[str, Any]:
        df = input_data.get("data")
        if df is None:
            raise ValueError("No input data received")
        
        columns_str = self.get_param("columns", "")
        columns = [c.strip() for c in columns_str.split(",") if c.strip()]
        
        missing = [c for c in columns if c not in df.columns]
        if missing:
            raise ValueError(f"Columns not found: {missing}")
        
        return {"data": df[columns].copy()}


@register_node
class RenameColumnsNode(BaseNode):
    """Node to rename columns"""
    
    node_type = "rename_columns"
    node_name = "重命名列"
    node_category = "转换"
    node_description = "重命名数据中的列"
    node_color = "#f59e0b"  # Amber
    
    def _setup_ports(self):
        self.add_input("data")
        self.add_output("data")
    
    def get_config_ui_schema(self) -> List[Dict[str, Any]]:
        return [
            {
                "key": "rename_map",
                "label": "重命名映射 (旧名:新名，每行一个)",
                "type": "textarea",
                "required": True,
                "placeholder": "旧列名:新列名\n列1:新列名1"
            }
        ]
    
    def execute(self, input_data: Dict[str, Any]) -> Dict[str, Any]:
        df = input_data.get("data")
        if df is None:
            raise ValueError("No input data received")
        
        df = df.copy()
        rename_str = self.get_param("rename_map", "")
        
        rename_dict = {}
        for line in rename_str.strip().split("\n"):
            if ":" in line:
                old, new = line.split(":", 1)
                rename_dict[old.strip()] = new.strip()
        
        if rename_dict:
            df = df.rename(columns=rename_dict)
        
        return {"data": df}


@register_node
class SortDataNode(BaseNode):
    """Node to sort data"""
    
    node_type = "sort_data"
    node_name = "排序数据"
    node_category = "转换"
    node_description = "按一列或多列对数据排序"
    node_color = "#f59e0b"  # Amber
    
    def _setup_ports(self):
        self.add_input("data")
        self.add_output("data")
    
    def get_config_ui_schema(self) -> List[Dict[str, Any]]:
        return [
            {
                "key": "columns",
                "label": "排序列 (逗号分隔)",
                "type": "text",
                "required": True
            },
            {
                "key": "ascending",
                "label": "升序",
                "type": "checkbox",
                "default": True
            }
        ]
    
    def execute(self, input_data: Dict[str, Any]) -> Dict[str, Any]:
        df = input_data.get("data")
        if df is None:
            raise ValueError("No input data received")
        
        columns_str = self.get_param("columns", "")
        columns = [c.strip() for c in columns_str.split(",") if c.strip()]
        ascending = self.get_param("ascending", True)
        
        df = df.sort_values(by=columns, ascending=ascending).reset_index(drop=True)
        
        return {"data": df}


@register_node
class RemoveDuplicatesNode(BaseNode):
    """Node to remove duplicate rows"""
    
    node_type = "remove_duplicates"
    node_name = "删除重复"
    node_category = "转换"
    node_description = "从数据中删除重复行"
    node_color = "#f59e0b"  # Amber
    
    def _setup_ports(self):
        self.add_input("data")
        self.add_output("data")
    
    def get_config_ui_schema(self) -> List[Dict[str, Any]]:
        return [
            {
                "key": "columns",
                "label": "基于列 (逗号分隔，留空表示所有列)",
                "type": "text",
                "default": "",
                "placeholder": "留空则检查所有列"
            },
            {
                "key": "keep",
                "label": "保留",
                "type": "select",
                "options": [
                    {"value": "first", "label": "第一次出现"},
                    {"value": "last", "label": "最后一次出现"},
                    {"value": "none", "label": "删除所有重复"}
                ],
                "default": "first"
            }
        ]
    
    def execute(self, input_data: Dict[str, Any]) -> Dict[str, Any]:
        df = input_data.get("data")
        if df is None:
            raise ValueError("No input data received")
        
        columns_str = self.get_param("columns", "")
        columns = [c.strip() for c in columns_str.split(",") if c.strip()] or None
        keep = self.get_param("keep", "first")
        
        if keep == "none":
            keep = False
        
        df = df.drop_duplicates(subset=columns, keep=keep).reset_index(drop=True)
        
        return {"data": df}


@register_node
class MergeDataNode(BaseNode):
    """Node to merge two dataframes"""
    
    node_type = "merge_data"
    node_name = "合并数据"
    node_category = "合并"
    node_description = "基于公共列合并两个数据集"
    node_color = "#8b5cf6"  # Purple
    
    def _setup_ports(self):
        self.add_input("left")
        self.add_input("right")
        self.add_output("data")
    
    def get_config_ui_schema(self) -> List[Dict[str, Any]]:
        return [
            {
                "key": "left_on",
                "label": "左表关键列",
                "type": "text",
                "required": True
            },
            {
                "key": "right_on",
                "label": "右表关键列",
                "type": "text",
                "required": True
            },
            {
                "key": "how",
                "label": "合并方式",
                "type": "select",
                "options": [
                    {"value": "inner", "label": "内连接 (仅匹配)"},
                    {"value": "left", "label": "左连接 (保留左表全部)"},
                    {"value": "right", "label": "右连接 (保留右表全部)"},
                    {"value": "outer", "label": "外连接 (保留两表全部)"}
                ],
                "default": "inner"
            }
        ]
    
    def execute(self, input_data: Dict[str, Any]) -> Dict[str, Any]:
        left_df = input_data.get("left")
        right_df = input_data.get("right")
        
        if left_df is None:
            raise ValueError("No left input data received")
        if right_df is None:
            raise ValueError("No right input data received")
        
        left_on = self.get_param("left_on")
        right_on = self.get_param("right_on")
        how = self.get_param("how", "inner")
        
        merged = pd.merge(left_df, right_df, left_on=left_on, right_on=right_on, how=how)
        
        return {"data": merged}


@register_node
class ConcatDataNode(BaseNode):
    """Node to concatenate dataframes"""
    
    node_type = "concat_data"
    node_name = "拼接数据"
    node_category = "合并"
    node_description = "将多个数据集垂直堆叠"
    node_color = "#8b5cf6"  # Purple
    
    def _setup_ports(self):
        self.add_input("data1")
        self.add_input("data2")
        self.add_output("data")
    
    def get_config_ui_schema(self) -> List[Dict[str, Any]]:
        return [
            {
                "key": "ignore_index",
                "label": "重置索引",
                "type": "checkbox",
                "default": True
            }
        ]
    
    def execute(self, input_data: Dict[str, Any]) -> Dict[str, Any]:
        dfs = []
        for key in ["data1", "data2"]:
            if key in input_data and input_data[key] is not None:
                dfs.append(input_data[key])
        
        if not dfs:
            raise ValueError("No input data received")
        
        ignore_index = self.get_param("ignore_index", True)
        result = pd.concat(dfs, ignore_index=ignore_index)
        
        return {"data": result}


@register_node
class AddColumnNode(BaseNode):
    """Node to add a new column"""
    
    node_type = "add_column"
    node_name = "添加列"
    node_category = "转换"
    node_description = "添加带有公式或值的新列"
    node_color = "#f59e0b"  # Amber
    
    def _setup_ports(self):
        self.add_input("data")
        self.add_output("data")
    
    def get_config_ui_schema(self) -> List[Dict[str, Any]]:
        return [
            {
                "key": "column_name",
                "label": "新列名称",
                "type": "text",
                "required": True
            },
            {
                "key": "value_type",
                "label": "值类型",
                "type": "select",
                "options": [
                    {"value": "constant", "label": "常量值"},
                    {"value": "formula", "label": "公式 (使用列名)"}
                ],
                "default": "constant"
            },
            {
                "key": "value",
                "label": "值 / 公式",
                "type": "text",
                "required": True,
                "placeholder": "例如: 100 或 列1 + 列2"
            }
        ]
    
    def execute(self, input_data: Dict[str, Any]) -> Dict[str, Any]:
        df = input_data.get("data")
        if df is None:
            raise ValueError("No input data received")
        
        df = df.copy()
        column_name = self.get_param("column_name")
        value_type = self.get_param("value_type", "constant")
        value = self.get_param("value", "")
        
        if value_type == "constant":
            # Try to convert to number
            try:
                num_value = float(value)
                if num_value.is_integer():
                    num_value = int(num_value)
                df[column_name] = num_value
            except ValueError:
                df[column_name] = value
        else:
            # Formula - evaluate using pandas eval
            try:
                df[column_name] = df.eval(value)
            except Exception as e:
                raise ValueError(f"Formula error: {e}")
        
        return {"data": df}


@register_node  
class GroupByNode(BaseNode):
    """Node to group and aggregate data"""
    
    node_type = "group_by"
    node_name = "分组汇总"
    node_category = "汇总"
    node_description = "对数据分组并应用汇总函数"
    node_color = "#ec4899"  # Pink
    
    def _setup_ports(self):
        self.add_input("data")
        self.add_output("data")
    
    def get_config_ui_schema(self) -> List[Dict[str, Any]]:
        return [
            {
                "key": "group_columns",
                "label": "分组列 (逗号分隔)",
                "type": "text",
                "required": True
            },
            {
                "key": "agg_column",
                "label": "汇总列",
                "type": "text",
                "required": True
            },
            {
                "key": "agg_function",
                "label": "汇总函数",
                "type": "select",
                "options": [
                    {"value": "sum", "label": "求和"},
                    {"value": "mean", "label": "平均值"},
                    {"value": "count", "label": "计数"},
                    {"value": "min", "label": "最小值"},
                    {"value": "max", "label": "最大值"},
                    {"value": "first", "label": "第一个"},
                    {"value": "last", "label": "最后一个"}
                ],
                "default": "sum"
            }
        ]
    
    def execute(self, input_data: Dict[str, Any]) -> Dict[str, Any]:
        df = input_data.get("data")
        if df is None:
            raise ValueError("No input data received")
        
        group_cols = [c.strip() for c in self.get_param("group_columns", "").split(",")]
        agg_col = self.get_param("agg_column")
        agg_func = self.get_param("agg_function", "sum")
        
        result = df.groupby(group_cols)[agg_col].agg(agg_func).reset_index()
        result.columns = list(group_cols) + [f"{agg_col}_{agg_func}"]
        
        return {"data": result}


@register_node
class FillNaNode(BaseNode):
    """Node to fill missing values"""
    
    node_type = "fill_na"
    node_name = "填充空值"
    node_category = "清洗"
    node_description = "填充列中的缺失/空值"
    node_color = "#06b6d4"  # Cyan
    
    def _setup_ports(self):
        self.add_input("data")
        self.add_output("data")
    
    def get_config_ui_schema(self) -> List[Dict[str, Any]]:
        return [
            {
                "key": "columns",
                "label": "列名 (逗号分隔，留空表示所有列)",
                "type": "text",
                "default": ""
            },
            {
                "key": "fill_method",
                "label": "填充方式",
                "type": "select",
                "options": [
                    {"value": "value", "label": "指定值"},
                    {"value": "ffill", "label": "向前填充"},
                    {"value": "bfill", "label": "向后填充"},
                    {"value": "mean", "label": "列平均值"},
                    {"value": "median", "label": "列中位数"}
                ],
                "default": "value"
            },
            {
                "key": "fill_value",
                "label": "填充值 (用于指定值方式)",
                "type": "text",
                "default": "0"
            }
        ]
    
    def execute(self, input_data: Dict[str, Any]) -> Dict[str, Any]:
        df = input_data.get("data")
        if df is None:
            raise ValueError("No input data received")
        
        df = df.copy()
        columns_str = self.get_param("columns", "")
        columns = [c.strip() for c in columns_str.split(",") if c.strip()] or df.columns.tolist()
        fill_method = self.get_param("fill_method", "value")
        fill_value = self.get_param("fill_value", "0")
        
        for col in columns:
            if col not in df.columns:
                continue
                
            if fill_method == "value":
                try:
                    fv = float(fill_value)
                    if fv.is_integer():
                        fv = int(fv)
                except ValueError:
                    fv = fill_value
                df[col] = df[col].fillna(fv)
            elif fill_method == "ffill":
                df[col] = df[col].ffill()
            elif fill_method == "bfill":
                df[col] = df[col].bfill()
            elif fill_method == "mean":
                df[col] = df[col].fillna(df[col].mean())
            elif fill_method == "median":
                df[col] = df[col].fillna(df[col].median())
        
        return {"data": df}


@register_node
class ReadAllSheetsNode(BaseNode):
    """Node to read all sheets from an Excel file"""
    
    node_type = "read_all_sheets"
    node_name = "读取所有工作表"
    node_category = "工作表操作"
    node_description = "从Excel文件读取所有工作表作为单独输出"
    node_color = "#10b981"  # Emerald
    
    def _setup_ports(self):
        self.add_output("sheets_dict")  # Dictionary of sheet_name: DataFrame
    
    def get_config_ui_schema(self) -> List[Dict[str, Any]]:
        return [
            {
                "key": "file_path",
                "label": "文件路径",
                "type": "file",
                "file_filter": "Excel文件 (*.xlsx *.xls);;所有文件 (*.*)",
                "required": True
            },
            {
                "key": "header_row",
                "label": "标题行",
                "type": "number",
                "default": 0,
                "min": 0
            }
        ]
    
    def validate(self) -> tuple[bool, str]:
        file_path = self.get_param("file_path", "")
        if not file_path:
            return False, "File path is required"
        if not Path(file_path).exists():
            return False, f"File not found: {file_path}"
        return True, ""
    
    def execute(self, input_data: Dict[str, Any]) -> Dict[str, Any]:
        file_path = self.get_param("file_path")
        header_row = self.get_param("header_row", 0)
        
        # Read all sheets
        sheets_dict = pd.read_excel(
            file_path,
            sheet_name=None,  # None means read all sheets
            header=header_row
        )
        
        return {"sheets_dict": sheets_dict}


@register_node
class GetSheetNode(BaseNode):
    """Node to get a specific sheet from sheets dictionary"""
    
    node_type = "get_sheet"
    node_name = "获取工作表"
    node_category = "工作表操作"
    node_description = "从工作表字典中提取特定工作表"
    node_color = "#10b981"  # Emerald
    
    def _setup_ports(self):
        self.add_input("sheets_dict")
        self.add_output("data")
    
    def get_config_ui_schema(self) -> List[Dict[str, Any]]:
        return [
            {
                "key": "sheet_name",
                "label": "工作表名称",
                "type": "text",
                "required": True,
                "placeholder": "输入要提取的工作表名称"
            }
        ]
    
    def validate(self) -> tuple[bool, str]:
        sheet_name = self.get_param("sheet_name", "")
        if not sheet_name:
            return False, "Sheet name is required"
        return True, ""
    
    def execute(self, input_data: Dict[str, Any]) -> Dict[str, Any]:
        sheets_dict = input_data.get("sheets_dict")
        if sheets_dict is None:
            raise ValueError("No sheets dictionary received")
        
        sheet_name = self.get_param("sheet_name")
        
        if sheet_name not in sheets_dict:
            available = list(sheets_dict.keys())
            raise ValueError(f"Sheet '{sheet_name}' not found. Available: {available}")
        
        return {"data": sheets_dict[sheet_name].copy()}


@register_node
class ListSheetsNode(BaseNode):
    """Node to list all sheet names in an Excel file"""
    
    node_type = "list_sheets"
    node_name = "列出工作表"
    node_category = "工作表操作"
    node_description = "获取Excel文件中所有工作表的名称列表"
    node_color = "#10b981"  # Emerald
    
    def _setup_ports(self):
        self.add_output("sheet_names")
        self.add_output("data")  # DataFrame with sheet info
    
    def get_config_ui_schema(self) -> List[Dict[str, Any]]:
        return [
            {
                "key": "file_path",
                "label": "文件路径",
                "type": "file",
                "file_filter": "Excel文件 (*.xlsx *.xls);;所有文件 (*.*)",
                "required": True
            }
        ]
    
    def validate(self) -> tuple[bool, str]:
        file_path = self.get_param("file_path", "")
        if not file_path:
            return False, "File path is required"
        if not Path(file_path).exists():
            return False, f"File not found: {file_path}"
        return True, ""
    
    def execute(self, input_data: Dict[str, Any]) -> Dict[str, Any]:
        from openpyxl import load_workbook
        
        file_path = self.get_param("file_path")
        
        # Load workbook to get sheet names
        wb = load_workbook(file_path, read_only=True)
        sheet_names = wb.sheetnames
        wb.close()
        
        # Create DataFrame with sheet info
        df = pd.DataFrame({
            "Sheet Index": range(len(sheet_names)),
            "Sheet Name": sheet_names
        })
        
        return {"sheet_names": sheet_names, "data": df}


@register_node
class CopySheetToFileNode(BaseNode):
    """Node to copy a sheet to another Excel file"""
    
    node_type = "copy_sheet_to_file"
    node_name = "复制工作表到文件"
    node_category = "工作表操作"
    node_description = "将数据复制到Excel文件的特定工作表 (创建或覆盖)"
    node_color = "#10b981"  # Emerald
    
    def _setup_ports(self):
        self.add_input("data")
        self.add_output("data")  # Pass through
    
    def get_config_ui_schema(self) -> List[Dict[str, Any]]:
        return [
            {
                "key": "target_file",
                "label": "目标Excel文件",
                "type": "file_save",
                "file_filter": "Excel文件 (*.xlsx);;所有文件 (*.*)",
                "required": True
            },
            {
                "key": "target_sheet",
                "label": "目标工作表名称",
                "type": "text",
                "required": True,
                "placeholder": "目标文件中的工作表名称"
            },
            {
                "key": "if_sheet_exists",
                "label": "如果工作表已存在",
                "type": "select",
                "options": [
                    {"value": "replace", "label": "替换 (覆盖)"},
                    {"value": "new", "label": "创建新的 (加后缀)"},
                    {"value": "error", "label": "报错"}
                ],
                "default": "replace"
            },
            {
                "key": "include_index",
                "label": "包含索引",
                "type": "checkbox",
                "default": False
            }
        ]
    
    def validate(self) -> tuple[bool, str]:
        target_file = self.get_param("target_file", "")
        target_sheet = self.get_param("target_sheet", "")
        if not target_file:
            return False, "Target file path is required"
        if not target_sheet:
            return False, "Target sheet name is required"
        return True, ""
    
    def execute(self, input_data: Dict[str, Any]) -> Dict[str, Any]:
        from openpyxl import load_workbook, Workbook
        
        df = input_data.get("data")
        if df is None:
            raise ValueError("No input data received")
        
        target_file = self.get_param("target_file")
        target_sheet = self.get_param("target_sheet")
        if_exists = self.get_param("if_sheet_exists", "replace")
        include_index = self.get_param("include_index", False)
        
        target_path = Path(target_file)
        
        # If file exists, we need to handle existing sheets
        if target_path.exists():
            with pd.ExcelWriter(target_file, engine='openpyxl', mode='a', 
                               if_sheet_exists=if_exists) as writer:
                df.to_excel(writer, sheet_name=target_sheet, index=include_index)
        else:
            # Create new file
            df.to_excel(target_file, sheet_name=target_sheet, index=include_index)
        
        return {"data": df}


@register_node  
class CopySheetBetweenFilesNode(BaseNode):
    """Node to copy a sheet from one Excel file to another"""
    
    node_type = "copy_sheet_between_files"
    node_name = "在文件间复制工作表"
    node_category = "工作表操作"
    node_description = "从源Excel文件复制工作表到目标Excel文件"
    node_color = "#10b981"  # Emerald
    
    def _setup_ports(self):
        self.add_output("data")  # The copied data
    
    def get_config_ui_schema(self) -> List[Dict[str, Any]]:
        return [
            {
                "key": "source_file",
                "label": "源Excel文件",
                "type": "file",
                "file_filter": "Excel文件 (*.xlsx *.xls);;所有文件 (*.*)",
                "required": True
            },
            {
                "key": "source_sheet",
                "label": "源工作表名称",
                "type": "text",
                "required": True,
                "placeholder": "要复制的工作表名称"
            },
            {
                "key": "target_file",
                "label": "目标Excel文件",
                "type": "file_save",
                "file_filter": "Excel文件 (*.xlsx);;所有文件 (*.*)",
                "required": True
            },
            {
                "key": "target_sheet",
                "label": "目标工作表名称",
                "type": "text",
                "required": True,
                "placeholder": "目标工作表名称 (留空=与源相同)"
            },
            {
                "key": "if_sheet_exists",
                "label": "如果目标工作表已存在",
                "type": "select",
                "options": [
                    {"value": "replace", "label": "替换 (覆盖)"},
                    {"value": "new", "label": "创建新的 (加后缀)"},
                    {"value": "error", "label": "报错"}
                ],
                "default": "replace"
            },
            {
                "key": "header_row",
                "label": "源文件标题行",
                "type": "number",
                "default": 0,
                "min": 0
            }
        ]
    
    def validate(self) -> tuple[bool, str]:
        source_file = self.get_param("source_file", "")
        source_sheet = self.get_param("source_sheet", "")
        target_file = self.get_param("target_file", "")
        
        if not source_file:
            return False, "Source file path is required"
        if not Path(source_file).exists():
            return False, f"Source file not found: {source_file}"
        if not source_sheet:
            return False, "Source sheet name is required"
        if not target_file:
            return False, "Target file path is required"
        return True, ""
    
    def execute(self, input_data: Dict[str, Any]) -> Dict[str, Any]:
        source_file = self.get_param("source_file")
        source_sheet = self.get_param("source_sheet")
        target_file = self.get_param("target_file")
        target_sheet = self.get_param("target_sheet") or source_sheet
        if_exists = self.get_param("if_sheet_exists", "replace")
        header_row = self.get_param("header_row", 0)
        
        # Read source sheet
        df = pd.read_excel(source_file, sheet_name=source_sheet, header=header_row)
        
        target_path = Path(target_file)
        
        # Write to target file
        if target_path.exists():
            with pd.ExcelWriter(target_file, engine='openpyxl', mode='a',
                               if_sheet_exists=if_exists) as writer:
                df.to_excel(writer, sheet_name=target_sheet, index=False)
        else:
            df.to_excel(target_file, sheet_name=target_sheet, index=False)
        
        return {"data": df}


@register_node
class WriteMultiSheetNode(BaseNode):
    """Node to write multiple dataframes to different sheets"""
    
    node_type = "write_multi_sheet"
    node_name = "写入多工作表Excel"
    node_category = "工作表操作"
    node_description = "将多个数据集写入同一Excel文件的不同工作表"
    node_color = "#10b981"  # Emerald
    
    def _setup_ports(self):
        self.add_input("data1")
        self.add_input("data2")
        self.add_input("data3")
        self.add_output("file_path")
    
    def get_config_ui_schema(self) -> List[Dict[str, Any]]:
        return [
            {
                "key": "file_path",
                "label": "输出Excel文件",
                "type": "file_save",
                "file_filter": "Excel文件 (*.xlsx);;所有文件 (*.*)",
                "required": True
            },
            {
                "key": "sheet1_name",
                "label": "工作表1名称",
                "type": "text",
                "default": "Sheet1"
            },
            {
                "key": "sheet2_name",
                "label": "工作表2名称",
                "type": "text",
                "default": "Sheet2"
            },
            {
                "key": "sheet3_name",
                "label": "工作表3名称",
                "type": "text",
                "default": "Sheet3"
            },
            {
                "key": "include_index",
                "label": "包含索引",
                "type": "checkbox",
                "default": False
            }
        ]
    
    def validate(self) -> tuple[bool, str]:
        file_path = self.get_param("file_path", "")
        if not file_path:
            return False, "Output file path is required"
        return True, ""
    
    def execute(self, input_data: Dict[str, Any]) -> Dict[str, Any]:
        file_path = self.get_param("file_path")
        include_index = self.get_param("include_index", False)
        
        sheet_names = [
            self.get_param("sheet1_name", "Sheet1"),
            self.get_param("sheet2_name", "Sheet2"),
            self.get_param("sheet3_name", "Sheet3")
        ]
        
        data_keys = ["data1", "data2", "data3"]
        
        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            for i, key in enumerate(data_keys):
                df = input_data.get(key)
                if df is not None:
                    df.to_excel(writer, sheet_name=sheet_names[i], index=include_index)
        
        return {"file_path": file_path}


@register_node
class MergeSheetsNode(BaseNode):
    """Node to merge multiple sheets from one file into one DataFrame"""
    
    node_type = "merge_sheets"
    node_name = "合并工作表"
    node_category = "工作表操作"
    node_description = "将Excel文件中的多个工作表合并成一个数据集"
    node_color = "#10b981"  # Emerald
    
    def _setup_ports(self):
        self.add_output("data")
    
    def get_config_ui_schema(self) -> List[Dict[str, Any]]:
        return [
            {
                "key": "file_path",
                "label": "Excel文件",
                "type": "file",
                "file_filter": "Excel文件 (*.xlsx *.xls);;所有文件 (*.*)",
                "required": True
            },
            {
                "key": "sheets_to_merge",
                "label": "要合并的工作表 (逗号分隔，留空表示全部)",
                "type": "text",
                "default": "",
                "placeholder": "Sheet1, Sheet2, Sheet3"
            },
            {
                "key": "add_sheet_column",
                "label": "添加源工作表列",
                "type": "checkbox",
                "default": True
            },
            {
                "key": "header_row",
                "label": "标题行",
                "type": "number",
                "default": 0,
                "min": 0
            }
        ]
    
    def validate(self) -> tuple[bool, str]:
        file_path = self.get_param("file_path", "")
        if not file_path:
            return False, "File path is required"
        if not Path(file_path).exists():
            return False, f"File not found: {file_path}"
        return True, ""
    
    def execute(self, input_data: Dict[str, Any]) -> Dict[str, Any]:
        file_path = self.get_param("file_path")
        sheets_str = self.get_param("sheets_to_merge", "")
        add_sheet_col = self.get_param("add_sheet_column", True)
        header_row = self.get_param("header_row", 0)
        
        # Get sheets to merge
        if sheets_str.strip():
            sheets = [s.strip() for s in sheets_str.split(",")]
        else:
            sheets = None  # Will read all sheets
        
        # Read all specified sheets
        all_sheets = pd.read_excel(file_path, sheet_name=sheets, header=header_row)
        
        # Combine all sheets
        dfs = []
        for sheet_name, df in all_sheets.items():
            if add_sheet_col:
                df = df.copy()
                df.insert(0, '_source_sheet', sheet_name)
            dfs.append(df)
        
        combined = pd.concat(dfs, ignore_index=True)
        
        return {"data": combined}


@register_node
class DataPreviewNode(BaseNode):
    """Node to preview data (terminal node)"""
    
    node_type = "data_preview"
    node_name = "数据预览"
    node_category = "输出"
    node_description = "预览数据 (无输出)"
    node_color = "#3b82f6"  # Blue
    
    def _setup_ports(self):
        self.add_input("data")
    
    def get_config_ui_schema(self) -> List[Dict[str, Any]]:
        return [
            {
                "key": "max_rows",
                "label": "最大显示行数",
                "type": "number",
                "default": 100,
                "min": 1,
                "max": 10000
            }
        ]
    
    def execute(self, input_data: Dict[str, Any]) -> Dict[str, Any]:
        # Preview node doesn't output anything, just stores for viewing
        df = input_data.get("data")
        if df is not None:
            max_rows = self.get_param("max_rows", 100)
            self._output_data["preview"] = df.head(max_rows)
        return {}


# ============================================================
# 数据清洗节点 (Data Cleaning)
# ============================================================

@register_node
class TrimWhitespaceNode(BaseNode):
    """Node to trim whitespace from text columns"""
    
    node_type = "trim_whitespace"
    node_name = "去除空白"
    node_category = "清洗"
    node_description = "删除文本列的前导/尾随空白"
    node_color = "#06b6d4"  # Cyan
    
    def _setup_ports(self):
        self.add_input("data")
        self.add_output("data")
    
    def get_config_ui_schema(self) -> List[Dict[str, Any]]:
        return [
            {
                "key": "columns",
                "label": "列名 (逗号分隔，留空表示所有文本列)",
                "type": "text",
                "default": ""
            },
            {
                "key": "remove_extra_spaces",
                "label": "同时删除词间多余空格",
                "type": "checkbox",
                "default": False
            }
        ]
    
    def execute(self, input_data: Dict[str, Any]) -> Dict[str, Any]:
        import re
        df = input_data.get("data")
        if df is None:
            raise ValueError("No input data received")
        
        df = df.copy()
        columns_str = self.get_param("columns", "")
        remove_extra = self.get_param("remove_extra_spaces", False)
        
        if columns_str.strip():
            columns = [c.strip() for c in columns_str.split(",")]
        else:
            columns = df.select_dtypes(include=['object']).columns.tolist()
        
        for col in columns:
            if col in df.columns:
                df[col] = df[col].astype(str).str.strip()
                if remove_extra:
                    df[col] = df[col].apply(lambda x: re.sub(r'\s+', ' ', x) if pd.notna(x) else x)
        
        return {"data": df}


@register_node
class RemoveEmptyRowsNode(BaseNode):
    """Node to remove rows with empty/null values"""
    
    node_type = "remove_empty_rows"
    node_name = "删除空行"
    node_category = "清洗"
    node_description = "删除包含空值或null的行"
    node_color = "#06b6d4"  # Cyan
    
    def _setup_ports(self):
        self.add_input("data")
        self.add_output("data")
    
    def get_config_ui_schema(self) -> List[Dict[str, Any]]:
        return [
            {
                "key": "how",
                "label": "删除条件",
                "type": "select",
                "options": [
                    {"value": "any", "label": "任一列为空"},
                    {"value": "all", "label": "所有列为空"}
                ],
                "default": "any"
            },
            {
                "key": "columns",
                "label": "检查列 (逗号分隔，留空表示所有列)",
                "type": "text",
                "default": ""
            }
        ]
    
    def execute(self, input_data: Dict[str, Any]) -> Dict[str, Any]:
        df = input_data.get("data")
        if df is None:
            raise ValueError("No input data received")
        
        how = self.get_param("how", "any")
        columns_str = self.get_param("columns", "")
        
        if columns_str.strip():
            columns = [c.strip() for c in columns_str.split(",")]
        else:
            columns = None
        
        df = df.dropna(how=how, subset=columns).reset_index(drop=True)
        
        return {"data": df}


@register_node
class FindReplaceNode(BaseNode):
    """Node to find and replace values"""
    
    node_type = "find_replace"
    node_name = "查找替换"
    node_category = "清洗"
    node_description = "在列中查找并替换值"
    node_color = "#06b6d4"  # Cyan
    
    def _setup_ports(self):
        self.add_input("data")
        self.add_output("data")
    
    def get_config_ui_schema(self) -> List[Dict[str, Any]]:
        return [
            {
                "key": "column",
                "label": "列名 (留空表示所有列)",
                "type": "text",
                "default": ""
            },
            {
                "key": "find_value",
                "label": "查找",
                "type": "text",
                "required": True
            },
            {
                "key": "replace_value",
                "label": "替换为",
                "type": "text",
                "default": ""
            },
            {
                "key": "use_regex",
                "label": "使用正则表达式",
                "type": "checkbox",
                "default": False
            },
            {
                "key": "case_sensitive",
                "label": "区分大小写",
                "type": "checkbox",
                "default": True
            }
        ]
    
    def execute(self, input_data: Dict[str, Any]) -> Dict[str, Any]:
        df = input_data.get("data")
        if df is None:
            raise ValueError("No input data received")
        
        df = df.copy()
        column = self.get_param("column", "")
        find_val = self.get_param("find_value", "")
        replace_val = self.get_param("replace_value", "")
        use_regex = self.get_param("use_regex", False)
        case_sensitive = self.get_param("case_sensitive", True)
        
        if column.strip():
            columns = [column.strip()]
        else:
            columns = df.columns.tolist()
        
        for col in columns:
            if col in df.columns:
                if use_regex:
                    df[col] = df[col].astype(str).str.replace(
                        find_val, replace_val, regex=True, case=case_sensitive
                    )
                else:
                    df[col] = df[col].replace(find_val, replace_val)
        
        return {"data": df}


@register_node
class ChangeDataTypeNode(BaseNode):
    """Node to change column data types"""
    
    node_type = "change_data_type"
    node_name = "转换数据类型"
    node_category = "清洗"
    node_description = "将列转换为不同的数据类型"
    node_color = "#06b6d4"  # Cyan
    
    def _setup_ports(self):
        self.add_input("data")
        self.add_output("data")
    
    def get_config_ui_schema(self) -> List[Dict[str, Any]]:
        return [
            {
                "key": "column",
                "label": "列名",
                "type": "text",
                "required": True
            },
            {
                "key": "target_type",
                "label": "转换为",
                "type": "select",
                "options": [
                    {"value": "string", "label": "文本 (字符串)"},
                    {"value": "int", "label": "整数"},
                    {"value": "float", "label": "小数 (浮点数)"},
                    {"value": "datetime", "label": "日期/时间"},
                    {"value": "bool", "label": "布尔值 (True/False)"}
                ],
                "default": "string"
            },
            {
                "key": "date_format",
                "label": "日期格式 (用于日期类型)",
                "type": "text",
                "default": "",
                "placeholder": "%Y-%m-%d 或留空自动识别"
            },
            {
                "key": "errors",
                "label": "出错时",
                "type": "select",
                "options": [
                    {"value": "coerce", "label": "设为空值"},
                    {"value": "ignore", "label": "保持原值"},
                    {"value": "raise", "label": "报错"}
                ],
                "default": "coerce"
            }
        ]
    
    def execute(self, input_data: Dict[str, Any]) -> Dict[str, Any]:
        df = input_data.get("data")
        if df is None:
            raise ValueError("No input data received")
        
        df = df.copy()
        column = self.get_param("column")
        target_type = self.get_param("target_type", "string")
        date_format = self.get_param("date_format", "") or None
        errors = self.get_param("errors", "coerce")
        
        if column not in df.columns:
            raise ValueError(f"Column '{column}' not found")
        
        if target_type == "string":
            df[column] = df[column].astype(str)
        elif target_type == "int":
            df[column] = pd.to_numeric(df[column], errors=errors).astype('Int64')
        elif target_type == "float":
            df[column] = pd.to_numeric(df[column], errors=errors)
        elif target_type == "datetime":
            df[column] = pd.to_datetime(df[column], format=date_format, errors=errors)
        elif target_type == "bool":
            df[column] = df[column].astype(bool)
        
        return {"data": df}


# ============================================================
# 文本处理节点 (Text Processing)
# ============================================================

@register_node
class TextCaseNode(BaseNode):
    """Node to change text case"""
    
    node_type = "text_case"
    node_name = "转换大小写"
    node_category = "文本"
    node_description = "将文本转换为大写、小写或首字母大写"
    node_color = "#a855f7"  # Purple
    
    def _setup_ports(self):
        self.add_input("data")
        self.add_output("data")
    
    def get_config_ui_schema(self) -> List[Dict[str, Any]]:
        return [
            {
                "key": "column",
                "label": "列名",
                "type": "text",
                "required": True
            },
            {
                "key": "case_type",
                "label": "转换为",
                "type": "select",
                "options": [
                    {"value": "upper", "label": "全部大写"},
                    {"value": "lower", "label": "全部小写"},
                    {"value": "title", "label": "首字母大写"},
                    {"value": "capitalize", "label": "句首大写"}
                ],
                "default": "upper"
            }
        ]
    
    def execute(self, input_data: Dict[str, Any]) -> Dict[str, Any]:
        df = input_data.get("data")
        if df is None:
            raise ValueError("No input data received")
        
        df = df.copy()
        column = self.get_param("column")
        case_type = self.get_param("case_type", "upper")
        
        if column not in df.columns:
            raise ValueError(f"Column '{column}' not found")
        
        if case_type == "upper":
            df[column] = df[column].astype(str).str.upper()
        elif case_type == "lower":
            df[column] = df[column].astype(str).str.lower()
        elif case_type == "title":
            df[column] = df[column].astype(str).str.title()
        elif case_type == "capitalize":
            df[column] = df[column].astype(str).str.capitalize()
        
        return {"data": df}


@register_node
class SplitColumnNode(BaseNode):
    """Node to split a column into multiple columns"""
    
    node_type = "split_column"
    node_name = "拆分列"
    node_category = "文本"
    node_description = "按分隔符将一列拆分为多列"
    node_color = "#a855f7"  # Purple
    
    def _setup_ports(self):
        self.add_input("data")
        self.add_output("data")
    
    def get_config_ui_schema(self) -> List[Dict[str, Any]]:
        return [
            {
                "key": "column",
                "label": "要拆分的列",
                "type": "text",
                "required": True
            },
            {
                "key": "delimiter",
                "label": "分隔符",
                "type": "text",
                "default": ","
            },
            {
                "key": "new_column_names",
                "label": "新列名 (逗号分隔)",
                "type": "text",
                "placeholder": "列1, 列2, 列3"
            },
            {
                "key": "max_splits",
                "label": "最大拆分次数 (0表示无限制)",
                "type": "number",
                "default": 0
            },
            {
                "key": "keep_original",
                "label": "保留原列",
                "type": "checkbox",
                "default": False
            }
        ]
    
    def execute(self, input_data: Dict[str, Any]) -> Dict[str, Any]:
        df = input_data.get("data")
        if df is None:
            raise ValueError("No input data received")
        
        df = df.copy()
        column = self.get_param("column")
        delimiter = self.get_param("delimiter", ",")
        new_names = self.get_param("new_column_names", "")
        max_splits = self.get_param("max_splits", 0)
        keep_original = self.get_param("keep_original", False)
        
        if column not in df.columns:
            raise ValueError(f"Column '{column}' not found")
        
        n = max_splits if max_splits > 0 else -1
        split_df = df[column].astype(str).str.split(delimiter, n=n, expand=True)
        
        if new_names.strip():
            names = [n.strip() for n in new_names.split(",")]
            split_df.columns = names[:len(split_df.columns)]
        else:
            split_df.columns = [f"{column}_{i+1}" for i in range(len(split_df.columns))]
        
        # Insert new columns after the original
        col_idx = df.columns.get_loc(column)
        for i, col in enumerate(split_df.columns):
            df.insert(col_idx + 1 + i, col, split_df[col])
        
        if not keep_original:
            df = df.drop(columns=[column])
        
        return {"data": df}


@register_node
class CombineColumnsNode(BaseNode):
    """Node to combine multiple columns into one"""
    
    node_type = "combine_columns"
    node_name = "合并列"
    node_category = "文本"
    node_description = "用分隔符将多列合并为一列"
    node_color = "#a855f7"  # Purple
    
    def _setup_ports(self):
        self.add_input("data")
        self.add_output("data")
    
    def get_config_ui_schema(self) -> List[Dict[str, Any]]:
        return [
            {
                "key": "columns",
                "label": "要合并的列 (逗号分隔)",
                "type": "text",
                "required": True
            },
            {
                "key": "new_column_name",
                "label": "新列名称",
                "type": "text",
                "required": True
            },
            {
                "key": "separator",
                "label": "分隔符",
                "type": "text",
                "default": " "
            },
            {
                "key": "drop_original",
                "label": "删除原列",
                "type": "checkbox",
                "default": False
            }
        ]
    
    def execute(self, input_data: Dict[str, Any]) -> Dict[str, Any]:
        df = input_data.get("data")
        if df is None:
            raise ValueError("No input data received")
        
        df = df.copy()
        columns = [c.strip() for c in self.get_param("columns", "").split(",")]
        new_name = self.get_param("new_column_name")
        separator = self.get_param("separator", " ")
        drop_original = self.get_param("drop_original", False)
        
        missing = [c for c in columns if c not in df.columns]
        if missing:
            raise ValueError(f"Columns not found: {missing}")
        
        df[new_name] = df[columns].astype(str).agg(separator.join, axis=1)
        
        if drop_original:
            df = df.drop(columns=columns)
        
        return {"data": df}


@register_node
class ExtractTextNode(BaseNode):
    """Node to extract text using patterns"""
    
    node_type = "extract_text"
    node_name = "提取文本"
    node_category = "文本"
    node_description = "使用模式提取文本 (正则或位置)"
    node_color = "#a855f7"  # Purple
    
    def _setup_ports(self):
        self.add_input("data")
        self.add_output("data")
    
    def get_config_ui_schema(self) -> List[Dict[str, Any]]:
        return [
            {
                "key": "column",
                "label": "源列",
                "type": "text",
                "required": True
            },
            {
                "key": "new_column",
                "label": "新列名称",
                "type": "text",
                "required": True
            },
            {
                "key": "method",
                "label": "提取方式",
                "type": "select",
                "options": [
                    {"value": "regex", "label": "正则表达式"},
                    {"value": "left", "label": "前N个字符"},
                    {"value": "right", "label": "后N个字符"},
                    {"value": "mid", "label": "中间 (起始位置, 长度)"}
                ],
                "default": "left"
            },
            {
                "key": "pattern",
                "label": "模式 / 起始位置",
                "type": "text",
                "required": True,
                "placeholder": "正则模式或数字"
            },
            {
                "key": "length",
                "label": "长度 (用于中间提取)",
                "type": "number",
                "default": 10
            }
        ]
    
    def execute(self, input_data: Dict[str, Any]) -> Dict[str, Any]:
        df = input_data.get("data")
        if df is None:
            raise ValueError("No input data received")
        
        df = df.copy()
        column = self.get_param("column")
        new_column = self.get_param("new_column")
        method = self.get_param("method", "left")
        pattern = self.get_param("pattern", "")
        length = self.get_param("length", 10)
        
        if column not in df.columns:
            raise ValueError(f"Column '{column}' not found")
        
        if method == "regex":
            df[new_column] = df[column].astype(str).str.extract(f'({pattern})', expand=False)
        elif method == "left":
            n = int(pattern)
            df[new_column] = df[column].astype(str).str[:n]
        elif method == "right":
            n = int(pattern)
            df[new_column] = df[column].astype(str).str[-n:]
        elif method == "mid":
            start = int(pattern)
            df[new_column] = df[column].astype(str).str[start:start+length]
        
        return {"data": df}


# ============================================================
# 日期时间节点 (Date/Time)
# ============================================================

@register_node
class ExtractDatePartsNode(BaseNode):
    """Node to extract parts from date columns"""
    
    node_type = "extract_date_parts"
    node_name = "提取日期部分"
    node_category = "日期/时间"
    node_description = "从日期列提取年、月、日等"
    node_color = "#f97316"  # Orange
    
    def _setup_ports(self):
        self.add_input("data")
        self.add_output("data")
    
    def get_config_ui_schema(self) -> List[Dict[str, Any]]:
        return [
            {
                "key": "column",
                "label": "日期列",
                "type": "text",
                "required": True
            },
            {
                "key": "extract_year",
                "label": "提取年",
                "type": "checkbox",
                "default": True
            },
            {
                "key": "extract_month",
                "label": "提取月",
                "type": "checkbox",
                "default": True
            },
            {
                "key": "extract_day",
                "label": "提取日",
                "type": "checkbox",
                "default": True
            },
            {
                "key": "extract_weekday",
                "label": "提取星期 (0=周一)",
                "type": "checkbox",
                "default": False
            },
            {
                "key": "extract_quarter",
                "label": "提取季度",
                "type": "checkbox",
                "default": False
            },
            {
                "key": "extract_week",
                "label": "提取周数",
                "type": "checkbox",
                "default": False
            }
        ]
    
    def execute(self, input_data: Dict[str, Any]) -> Dict[str, Any]:
        df = input_data.get("data")
        if df is None:
            raise ValueError("No input data received")
        
        df = df.copy()
        column = self.get_param("column")
        
        if column not in df.columns:
            raise ValueError(f"Column '{column}' not found")
        
        # Convert to datetime if not already
        dt_col = pd.to_datetime(df[column], errors='coerce')
        
        if self.get_param("extract_year", True):
            df[f"{column}_year"] = dt_col.dt.year
        if self.get_param("extract_month", True):
            df[f"{column}_month"] = dt_col.dt.month
        if self.get_param("extract_day", True):
            df[f"{column}_day"] = dt_col.dt.day
        if self.get_param("extract_weekday", False):
            df[f"{column}_weekday"] = dt_col.dt.weekday
        if self.get_param("extract_quarter", False):
            df[f"{column}_quarter"] = dt_col.dt.quarter
        if self.get_param("extract_week", False):
            df[f"{column}_week"] = dt_col.dt.isocalendar().week
        
        return {"data": df}


@register_node
class FormatDateNode(BaseNode):
    """Node to format date columns"""
    
    node_type = "format_date"
    node_name = "格式化日期"
    node_category = "日期/时间"
    node_description = "将日期转换为特定格式"
    node_color = "#f97316"  # Orange
    
    def _setup_ports(self):
        self.add_input("data")
        self.add_output("data")
    
    def get_config_ui_schema(self) -> List[Dict[str, Any]]:
        return [
            {
                "key": "column",
                "label": "日期列",
                "type": "text",
                "required": True
            },
            {
                "key": "output_format",
                "label": "输出格式",
                "type": "select",
                "options": [
                    {"value": "%Y-%m-%d", "label": "2024-01-31"},
                    {"value": "%d/%m/%Y", "label": "31/01/2024"},
                    {"value": "%m/%d/%Y", "label": "01/31/2024"},
                    {"value": "%Y/%m/%d", "label": "2024/01/31"},
                    {"value": "%d-%m-%Y", "label": "31-01-2024"},
                    {"value": "%Y年%m月%d日", "label": "2024年01月31日"},
                    {"value": "%B %d, %Y", "label": "January 31, 2024"},
                    {"value": "custom", "label": "自定义格式"}
                ],
                "default": "%Y-%m-%d"
            },
            {
                "key": "custom_format",
                "label": "自定义格式 (如果选择)",
                "type": "text",
                "placeholder": "%Y-%m-%d %H:%M:%S"
            },
            {
                "key": "new_column",
                "label": "新列名称 (留空则覆盖原列)",
                "type": "text",
                "default": ""
            }
        ]
    
    def execute(self, input_data: Dict[str, Any]) -> Dict[str, Any]:
        df = input_data.get("data")
        if df is None:
            raise ValueError("No input data received")
        
        df = df.copy()
        column = self.get_param("column")
        output_format = self.get_param("output_format", "%Y-%m-%d")
        custom_format = self.get_param("custom_format", "")
        new_column = self.get_param("new_column", "")
        
        if column not in df.columns:
            raise ValueError(f"Column '{column}' not found")
        
        if output_format == "custom" and custom_format:
            output_format = custom_format
        
        dt_col = pd.to_datetime(df[column], errors='coerce')
        formatted = dt_col.dt.strftime(output_format)
        
        if new_column.strip():
            df[new_column] = formatted
        else:
            df[column] = formatted
        
        return {"data": df}


@register_node
class DateDifferenceNode(BaseNode):
    """Node to calculate difference between dates"""
    
    node_type = "date_difference"
    node_name = "日期差异"
    node_category = "日期/时间"
    node_description = "计算两个日期列之间的差异"
    node_color = "#f97316"  # Orange
    
    def _setup_ports(self):
        self.add_input("data")
        self.add_output("data")
    
    def get_config_ui_schema(self) -> List[Dict[str, Any]]:
        return [
            {
                "key": "start_column",
                "label": "开始日期列",
                "type": "text",
                "required": True
            },
            {
                "key": "end_column",
                "label": "结束日期列",
                "type": "text",
                "required": True
            },
            {
                "key": "new_column",
                "label": "新列名称",
                "type": "text",
                "default": "date_diff"
            },
            {
                "key": "unit",
                "label": "单位",
                "type": "select",
                "options": [
                    {"value": "days", "label": "天"},
                    {"value": "weeks", "label": "周"},
                    {"value": "months", "label": "月 (约)"},
                    {"value": "years", "label": "年 (约)"},
                    {"value": "hours", "label": "小时"},
                    {"value": "minutes", "label": "分钟"}
                ],
                "default": "days"
            }
        ]
    
    def execute(self, input_data: Dict[str, Any]) -> Dict[str, Any]:
        df = input_data.get("data")
        if df is None:
            raise ValueError("No input data received")
        
        df = df.copy()
        start_col = self.get_param("start_column")
        end_col = self.get_param("end_column")
        new_column = self.get_param("new_column", "date_diff")
        unit = self.get_param("unit", "days")
        
        start_dt = pd.to_datetime(df[start_col], errors='coerce')
        end_dt = pd.to_datetime(df[end_col], errors='coerce')
        
        diff = end_dt - start_dt
        
        if unit == "days":
            df[new_column] = diff.dt.days
        elif unit == "weeks":
            df[new_column] = diff.dt.days / 7
        elif unit == "months":
            df[new_column] = diff.dt.days / 30.44
        elif unit == "years":
            df[new_column] = diff.dt.days / 365.25
        elif unit == "hours":
            df[new_column] = diff.dt.total_seconds() / 3600
        elif unit == "minutes":
            df[new_column] = diff.dt.total_seconds() / 60
        
        return {"data": df}


# ============================================================
# 条件逻辑节点 (Conditional Logic)
# ============================================================

@register_node
class ConditionalColumnNode(BaseNode):
    """Node to create column based on conditions (IF/ELSE)"""
    
    node_type = "conditional_column"
    node_name = "IF条件"
    node_category = "逻辑"
    node_description = "根据IF/ELSE条件创建新列"
    node_color = "#eab308"  # Yellow
    
    def _setup_ports(self):
        self.add_input("data")
        self.add_output("data")
    
    def get_config_ui_schema(self) -> List[Dict[str, Any]]:
        return [
            {
                "key": "new_column",
                "label": "新列名称",
                "type": "text",
                "required": True
            },
            {
                "key": "condition_column",
                "label": "要检查的列",
                "type": "text",
                "required": True
            },
            {
                "key": "operator",
                "label": "条件",
                "type": "select",
                "options": [
                    {"value": "==", "label": "等于"},
                    {"value": "!=", "label": "不等于"},
                    {"value": ">", "label": "大于"},
                    {"value": ">=", "label": "大于等于"},
                    {"value": "<", "label": "小于"},
                    {"value": "<=", "label": "小于等于"},
                    {"value": "contains", "label": "包含"},
                    {"value": "isnull", "label": "为空"},
                    {"value": "notnull", "label": "不为空"}
                ],
                "default": "=="
            },
            {
                "key": "condition_value",
                "label": "比较值",
                "type": "text",
                "default": ""
            },
            {
                "key": "true_value",
                "label": "TRUE时的值",
                "type": "text",
                "required": True
            },
            {
                "key": "false_value",
                "label": "FALSE时的值",
                "type": "text",
                "required": True
            }
        ]
    
    def execute(self, input_data: Dict[str, Any]) -> Dict[str, Any]:
        import numpy as np
        df = input_data.get("data")
        if df is None:
            raise ValueError("No input data received")
        
        df = df.copy()
        new_column = self.get_param("new_column")
        cond_column = self.get_param("condition_column")
        operator = self.get_param("operator", "==")
        cond_value = self.get_param("condition_value", "")
        true_value = self.get_param("true_value")
        false_value = self.get_param("false_value")
        
        if cond_column not in df.columns:
            raise ValueError(f"Column '{cond_column}' not found")
        
        # Try to convert to numeric
        try:
            cond_value = float(cond_value)
            if cond_value.is_integer():
                cond_value = int(cond_value)
        except (ValueError, TypeError):
            pass
        
        col = df[cond_column]
        
        if operator == "==":
            mask = col == cond_value
        elif operator == "!=":
            mask = col != cond_value
        elif operator == ">":
            mask = col > cond_value
        elif operator == ">=":
            mask = col >= cond_value
        elif operator == "<":
            mask = col < cond_value
        elif operator == "<=":
            mask = col <= cond_value
        elif operator == "contains":
            mask = col.astype(str).str.contains(str(cond_value), case=False, na=False)
        elif operator == "isnull":
            mask = col.isna()
        elif operator == "notnull":
            mask = col.notna()
        else:
            mask = pd.Series([True] * len(df))
        
        df[new_column] = np.where(mask, true_value, false_value)
        
        return {"data": df}


@register_node
class MultiConditionNode(BaseNode):
    """Node for multiple conditions (like nested IF or CASE/SWITCH)"""
    
    node_type = "multi_condition"
    node_name = "多条件判断"
    node_category = "逻辑"
    node_description = "应用多个条件 (CASE/SWITCH风格)"
    node_color = "#eab308"  # Yellow
    
    def _setup_ports(self):
        self.add_input("data")
        self.add_output("data")
    
    def get_config_ui_schema(self) -> List[Dict[str, Any]]:
        return [
            {
                "key": "new_column",
                "label": "新列名称",
                "type": "text",
                "required": True
            },
            {
                "key": "condition_column",
                "label": "要检查的列",
                "type": "text",
                "required": True
            },
            {
                "key": "conditions",
                "label": "条件 (值:结果，每行一个)",
                "type": "textarea",
                "required": True,
                "placeholder": "A:类别A\nB:类别B\nC:类别C"
            },
            {
                "key": "default_value",
                "label": "默认值 (无匹配时)",
                "type": "text",
                "default": "其他"
            }
        ]
    
    def execute(self, input_data: Dict[str, Any]) -> Dict[str, Any]:
        df = input_data.get("data")
        if df is None:
            raise ValueError("No input data received")
        
        df = df.copy()
        new_column = self.get_param("new_column")
        cond_column = self.get_param("condition_column")
        conditions_str = self.get_param("conditions", "")
        default_value = self.get_param("default_value", "Other")
        
        if cond_column not in df.columns:
            raise ValueError(f"Column '{cond_column}' not found")
        
        # Parse conditions
        mapping = {}
        for line in conditions_str.strip().split("\n"):
            if ":" in line:
                key, val = line.split(":", 1)
                mapping[key.strip()] = val.strip()
        
        df[new_column] = df[cond_column].astype(str).map(mapping).fillna(default_value)
        
        return {"data": df}


# ============================================================
# 数据分析节点 (Data Analysis)
# ============================================================

@register_node
class PivotTableNode(BaseNode):
    """Node to create pivot table"""
    
    node_type = "pivot_table"
    node_name = "数据透视表"
    node_category = "汇总"
    node_description = "创建数据透视表汇总"
    node_color = "#ec4899"  # Pink
    
    def _setup_ports(self):
        self.add_input("data")
        self.add_output("data")
    
    def get_config_ui_schema(self) -> List[Dict[str, Any]]:
        return [
            {
                "key": "index_columns",
                "label": "行标签 (逗号分隔)",
                "type": "text",
                "required": True
            },
            {
                "key": "column_labels",
                "label": "列标签 (逗号分隔)",
                "type": "text",
                "default": ""
            },
            {
                "key": "value_column",
                "label": "值列",
                "type": "text",
                "required": True
            },
            {
                "key": "agg_function",
                "label": "汇总方式",
                "type": "select",
                "options": [
                    {"value": "sum", "label": "求和"},
                    {"value": "mean", "label": "平均值"},
                    {"value": "count", "label": "计数"},
                    {"value": "min", "label": "最小值"},
                    {"value": "max", "label": "最大值"}
                ],
                "default": "sum"
            },
            {
                "key": "fill_value",
                "label": "空值填充",
                "type": "text",
                "default": "0"
            }
        ]
    
    def execute(self, input_data: Dict[str, Any]) -> Dict[str, Any]:
        df = input_data.get("data")
        if df is None:
            raise ValueError("No input data received")
        
        index_cols = [c.strip() for c in self.get_param("index_columns", "").split(",")]
        col_labels = self.get_param("column_labels", "").strip()
        value_col = self.get_param("value_column")
        agg_func = self.get_param("agg_function", "sum")
        fill_value = self.get_param("fill_value", "0")
        
        try:
            fill_value = float(fill_value)
        except ValueError:
            pass
        
        columns = [c.strip() for c in col_labels.split(",")] if col_labels else None
        
        pivot = pd.pivot_table(
            df,
            index=index_cols,
            columns=columns,
            values=value_col,
            aggfunc=agg_func,
            fill_value=fill_value
        ).reset_index()
        
        # Flatten column names if multi-level
        if isinstance(pivot.columns, pd.MultiIndex):
            pivot.columns = ['_'.join(str(c) for c in col).strip('_') for col in pivot.columns]
        
        return {"data": pivot}


@register_node
class VLookupNode(BaseNode):
    """Node to perform VLOOKUP-like operation"""
    
    node_type = "vlookup"
    node_name = "查找匹配(VLOOKUP)"
    node_category = "合并"
    node_description = "从另一个表查找值 (类似Excel VLOOKUP)"
    node_color = "#8b5cf6"  # Purple
    
    def _setup_ports(self):
        self.add_input("main_data")
        self.add_input("lookup_data")
        self.add_output("data")
    
    def get_config_ui_schema(self) -> List[Dict[str, Any]]:
        return [
            {
                "key": "main_key",
                "label": "主表关键列",
                "type": "text",
                "required": True
            },
            {
                "key": "lookup_key",
                "label": "查找表关键列",
                "type": "text",
                "required": True
            },
            {
                "key": "return_columns",
                "label": "返回列 (逗号分隔)",
                "type": "text",
                "required": True,
                "placeholder": "列1, 列2"
            },
            {
                "key": "no_match_value",
                "label": "无匹配时的值",
                "type": "text",
                "default": ""
            }
        ]
    
    def execute(self, input_data: Dict[str, Any]) -> Dict[str, Any]:
        main_df = input_data.get("main_data")
        lookup_df = input_data.get("lookup_data")
        
        if main_df is None:
            raise ValueError("No main data received")
        if lookup_df is None:
            raise ValueError("No lookup data received")
        
        main_key = self.get_param("main_key")
        lookup_key = self.get_param("lookup_key")
        return_cols = [c.strip() for c in self.get_param("return_columns", "").split(",")]
        no_match = self.get_param("no_match_value", "")
        
        # Select only needed columns from lookup
        lookup_subset = lookup_df[[lookup_key] + return_cols].drop_duplicates(subset=[lookup_key])
        
        # Merge
        result = main_df.merge(
            lookup_subset,
            left_on=main_key,
            right_on=lookup_key,
            how='left'
        )
        
        # Fill missing values
        if no_match:
            for col in return_cols:
                if col in result.columns:
                    result[col] = result[col].fillna(no_match)
        
        # Remove duplicate key column if different names
        if main_key != lookup_key and lookup_key in result.columns:
            result = result.drop(columns=[lookup_key])
        
        return {"data": result}


@register_node
class StatisticsSummaryNode(BaseNode):
    """Node to generate statistics summary"""
    
    node_type = "statistics_summary"
    node_name = "统计摘要"
    node_category = "汇总"
    node_description = "生成数值列的统计摘要"
    node_color = "#ec4899"  # Pink
    
    def _setup_ports(self):
        self.add_input("data")
        self.add_output("data")
        self.add_output("summary")
    
    def get_config_ui_schema(self) -> List[Dict[str, Any]]:
        return [
            {
                "key": "columns",
                "label": "列名 (逗号分隔，留空表示所有数值列)",
                "type": "text",
                "default": ""
            }
        ]
    
    def execute(self, input_data: Dict[str, Any]) -> Dict[str, Any]:
        df = input_data.get("data")
        if df is None:
            raise ValueError("No input data received")
        
        columns_str = self.get_param("columns", "")
        
        if columns_str.strip():
            columns = [c.strip() for c in columns_str.split(",")]
        else:
            columns = df.select_dtypes(include=['number']).columns.tolist()
        
        # Generate summary
        summary_data = []
        for col in columns:
            if col in df.columns:
                stats = {
                    'Column': col,
                    'Count': df[col].count(),
                    'Mean': df[col].mean(),
                    'Std': df[col].std(),
                    'Min': df[col].min(),
                    '25%': df[col].quantile(0.25),
                    '50%': df[col].quantile(0.50),
                    '75%': df[col].quantile(0.75),
                    'Max': df[col].max(),
                    'Null Count': df[col].isna().sum()
                }
                summary_data.append(stats)
        
        summary_df = pd.DataFrame(summary_data)
        
        return {"data": df, "summary": summary_df}


@register_node
class SampleDataNode(BaseNode):
    """Node to sample data"""
    
    node_type = "sample_data"
    node_name = "数据抽样"
    node_category = "转换"
    node_description = "随机或系统地抽取数据样本"
    node_color = "#f59e0b"  # Amber
    
    def _setup_ports(self):
        self.add_input("data")
        self.add_output("data")
    
    def get_config_ui_schema(self) -> List[Dict[str, Any]]:
        return [
            {
                "key": "method",
                "label": "抽样方式",
                "type": "select",
                "options": [
                    {"value": "n", "label": "固定行数"},
                    {"value": "frac", "label": "百分比"},
                    {"value": "first", "label": "前N行"},
                    {"value": "last", "label": "后N行"},
                    {"value": "every_nth", "label": "每N行"}
                ],
                "default": "n"
            },
            {
                "key": "value",
                "label": "值 (数字或百分比)",
                "type": "text",
                "required": True,
                "placeholder": "100 或 0.1 表示10%"
            },
            {
                "key": "random_seed",
                "label": "随机种子 (用于可重复性)",
                "type": "number",
                "default": 42
            }
        ]
    
    def execute(self, input_data: Dict[str, Any]) -> Dict[str, Any]:
        df = input_data.get("data")
        if df is None:
            raise ValueError("No input data received")
        
        method = self.get_param("method", "n")
        value = self.get_param("value", "100")
        seed = self.get_param("random_seed", 42)
        
        if method == "n":
            n = min(int(value), len(df))
            result = df.sample(n=n, random_state=seed)
        elif method == "frac":
            frac = float(value)
            result = df.sample(frac=frac, random_state=seed)
        elif method == "first":
            n = int(value)
            result = df.head(n)
        elif method == "last":
            n = int(value)
            result = df.tail(n)
        elif method == "every_nth":
            n = int(value)
            result = df.iloc[::n]
        else:
            result = df
        
        return {"data": result.reset_index(drop=True)}


@register_node
class DuplicateReportNode(BaseNode):
    """Node to find and report duplicates"""
    
    node_type = "duplicate_report"
    node_name = "查找重复"
    node_category = "验证"
    node_description = "查找并报告重复行"
    node_color = "#ef4444"  # Red
    
    def _setup_ports(self):
        self.add_input("data")
        self.add_output("duplicates")  # Only duplicated rows
        self.add_output("unique")  # Only unique rows
        self.add_output("all_data")  # Original with duplicate flag
    
    def get_config_ui_schema(self) -> List[Dict[str, Any]]:
        return [
            {
                "key": "columns",
                "label": "检查列 (逗号分隔，留空表示所有列)",
                "type": "text",
                "default": ""
            },
            {
                "key": "keep",
                "label": "标记为重复",
                "type": "select",
                "options": [
                    {"value": "first", "label": "除第一次出现外全部"},
                    {"value": "last", "label": "除最后一次出现外全部"},
                    {"value": False, "label": "所有重复项"}
                ],
                "default": "first"
            }
        ]
    
    def execute(self, input_data: Dict[str, Any]) -> Dict[str, Any]:
        df = input_data.get("data")
        if df is None:
            raise ValueError("No input data received")
        
        df = df.copy()
        columns_str = self.get_param("columns", "")
        keep = self.get_param("keep", "first")
        
        if keep == "False":
            keep = False
        
        columns = [c.strip() for c in columns_str.split(",") if c.strip()] or None
        
        # Find duplicates
        is_dup = df.duplicated(subset=columns, keep=keep)
        
        df['_is_duplicate'] = is_dup
        
        duplicates_df = df[is_dup].drop(columns=['_is_duplicate']).reset_index(drop=True)
        unique_df = df[~is_dup].drop(columns=['_is_duplicate']).reset_index(drop=True)
        
        return {
            "duplicates": duplicates_df,
            "unique": unique_df,
            "all_data": df
        }


@register_node
class DataValidationNode(BaseNode):
    """Node to validate data and create report"""
    
    node_type = "data_validation"
    node_name = "数据验证"
    node_category = "验证"
    node_description = "验证数据并生成报告"
    node_color = "#ef4444"  # Red
    
    def _setup_ports(self):
        self.add_input("data")
        self.add_output("valid")
        self.add_output("invalid")
        self.add_output("report")
    
    def get_config_ui_schema(self) -> List[Dict[str, Any]]:
        return [
            {
                "key": "column",
                "label": "要验证的列",
                "type": "text",
                "required": True
            },
            {
                "key": "rule",
                "label": "验证规则",
                "type": "select",
                "options": [
                    {"value": "not_empty", "label": "不能为空"},
                    {"value": "is_number", "label": "必须是数字"},
                    {"value": "is_email", "label": "必须是有效邮箱"},
                    {"value": "min_length", "label": "最小文本长度"},
                    {"value": "max_length", "label": "最大文本长度"},
                    {"value": "in_range", "label": "数字在范围内"},
                    {"value": "in_list", "label": "值在列表中"},
                    {"value": "regex", "label": "匹配模式"}
                ],
                "default": "not_empty"
            },
            {
                "key": "rule_value",
                "label": "规则值 (根据规则)",
                "type": "text",
                "placeholder": "最小,最大 或 列表 或 模式"
            }
        ]
    
    def execute(self, input_data: Dict[str, Any]) -> Dict[str, Any]:
        import re
        df = input_data.get("data")
        if df is None:
            raise ValueError("No input data received")
        
        df = df.copy()
        column = self.get_param("column")
        rule = self.get_param("rule", "not_empty")
        rule_value = self.get_param("rule_value", "")
        
        if column not in df.columns:
            raise ValueError(f"Column '{column}' not found")
        
        col = df[column]
        
        if rule == "not_empty":
            valid_mask = col.notna() & (col.astype(str).str.strip() != '')
        elif rule == "is_number":
            valid_mask = pd.to_numeric(col, errors='coerce').notna()
        elif rule == "is_email":
            email_pattern = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
            valid_mask = col.astype(str).str.match(email_pattern)
        elif rule == "min_length":
            min_len = int(rule_value)
            valid_mask = col.astype(str).str.len() >= min_len
        elif rule == "max_length":
            max_len = int(rule_value)
            valid_mask = col.astype(str).str.len() <= max_len
        elif rule == "in_range":
            min_val, max_val = map(float, rule_value.split(","))
            numeric_col = pd.to_numeric(col, errors='coerce')
            valid_mask = (numeric_col >= min_val) & (numeric_col <= max_val)
        elif rule == "in_list":
            allowed = [v.strip() for v in rule_value.split(",")]
            valid_mask = col.astype(str).isin(allowed)
        elif rule == "regex":
            valid_mask = col.astype(str).str.match(rule_value)
        else:
            valid_mask = pd.Series([True] * len(df))
        
        valid_df = df[valid_mask].reset_index(drop=True)
        invalid_df = df[~valid_mask].reset_index(drop=True)
        
        report = pd.DataFrame({
            'Metric': ['Total Rows', 'Valid Rows', 'Invalid Rows', 'Valid %'],
            'Value': [
                len(df),
                len(valid_df),
                len(invalid_df),
                f"{len(valid_df)/len(df)*100:.1f}%" if len(df) > 0 else "N/A"
            ]
        })
        
        return {"valid": valid_df, "invalid": invalid_df, "report": report}


@register_node
class NumberFormatNode(BaseNode):
    """Node to format numbers"""
    
    node_type = "number_format"
    node_name = "格式化数字"
    node_category = "转换"
    node_description = "格式化数字 (小数位、货币、百分比)"
    node_color = "#f59e0b"  # Amber
    
    def _setup_ports(self):
        self.add_input("data")
        self.add_output("data")
    
    def get_config_ui_schema(self) -> List[Dict[str, Any]]:
        return [
            {
                "key": "column",
                "label": "列名",
                "type": "text",
                "required": True
            },
            {
                "key": "format_type",
                "label": "格式类型",
                "type": "select",
                "options": [
                    {"value": "decimal", "label": "小数位数"},
                    {"value": "currency", "label": "货币"},
                    {"value": "percentage", "label": "百分比"},
                    {"value": "thousands", "label": "千位分隔符"},
                    {"value": "scientific", "label": "科学记数法"}
                ],
                "default": "decimal"
            },
            {
                "key": "decimal_places",
                "label": "小数位数",
                "type": "number",
                "default": 2
            },
            {
                "key": "currency_symbol",
                "label": "货币符号",
                "type": "text",
                "default": "￥"
            }
        ]
    
    def execute(self, input_data: Dict[str, Any]) -> Dict[str, Any]:
        df = input_data.get("data")
        if df is None:
            raise ValueError("No input data received")
        
        df = df.copy()
        column = self.get_param("column")
        format_type = self.get_param("format_type", "decimal")
        decimal_places = self.get_param("decimal_places", 2)
        currency_symbol = self.get_param("currency_symbol", "$")
        
        if column not in df.columns:
            raise ValueError(f"Column '{column}' not found")
        
        col = pd.to_numeric(df[column], errors='coerce')
        
        if format_type == "decimal":
            df[column] = col.round(decimal_places)
        elif format_type == "currency":
            df[column] = col.apply(lambda x: f"{currency_symbol}{x:,.{decimal_places}f}" if pd.notna(x) else "")
        elif format_type == "percentage":
            df[column] = col.apply(lambda x: f"{x*100:.{decimal_places}f}%" if pd.notna(x) else "")
        elif format_type == "thousands":
            df[column] = col.apply(lambda x: f"{x:,.{decimal_places}f}" if pd.notna(x) else "")
        elif format_type == "scientific":
            df[column] = col.apply(lambda x: f"{x:.{decimal_places}e}" if pd.notna(x) else "")
        
        return {"data": df}


@register_node
class RankNode(BaseNode):
    """Node to add ranking column"""
    
    node_type = "rank_data"
    node_name = "添加排名"
    node_category = "转换"
    node_description = "根据值添加排名列"
    node_color = "#f59e0b"  # Amber
    
    def _setup_ports(self):
        self.add_input("data")
        self.add_output("data")
    
    def get_config_ui_schema(self) -> List[Dict[str, Any]]:
        return [
            {
                "key": "column",
                "label": "排名依据列",
                "type": "text",
                "required": True
            },
            {
                "key": "new_column",
                "label": "新排名列名称",
                "type": "text",
                "default": "rank"
            },
            {
                "key": "ascending",
                "label": "升序 (1 = 最小值)",
                "type": "checkbox",
                "default": False
            },
            {
                "key": "method",
                "label": "并列处理方式",
                "type": "select",
                "options": [
                    {"value": "average", "label": "平均排名"},
                    {"value": "min", "label": "最低排名"},
                    {"value": "max", "label": "最高排名"},
                    {"value": "first", "label": "按出现顺序"},
                    {"value": "dense", "label": "密集排名 (无间隔)"}
                ],
                "default": "average"
            },
            {
                "key": "group_by",
                "label": "分组列 (可选)",
                "type": "text",
                "default": ""
            }
        ]
    
    def execute(self, input_data: Dict[str, Any]) -> Dict[str, Any]:
        df = input_data.get("data")
        if df is None:
            raise ValueError("No input data received")
        
        df = df.copy()
        column = self.get_param("column")
        new_column = self.get_param("new_column", "rank")
        ascending = self.get_param("ascending", False)
        method = self.get_param("method", "average")
        group_by = self.get_param("group_by", "").strip()
        
        if column not in df.columns:
            raise ValueError(f"Column '{column}' not found")
        
        if group_by and group_by in df.columns:
            df[new_column] = df.groupby(group_by)[column].rank(
                ascending=ascending, method=method
            )
        else:
            df[new_column] = df[column].rank(ascending=ascending, method=method)
        
        return {"data": df}


# ============================================================
# 数据处理逻辑节点
# ============================================================

@register_node
class MergeDataNode(BaseNode):
    """Node to merge two dataframes"""
    
    node_type = "merge_data"
    node_name = "合并数据"
    node_category = "数据处理"
    node_description = "根据键列合并两个数据表"
    node_color = "#8b5cf6"  # Purple
    
    def _setup_ports(self):
        self.add_input("left")
        self.add_input("right")
        self.add_output("data")
    
    def get_config_ui_schema(self) -> List[Dict[str, Any]]:
        return [
            {
                "key": "left_on",
                "label": "左表键列",
                "type": "text",
                "required": True
            },
            {
                "key": "right_on",
                "label": "右表键列",
                "type": "text",
                "required": True
            },
            {
                "key": "how",
                "label": "合并方式",
                "type": "select",
                "options": [
                    {"value": "inner", "label": "内连接 (仅匹配行)"},
                    {"value": "left", "label": "左连接 (保留左表所有行)"},
                    {"value": "right", "label": "右连接 (保留右表所有行)"},
                    {"value": "outer", "label": "外连接 (保留所有行)"}
                ],
                "default": "inner"
            },
            {
                "key": "suffixes",
                "label": "重复列后缀 (逗号分隔)",
                "type": "text",
                "default": "_左,_右"
            }
        ]
    
    def validate(self) -> tuple[bool, str]:
        if not self.get_param("left_on"):
            return False, "Left key column is required"
        if not self.get_param("right_on"):
            return False, "Right key column is required"
        return True, ""
    
    def execute(self, input_data: Dict[str, Any]) -> Dict[str, Any]:
        left_df = input_data.get("left")
        right_df = input_data.get("right")
        
        if left_df is None:
            raise ValueError("No left input data received")
        if right_df is None:
            raise ValueError("No right input data received")
        
        left_on = self.get_param("left_on")
        right_on = self.get_param("right_on")
        how = self.get_param("how", "inner")
        suffixes_str = self.get_param("suffixes", "_左,_右")
        suffixes = tuple(s.strip() for s in suffixes_str.split(","))
        if len(suffixes) != 2:
            suffixes = ("_左", "_右")
        
        result = pd.merge(
            left_df, right_df,
            left_on=left_on, right_on=right_on,
            how=how, suffixes=suffixes
        )
        
        return {"data": result}


@register_node
class ConcatDataNode(BaseNode):
    """Node to concatenate dataframes"""
    
    node_type = "concat_data"
    node_name = "拼接数据"
    node_category = "数据处理"
    node_description = "纵向或横向拼接多个数据表"
    node_color = "#8b5cf6"  # Purple
    
    def _setup_ports(self):
        self.add_input("data1")
        self.add_input("data2")
        self.add_output("data")
    
    def get_config_ui_schema(self) -> List[Dict[str, Any]]:
        return [
            {
                "key": "axis",
                "label": "拼接方向",
                "type": "select",
                "options": [
                    {"value": "0", "label": "纵向 (上下拼接)"},
                    {"value": "1", "label": "横向 (左右拼接)"}
                ],
                "default": "0"
            },
            {
                "key": "ignore_index",
                "label": "重置索引",
                "type": "checkbox",
                "default": True
            }
        ]
    
    def execute(self, input_data: Dict[str, Any]) -> Dict[str, Any]:
        dfs = []
        for key in ["data1", "data2"]:
            if input_data.get(key) is not None:
                dfs.append(input_data[key])
        
        if not dfs:
            raise ValueError("No input data received")
        
        axis = int(self.get_param("axis", "0"))
        ignore_index = self.get_param("ignore_index", True)
        
        result = pd.concat(dfs, axis=axis, ignore_index=ignore_index)
        
        return {"data": result}


@register_node
class GroupAggregateNode(BaseNode):
    """Node to group and aggregate data"""
    
    node_type = "group_aggregate"
    node_name = "分组聚合"
    node_category = "数据处理"
    node_description = "按列分组并进行聚合计算"
    node_color = "#8b5cf6"  # Purple
    
    def _setup_ports(self):
        self.add_input("data")
        self.add_output("data")
    
    def get_config_ui_schema(self) -> List[Dict[str, Any]]:
        return [
            {
                "key": "group_by",
                "label": "分组列 (逗号分隔)",
                "type": "text",
                "required": True
            },
            {
                "key": "agg_column",
                "label": "聚合列",
                "type": "text",
                "required": True
            },
            {
                "key": "agg_func",
                "label": "聚合函数",
                "type": "select",
                "options": [
                    {"value": "sum", "label": "求和 (sum)"},
                    {"value": "mean", "label": "平均值 (mean)"},
                    {"value": "count", "label": "计数 (count)"},
                    {"value": "min", "label": "最小值 (min)"},
                    {"value": "max", "label": "最大值 (max)"},
                    {"value": "median", "label": "中位数 (median)"},
                    {"value": "std", "label": "标准差 (std)"},
                    {"value": "var", "label": "方差 (var)"},
                    {"value": "first", "label": "第一个值"},
                    {"value": "last", "label": "最后一个值"}
                ],
                "default": "sum"
            },
            {
                "key": "new_column",
                "label": "结果列名称",
                "type": "text",
                "default": ""
            }
        ]
    
    def validate(self) -> tuple[bool, str]:
        if not self.get_param("group_by"):
            return False, "Group by columns required"
        if not self.get_param("agg_column"):
            return False, "Aggregation column required"
        return True, ""
    
    def execute(self, input_data: Dict[str, Any]) -> Dict[str, Any]:
        df = input_data.get("data")
        if df is None:
            raise ValueError("No input data received")
        
        group_cols = [c.strip() for c in self.get_param("group_by").split(",")]
        agg_col = self.get_param("agg_column")
        agg_func = self.get_param("agg_func", "sum")
        new_col = self.get_param("new_column", "").strip() or f"{agg_col}_{agg_func}"
        
        result = df.groupby(group_cols)[agg_col].agg(agg_func).reset_index()
        result.columns = list(group_cols) + [new_col]
        
        return {"data": result}


@register_node
class PivotTableNode(BaseNode):
    """Node to create pivot table"""
    
    node_type = "pivot_table"
    node_name = "数据透视表"
    node_category = "数据处理"
    node_description = "创建数据透视表"
    node_color = "#8b5cf6"  # Purple
    
    def _setup_ports(self):
        self.add_input("data")
        self.add_output("data")
    
    def get_config_ui_schema(self) -> List[Dict[str, Any]]:
        return [
            {
                "key": "index",
                "label": "行标签列",
                "type": "text",
                "required": True
            },
            {
                "key": "columns",
                "label": "列标签列",
                "type": "text",
                "required": True
            },
            {
                "key": "values",
                "label": "值列",
                "type": "text",
                "required": True
            },
            {
                "key": "aggfunc",
                "label": "聚合函数",
                "type": "select",
                "options": [
                    {"value": "sum", "label": "求和"},
                    {"value": "mean", "label": "平均值"},
                    {"value": "count", "label": "计数"},
                    {"value": "min", "label": "最小值"},
                    {"value": "max", "label": "最大值"}
                ],
                "default": "sum"
            },
            {
                "key": "fill_value",
                "label": "空值填充",
                "type": "text",
                "default": "0"
            }
        ]
    
    def validate(self) -> tuple[bool, str]:
        if not self.get_param("index"):
            return False, "Index column required"
        if not self.get_param("columns"):
            return False, "Columns required"
        if not self.get_param("values"):
            return False, "Values column required"
        return True, ""
    
    def execute(self, input_data: Dict[str, Any]) -> Dict[str, Any]:
        df = input_data.get("data")
        if df is None:
            raise ValueError("No input data received")
        
        index = self.get_param("index")
        columns = self.get_param("columns")
        values = self.get_param("values")
        aggfunc = self.get_param("aggfunc", "sum")
        fill_value = self.get_param("fill_value", "0")
        
        try:
            fill_value = float(fill_value)
        except ValueError:
            fill_value = 0
        
        result = pd.pivot_table(
            df, index=index, columns=columns, values=values,
            aggfunc=aggfunc, fill_value=fill_value
        ).reset_index()
        
        return {"data": result}


@register_node
class ConditionalNode(BaseNode):
    """Node to apply conditional logic"""
    
    node_type = "conditional"
    node_name = "条件判断"
    node_category = "数据处理"
    node_description = "根据条件创建新列或修改值"
    node_color = "#ec4899"  # Pink
    
    def _setup_ports(self):
        self.add_input("data")
        self.add_output("data")
    
    def get_config_ui_schema(self) -> List[Dict[str, Any]]:
        return [
            {
                "key": "condition_column",
                "label": "条件列",
                "type": "text",
                "required": True
            },
            {
                "key": "operator",
                "label": "条件运算符",
                "type": "select",
                "options": [
                    {"value": "==", "label": "等于"},
                    {"value": "!=", "label": "不等于"},
                    {"value": ">", "label": "大于"},
                    {"value": ">=", "label": "大于等于"},
                    {"value": "<", "label": "小于"},
                    {"value": "<=", "label": "小于等于"},
                    {"value": "contains", "label": "包含"},
                    {"value": "isnull", "label": "为空"}
                ],
                "default": "=="
            },
            {
                "key": "condition_value",
                "label": "条件值",
                "type": "text",
                "default": ""
            },
            {
                "key": "new_column",
                "label": "新列名称",
                "type": "text",
                "required": True
            },
            {
                "key": "true_value",
                "label": "条件为真时的值",
                "type": "text",
                "default": "是"
            },
            {
                "key": "false_value",
                "label": "条件为假时的值",
                "type": "text",
                "default": "否"
            }
        ]
    
    def validate(self) -> tuple[bool, str]:
        if not self.get_param("condition_column"):
            return False, "Condition column required"
        if not self.get_param("new_column"):
            return False, "New column name required"
        return True, ""
    
    def execute(self, input_data: Dict[str, Any]) -> Dict[str, Any]:
        df = input_data.get("data")
        if df is None:
            raise ValueError("No input data received")
        
        df = df.copy()
        col = self.get_param("condition_column")
        op = self.get_param("operator", "==")
        val = self.get_param("condition_value", "")
        new_col = self.get_param("new_column")
        true_val = self.get_param("true_value", "是")
        false_val = self.get_param("false_value", "否")
        
        # Try numeric conversion
        try:
            val = float(val)
            if val.is_integer():
                val = int(val)
        except (ValueError, TypeError):
            pass
        
        if op == "==":
            mask = df[col] == val
        elif op == "!=":
            mask = df[col] != val
        elif op == ">":
            mask = df[col] > val
        elif op == ">=":
            mask = df[col] >= val
        elif op == "<":
            mask = df[col] < val
        elif op == "<=":
            mask = df[col] <= val
        elif op == "contains":
            mask = df[col].astype(str).str.contains(str(val), case=False, na=False)
        elif op == "isnull":
            mask = df[col].isna()
        else:
            mask = pd.Series([False] * len(df))
        
        df[new_col] = false_val
        df.loc[mask, new_col] = true_val
        
        return {"data": df}


@register_node
class CalculateColumnNode(BaseNode):
    """Node to calculate new column from expression"""
    
    node_type = "calculate_column"
    node_name = "计算列"
    node_category = "数据处理"
    node_description = "通过表达式计算新列"
    node_color = "#ec4899"  # Pink
    
    def _setup_ports(self):
        self.add_input("data")
        self.add_output("data")
    
    def get_config_ui_schema(self) -> List[Dict[str, Any]]:
        return [
            {
                "key": "new_column",
                "label": "新列名称",
                "type": "text",
                "required": True
            },
            {
                "key": "expression",
                "label": "表达式",
                "type": "text",
                "required": True,
                "placeholder": "例如: [列A] + [列B] * 2"
            },
            {
                "key": "help_text",
                "label": "提示",
                "type": "label",
                "default": "使用 [列名] 引用列, 支持 +, -, *, /, **, //"
            }
        ]
    
    def validate(self) -> tuple[bool, str]:
        if not self.get_param("new_column"):
            return False, "New column name required"
        if not self.get_param("expression"):
            return False, "Expression required"
        return True, ""
    
    def execute(self, input_data: Dict[str, Any]) -> Dict[str, Any]:
        df = input_data.get("data")
        if df is None:
            raise ValueError("No input data received")
        
        df = df.copy()
        new_col = self.get_param("new_column")
        expr = self.get_param("expression", "")
        
        # Replace [column] with df['column']
        import re
        def replace_col(match):
            col_name = match.group(1)
            if col_name in df.columns:
                return f"df['{col_name}']"
            else:
                raise ValueError(f"Column '{col_name}' not found")
        
        expr = re.sub(r'\[([^\]]+)\]', replace_col, expr)
        
        # Safely evaluate the expression
        try:
            df[new_col] = eval(expr)
        except Exception as e:
            raise ValueError(f"Expression error: {e}")
        
        return {"data": df}


@register_node
class SplitDataNode(BaseNode):
    """Node to split data based on condition"""
    
    node_type = "split_data"
    node_name = "拆分数据"
    node_category = "数据处理"
    node_description = "根据条件拆分数据为两个输出"
    node_color = "#06b6d4"  # Cyan
    
    def _setup_ports(self):
        self.add_input("data")
        self.add_output("matched")
        self.add_output("unmatched")
    
    def get_config_ui_schema(self) -> List[Dict[str, Any]]:
        return [
            {
                "key": "column",
                "label": "条件列",
                "type": "text",
                "required": True
            },
            {
                "key": "operator",
                "label": "运算符",
                "type": "select",
                "options": [
                    {"value": "==", "label": "等于"},
                    {"value": "!=", "label": "不等于"},
                    {"value": ">", "label": "大于"},
                    {"value": "<", "label": "小于"},
                    {"value": "contains", "label": "包含"},
                    {"value": "isnull", "label": "为空"}
                ],
                "default": "=="
            },
            {
                "key": "value",
                "label": "值",
                "type": "text",
                "default": ""
            }
        ]
    
    def validate(self) -> tuple[bool, str]:
        if not self.get_param("column"):
            return False, "Column required"
        return True, ""
    
    def execute(self, input_data: Dict[str, Any]) -> Dict[str, Any]:
        df = input_data.get("data")
        if df is None:
            raise ValueError("No input data received")
        
        col = self.get_param("column")
        op = self.get_param("operator", "==")
        val = self.get_param("value", "")
        
        try:
            val = float(val)
            if val.is_integer():
                val = int(val)
        except (ValueError, TypeError):
            pass
        
        if op == "==":
            mask = df[col] == val
        elif op == "!=":
            mask = df[col] != val
        elif op == ">":
            mask = df[col] > val
        elif op == "<":
            mask = df[col] < val
        elif op == "contains":
            mask = df[col].astype(str).str.contains(str(val), case=False, na=False)
        elif op == "isnull":
            mask = df[col].isna()
        else:
            mask = pd.Series([False] * len(df))
        
        return {
            "matched": df[mask].reset_index(drop=True),
            "unmatched": df[~mask].reset_index(drop=True)
        }


@register_node
class FillNullNode(BaseNode):
    """Node to fill null values"""
    
    node_type = "fill_null"
    node_name = "填充空值"
    node_category = "数据处理"
    node_description = "填充数据中的空值"
    node_color = "#06b6d4"  # Cyan
    
    def _setup_ports(self):
        self.add_input("data")
        self.add_output("data")
    
    def get_config_ui_schema(self) -> List[Dict[str, Any]]:
        return [
            {
                "key": "column",
                "label": "列名 (留空表示所有列)",
                "type": "text",
                "default": ""
            },
            {
                "key": "method",
                "label": "填充方式",
                "type": "select",
                "options": [
                    {"value": "value", "label": "固定值"},
                    {"value": "ffill", "label": "向前填充 (用上一行)"},
                    {"value": "bfill", "label": "向后填充 (用下一行)"},
                    {"value": "mean", "label": "平均值"},
                    {"value": "median", "label": "中位数"},
                    {"value": "mode", "label": "众数"}
                ],
                "default": "value"
            },
            {
                "key": "fill_value",
                "label": "填充值 (固定值时使用)",
                "type": "text",
                "default": "0"
            }
        ]
    
    def execute(self, input_data: Dict[str, Any]) -> Dict[str, Any]:
        df = input_data.get("data")
        if df is None:
            raise ValueError("No input data received")
        
        df = df.copy()
        column = self.get_param("column", "").strip()
        method = self.get_param("method", "value")
        fill_val = self.get_param("fill_value", "0")
        
        cols = [column] if column and column in df.columns else df.columns
        
        for col in cols:
            if method == "value":
                try:
                    val = float(fill_val)
                    if val.is_integer():
                        val = int(val)
                except ValueError:
                    val = fill_val
                df[col] = df[col].fillna(val)
            elif method == "ffill":
                df[col] = df[col].ffill()
            elif method == "bfill":
                df[col] = df[col].bfill()
            elif method == "mean":
                if pd.api.types.is_numeric_dtype(df[col]):
                    df[col] = df[col].fillna(df[col].mean())
            elif method == "median":
                if pd.api.types.is_numeric_dtype(df[col]):
                    df[col] = df[col].fillna(df[col].median())
            elif method == "mode":
                mode_val = df[col].mode()
                if len(mode_val) > 0:
                    df[col] = df[col].fillna(mode_val[0])
        
        return {"data": df}


@register_node
class TypeConvertNode(BaseNode):
    """Node to convert column data types"""
    
    node_type = "type_convert"
    node_name = "类型转换"
    node_category = "数据处理"
    node_description = "转换列的数据类型"
    node_color = "#06b6d4"  # Cyan
    
    def _setup_ports(self):
        self.add_input("data")
        self.add_output("data")
    
    def get_config_ui_schema(self) -> List[Dict[str, Any]]:
        return [
            {
                "key": "column",
                "label": "列名",
                "type": "text",
                "required": True
            },
            {
                "key": "target_type",
                "label": "目标类型",
                "type": "select",
                "options": [
                    {"value": "str", "label": "文本 (string)"},
                    {"value": "int", "label": "整数 (int)"},
                    {"value": "float", "label": "小数 (float)"},
                    {"value": "datetime", "label": "日期时间"},
                    {"value": "bool", "label": "布尔值"}
                ],
                "default": "str"
            },
            {
                "key": "date_format",
                "label": "日期格式 (日期时间类型时使用)",
                "type": "text",
                "default": "%Y-%m-%d",
                "placeholder": "%Y-%m-%d %H:%M:%S"
            }
        ]
    
    def validate(self) -> tuple[bool, str]:
        if not self.get_param("column"):
            return False, "Column required"
        return True, ""
    
    def execute(self, input_data: Dict[str, Any]) -> Dict[str, Any]:
        df = input_data.get("data")
        if df is None:
            raise ValueError("No input data received")
        
        df = df.copy()
        col = self.get_param("column")
        target = self.get_param("target_type", "str")
        date_fmt = self.get_param("date_format", "%Y-%m-%d")
        
        if col not in df.columns:
            raise ValueError(f"Column '{col}' not found")
        
        if target == "str":
            df[col] = df[col].astype(str)
        elif target == "int":
            df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0).astype(int)
        elif target == "float":
            df[col] = pd.to_numeric(df[col], errors='coerce')
        elif target == "datetime":
            df[col] = pd.to_datetime(df[col], format=date_fmt, errors='coerce')
        elif target == "bool":
            df[col] = df[col].astype(bool)
        
        return {"data": df}


@register_node
class SampleDataNode(BaseNode):
    """Node to sample data"""
    
    node_type = "sample_data"
    node_name = "抽样数据"
    node_category = "数据处理"
    node_description = "从数据中随机抽取样本"
    node_color = "#14b8a6"  # Teal
    
    def _setup_ports(self):
        self.add_input("data")
        self.add_output("data")
    
    def get_config_ui_schema(self) -> List[Dict[str, Any]]:
        return [
            {
                "key": "method",
                "label": "抽样方式",
                "type": "select",
                "options": [
                    {"value": "n", "label": "固定数量"},
                    {"value": "frac", "label": "百分比"}
                ],
                "default": "n"
            },
            {
                "key": "value",
                "label": "数量/百分比",
                "type": "number",
                "default": 100
            },
            {
                "key": "random_state",
                "label": "随机种子 (可选)",
                "type": "number",
                "default": ""
            }
        ]
    
    def execute(self, input_data: Dict[str, Any]) -> Dict[str, Any]:
        df = input_data.get("data")
        if df is None:
            raise ValueError("No input data received")
        
        method = self.get_param("method", "n")
        value = self.get_param("value", 100)
        random_state = self.get_param("random_state", "")
        
        try:
            random_state = int(random_state) if random_state else None
        except ValueError:
            random_state = None
        
        if method == "n":
            n = min(int(value), len(df))
            result = df.sample(n=n, random_state=random_state)
        else:
            frac = float(value) / 100 if value > 1 else float(value)
            result = df.sample(frac=frac, random_state=random_state)
        
        return {"data": result.reset_index(drop=True)}


@register_node
class UniqueValuesNode(BaseNode):
    """Node to get unique values"""
    
    node_type = "unique_values"
    node_name = "唯一值"
    node_category = "数据处理"
    node_description = "获取列的唯一值"
    node_color = "#14b8a6"  # Teal
    
    def _setup_ports(self):
        self.add_input("data")
        self.add_output("data")
    
    def get_config_ui_schema(self) -> List[Dict[str, Any]]:
        return [
            {
                "key": "columns",
                "label": "列名 (逗号分隔)",
                "type": "text",
                "required": True
            },
            {
                "key": "include_count",
                "label": "包含计数",
                "type": "checkbox",
                "default": True
            }
        ]
    
    def validate(self) -> tuple[bool, str]:
        if not self.get_param("columns"):
            return False, "Columns required"
        return True, ""
    
    def execute(self, input_data: Dict[str, Any]) -> Dict[str, Any]:
        df = input_data.get("data")
        if df is None:
            raise ValueError("No input data received")
        
        cols = [c.strip() for c in self.get_param("columns", "").split(",")]
        include_count = self.get_param("include_count", True)
        
        valid_cols = [c for c in cols if c in df.columns]
        if not valid_cols:
            raise ValueError("No valid columns found")
        
        if include_count:
            result = df.groupby(valid_cols).size().reset_index(name='count')
        else:
            result = df[valid_cols].drop_duplicates().reset_index(drop=True)
        
        return {"data": result}


@register_node
class TextProcessNode(BaseNode):
    """Node for text processing"""
    
    node_type = "text_process"
    node_name = "文本处理"
    node_category = "数据处理"
    node_description = "对文本列进行处理"
    node_color = "#14b8a6"  # Teal
    
    def _setup_ports(self):
        self.add_input("data")
        self.add_output("data")
    
    def get_config_ui_schema(self) -> List[Dict[str, Any]]:
        return [
            {
                "key": "column",
                "label": "列名",
                "type": "text",
                "required": True
            },
            {
                "key": "operation",
                "label": "操作",
                "type": "select",
                "options": [
                    {"value": "upper", "label": "转大写"},
                    {"value": "lower", "label": "转小写"},
                    {"value": "title", "label": "首字母大写"},
                    {"value": "strip", "label": "去除空白"},
                    {"value": "replace", "label": "替换文本"},
                    {"value": "split", "label": "分割取值"},
                    {"value": "extract", "label": "正则提取"},
                    {"value": "len", "label": "文本长度"}
                ],
                "default": "strip"
            },
            {
                "key": "param1",
                "label": "参数1 (替换:原文本/分割:分隔符/提取:正则)",
                "type": "text",
                "default": ""
            },
            {
                "key": "param2",
                "label": "参数2 (替换:新文本/分割:索引)",
                "type": "text",
                "default": ""
            },
            {
                "key": "new_column",
                "label": "新列名 (留空则覆盖原列)",
                "type": "text",
                "default": ""
            }
        ]
    
    def validate(self) -> tuple[bool, str]:
        if not self.get_param("column"):
            return False, "Column required"
        return True, ""
    
    def execute(self, input_data: Dict[str, Any]) -> Dict[str, Any]:
        df = input_data.get("data")
        if df is None:
            raise ValueError("No input data received")
        
        df = df.copy()
        col = self.get_param("column")
        op = self.get_param("operation", "strip")
        param1 = self.get_param("param1", "")
        param2 = self.get_param("param2", "")
        new_col = self.get_param("new_column", "").strip() or col
        
        if col not in df.columns:
            raise ValueError(f"Column '{col}' not found")
        
        s = df[col].astype(str)
        
        if op == "upper":
            df[new_col] = s.str.upper()
        elif op == "lower":
            df[new_col] = s.str.lower()
        elif op == "title":
            df[new_col] = s.str.title()
        elif op == "strip":
            df[new_col] = s.str.strip()
        elif op == "replace":
            df[new_col] = s.str.replace(param1, param2, regex=False)
        elif op == "split":
            try:
                idx = int(param2) if param2 else 0
                df[new_col] = s.str.split(param1).str[idx]
            except (ValueError, IndexError):
                df[new_col] = s
        elif op == "extract":
            df[new_col] = s.str.extract(f'({param1})', expand=False)
        elif op == "len":
            df[new_col] = s.str.len()
        
        return {"data": df}


@register_node
class PivotTableNode(BaseNode):
    """Node to create a pivot table from Excel data"""
    
    node_type = "pivot_table"
    node_name = "数据透视表"
    node_category = "汇总"
    node_description = "从数据创建数据透视表，进行多维汇总分析"
    node_color = "#8b5cf6"  # Purple
    
    def _setup_ports(self):
        self.add_input("data")
        self.add_output("data")
    
    def get_config_ui_schema(self) -> List[Dict[str, Any]]:
        return [
            {
                "key": "index_columns",
                "label": "行标签 (逗号分隔)",
                "type": "text",
                "required": True,
                "placeholder": "例如: 部门, 年份"
            },
            {
                "key": "column_columns",
                "label": "列标签 (逗号分隔，可选)",
                "type": "text",
                "default": "",
                "placeholder": "例如: 季度, 月份"
            },
            {
                "key": "value_column",
                "label": "值列",
                "type": "text",
                "required": True,
                "placeholder": "例如: 销售额"
            },
            {
                "key": "agg_function",
                "label": "汇总函数",
                "type": "select",
                "options": [
                    {"value": "sum", "label": "求和"},
                    {"value": "mean", "label": "平均值"},
                    {"value": "count", "label": "计数"},
                    {"value": "min", "label": "最小值"},
                    {"value": "max", "label": "最大值"},
                    {"value": "median", "label": "中位数"},
                    {"value": "std", "label": "标准差"},
                    {"value": "var", "label": "方差"}
                ],
                "default": "sum"
            },
            {
                "key": "fill_value",
                "label": "空值填充",
                "type": "text",
                "default": "0",
                "placeholder": "填充透视表中的空值"
            },
            {
                "key": "margins",
                "label": "显示合计行/列",
                "type": "checkbox",
                "default": False
            },
            {
                "key": "margins_name",
                "label": "合计名称",
                "type": "text",
                "default": "合计"
            },
            {
                "key": "flatten_columns",
                "label": "扁平化列名",
                "type": "checkbox",
                "default": True,
                "description": "将多级列名转换为单级列名"
            }
        ]
    
    def validate(self) -> tuple[bool, str]:
        if not self.get_param("index_columns"):
            return False, "行标签列是必需的"
        if not self.get_param("value_column"):
            return False, "值列是必需的"
        return True, ""
    
    def execute(self, input_data: Dict[str, Any]) -> Dict[str, Any]:
        df = input_data.get("data")
        if df is None:
            raise ValueError("No input data received")
        
        # Parse parameters
        index_cols = [c.strip() for c in self.get_param("index_columns", "").split(",") if c.strip()]
        col_cols_str = self.get_param("column_columns", "")
        col_cols = [c.strip() for c in col_cols_str.split(",") if c.strip()] if col_cols_str else None
        value_col = self.get_param("value_column", "").strip()
        agg_func = self.get_param("agg_function", "sum")
        fill_value = self.get_param("fill_value", "0")
        margins = self.get_param("margins", False)
        margins_name = self.get_param("margins_name", "合计")
        flatten = self.get_param("flatten_columns", True)
        
        # Validate columns exist
        for col in index_cols:
            if col not in df.columns:
                raise ValueError(f"行标签列 '{col}' 不存在")
        
        if col_cols:
            for col in col_cols:
                if col not in df.columns:
                    raise ValueError(f"列标签列 '{col}' 不存在")
        
        if value_col not in df.columns:
            raise ValueError(f"值列 '{value_col}' 不存在")
        
        # Try to parse fill_value as number
        try:
            if '.' in fill_value:
                fill_val = float(fill_value)
            else:
                fill_val = int(fill_value)
        except ValueError:
            fill_val = fill_value  # Keep as string
        
        # Create pivot table
        import pandas as pd
        
        pivot_df = pd.pivot_table(
            df,
            index=index_cols,
            columns=col_cols,
            values=value_col,
            aggfunc=agg_func,
            fill_value=fill_val,
            margins=margins,
            margins_name=margins_name
        )
        
        # Reset index to make it a regular DataFrame
        pivot_df = pivot_df.reset_index()
        
        # Flatten multi-level columns if requested
        if flatten and isinstance(pivot_df.columns, pd.MultiIndex):
            pivot_df.columns = ['_'.join(str(c) for c in col if str(c) != '').strip('_') 
                               if isinstance(col, tuple) else col 
                               for col in pivot_df.columns]
        elif hasattr(pivot_df.columns, 'to_flat_index'):
            # Handle case where columns might be MultiIndex
            if isinstance(pivot_df.columns, pd.MultiIndex) and flatten:
                pivot_df.columns = ['_'.join(map(str, col)).strip('_') for col in pivot_df.columns]
        
        return {"data": pivot_df}


@register_node
class UnpivotNode(BaseNode):
    """Node to unpivot (melt) data - convert wide format to long format"""
    
    node_type = "unpivot"
    node_name = "逆透视表"
    node_category = "汇总"
    node_description = "将宽表转换为长表格式（逆透视/融合）"
    node_color = "#a855f7"  # Light Purple
    
    def _setup_ports(self):
        self.add_input("data")
        self.add_output("data")
    
    def get_config_ui_schema(self) -> List[Dict[str, Any]]:
        return [
            {
                "key": "id_columns",
                "label": "标识列 (逗号分隔)",
                "type": "text",
                "required": True,
                "placeholder": "保持不变的列，例如: 产品名称, 地区"
            },
            {
                "key": "value_columns",
                "label": "值列 (逗号分隔，可选)",
                "type": "text",
                "default": "",
                "placeholder": "要逆透视的列，留空表示所有其他列"
            },
            {
                "key": "var_name",
                "label": "变量列名",
                "type": "text",
                "default": "变量",
                "placeholder": "存储原列名的新列名称"
            },
            {
                "key": "value_name",
                "label": "值列名",
                "type": "text",
                "default": "值",
                "placeholder": "存储值的新列名称"
            }
        ]
    
    def validate(self) -> tuple[bool, str]:
        if not self.get_param("id_columns"):
            return False, "标识列是必需的"
        return True, ""
    
    def execute(self, input_data: Dict[str, Any]) -> Dict[str, Any]:
        df = input_data.get("data")
        if df is None:
            raise ValueError("No input data received")
        
        import pandas as pd
        
        # Parse parameters
        id_cols = [c.strip() for c in self.get_param("id_columns", "").split(",") if c.strip()]
        value_cols_str = self.get_param("value_columns", "")
        value_cols = [c.strip() for c in value_cols_str.split(",") if c.strip()] if value_cols_str else None
        var_name = self.get_param("var_name", "变量")
        value_name = self.get_param("value_name", "值")
        
        # Validate id columns exist
        for col in id_cols:
            if col not in df.columns:
                raise ValueError(f"标识列 '{col}' 不存在")
        
        # Validate value columns exist if specified
        if value_cols:
            for col in value_cols:
                if col not in df.columns:
                    raise ValueError(f"值列 '{col}' 不存在")
        
        # Perform melt (unpivot)
        result = pd.melt(
            df,
            id_vars=id_cols,
            value_vars=value_cols,
            var_name=var_name,
            value_name=value_name
        )
        
        return {"data": result}


@register_node
class CrossTabNode(BaseNode):
    """Node to create a cross-tabulation table"""
    
    node_type = "cross_tab"
    node_name = "交叉表"
    node_category = "汇总"
    node_description = "创建交叉表，统计两个或多个变量的频率分布"
    node_color = "#7c3aed"  # Violet
    
    def _setup_ports(self):
        self.add_input("data")
        self.add_output("data")
    
    def get_config_ui_schema(self) -> List[Dict[str, Any]]:
        return [
            {
                "key": "row_column",
                "label": "行变量",
                "type": "text",
                "required": True,
                "placeholder": "例如: 性别"
            },
            {
                "key": "col_column",
                "label": "列变量",
                "type": "text",
                "required": True,
                "placeholder": "例如: 年龄段"
            },
            {
                "key": "value_column",
                "label": "值列 (可选)",
                "type": "text",
                "default": "",
                "placeholder": "用于汇总的数值列，留空则计数"
            },
            {
                "key": "agg_function",
                "label": "汇总函数",
                "type": "select",
                "options": [
                    {"value": "count", "label": "计数"},
                    {"value": "sum", "label": "求和"},
                    {"value": "mean", "label": "平均值"},
                    {"value": "min", "label": "最小值"},
                    {"value": "max", "label": "最大值"}
                ],
                "default": "count"
            },
            {
                "key": "normalize",
                "label": "归一化",
                "type": "select",
                "options": [
                    {"value": "none", "label": "不归一化"},
                    {"value": "all", "label": "按总计"},
                    {"value": "index", "label": "按行"},
                    {"value": "columns", "label": "按列"}
                ],
                "default": "none"
            },
            {
                "key": "margins",
                "label": "显示合计",
                "type": "checkbox",
                "default": False
            }
        ]
    
    def validate(self) -> tuple[bool, str]:
        if not self.get_param("row_column"):
            return False, "行变量是必需的"
        if not self.get_param("col_column"):
            return False, "列变量是必需的"
        return True, ""
    
    def execute(self, input_data: Dict[str, Any]) -> Dict[str, Any]:
        df = input_data.get("data")
        if df is None:
            raise ValueError("No input data received")
        
        import pandas as pd
        import numpy as np
        
        row_col = self.get_param("row_column", "").strip()
        col_col = self.get_param("col_column", "").strip()
        value_col = self.get_param("value_column", "").strip()
        agg_func = self.get_param("agg_function", "count")
        normalize = self.get_param("normalize", "none")
        margins = self.get_param("margins", False)
        
        # Validate columns
        if row_col not in df.columns:
            raise ValueError(f"行变量列 '{row_col}' 不存在")
        if col_col not in df.columns:
            raise ValueError(f"列变量列 '{col_col}' 不存在")
        if value_col and value_col not in df.columns:
            raise ValueError(f"值列 '{value_col}' 不存在")
        
        # Determine normalize parameter
        norm_param = normalize if normalize != 'none' else False
        
        # Create cross tabulation
        if value_col and agg_func != 'count':
            # Use pivot_table for aggregation with values
            result = pd.pivot_table(
                df,
                index=row_col,
                columns=col_col,
                values=value_col,
                aggfunc=agg_func,
                margins=margins,
                margins_name='合计'
            )
        else:
            # Use crosstab for counting
            result = pd.crosstab(
                df[row_col],
                df[col_col],
                margins=margins,
                margins_name='合计',
                normalize=norm_param
            )
        
        # Reset index to make it a regular DataFrame
        result = result.reset_index()
        
        return {"data": result}


# ============================================================================
# 批量文件处理节点
# ============================================================================

@register_node
class BatchReadExcelNode(BaseNode):
    """Node to read multiple Excel files from a folder"""
    
    node_type = "batch_read_excel"
    node_name = "批量读取Excel"
    node_category = "批量处理"
    node_description = "从文件夹批量读取多个Excel文件并合并"
    node_color = "#f59e0b"  # Amber
    
    def _setup_ports(self):
        self.add_output("data")
        self.add_output("file_list")
    
    def get_config_ui_schema(self) -> List[Dict[str, Any]]:
        return [
            {
                "key": "folder_path",
                "label": "文件夹路径",
                "type": "folder",
                "required": True
            },
            {
                "key": "pattern",
                "label": "文件匹配模式",
                "type": "text",
                "default": "*.xlsx",
                "placeholder": "例如: *.xlsx, report_*.xls"
            },
            {
                "key": "recursive",
                "label": "包含子文件夹",
                "type": "checkbox",
                "default": False
            },
            {
                "key": "sheet_name",
                "label": "工作表名称",
                "type": "text",
                "default": "",
                "placeholder": "留空则读取第一个工作表"
            },
            {
                "key": "add_filename_column",
                "label": "添加文件名列",
                "type": "checkbox",
                "default": True
            },
            {
                "key": "merge_mode",
                "label": "合并模式",
                "type": "select",
                "options": [
                    {"value": "concat", "label": "纵向合并 (追加行)"},
                    {"value": "separate", "label": "不合并 (返回第一个文件)"}
                ],
                "default": "concat"
            }
        ]
    
    def validate(self) -> tuple[bool, str]:
        folder_path = self.get_param("folder_path", "")
        if not folder_path:
            return False, "文件夹路径是必需的"
        if not Path(folder_path).exists():
            return False, f"文件夹不存在: {folder_path}"
        return True, ""
    
    def execute(self, input_data: Dict[str, Any]) -> Dict[str, Any]:
        import glob
        
        folder_path = Path(self.get_param("folder_path"))
        pattern = self.get_param("pattern", "*.xlsx")
        recursive = self.get_param("recursive", False)
        sheet_name = self.get_param("sheet_name", "") or 0
        add_filename = self.get_param("add_filename_column", True)
        merge_mode = self.get_param("merge_mode", "concat")
        
        # Find all matching files
        if recursive:
            files = list(folder_path.rglob(pattern))
        else:
            files = list(folder_path.glob(pattern))
        
        if not files:
            raise ValueError(f"没有找到匹配 '{pattern}' 的文件")
        
        # Sort files by name
        files = sorted(files)
        
        # Read all files
        dfs = []
        file_names = []
        
        for file_path in files:
            try:
                df = pd.read_excel(file_path, sheet_name=sheet_name)
                if add_filename:
                    df['_文件名'] = file_path.name
                    df['_文件路径'] = str(file_path)
                dfs.append(df)
                file_names.append(str(file_path))
            except Exception as e:
                print(f"读取文件失败 {file_path}: {e}")
        
        if not dfs:
            raise ValueError("没有成功读取任何文件")
        
        # Merge or return first
        if merge_mode == "concat":
            result = pd.concat(dfs, ignore_index=True)
        else:
            result = dfs[0]
        
        # Create file list DataFrame
        file_list_df = pd.DataFrame({
            '文件名': [Path(f).name for f in file_names],
            '文件路径': file_names,
            '行数': [len(df) for df in dfs]
        })
        
        return {"data": result, "file_list": file_list_df}


@register_node
class BatchWriteExcelNode(BaseNode):
    """Node to write data to multiple Excel files based on grouping"""
    
    node_type = "batch_write_excel"
    node_name = "批量写入Excel"
    node_category = "批量处理"
    node_description = "按分组列将数据拆分并写入多个Excel文件"
    node_color = "#f59e0b"  # Amber
    
    def _setup_ports(self):
        self.add_input("data")
        self.add_output("summary")
    
    def get_config_ui_schema(self) -> List[Dict[str, Any]]:
        return [
            {
                "key": "output_folder",
                "label": "输出文件夹",
                "type": "folder",
                "required": True
            },
            {
                "key": "group_column",
                "label": "分组列",
                "type": "text",
                "required": True,
                "placeholder": "按此列的值拆分为多个文件"
            },
            {
                "key": "filename_prefix",
                "label": "文件名前缀",
                "type": "text",
                "default": "output_"
            },
            {
                "key": "include_group_in_filename",
                "label": "文件名包含分组值",
                "type": "checkbox",
                "default": True
            },
            {
                "key": "sheet_name",
                "label": "工作表名称",
                "type": "text",
                "default": "Sheet1"
            }
        ]
    
    def validate(self) -> tuple[bool, str]:
        if not self.get_param("output_folder"):
            return False, "输出文件夹是必需的"
        if not self.get_param("group_column"):
            return False, "分组列是必需的"
        return True, ""
    
    def execute(self, input_data: Dict[str, Any]) -> Dict[str, Any]:
        df = input_data.get("data")
        if df is None:
            raise ValueError("No input data received")
        
        output_folder = Path(self.get_param("output_folder"))
        group_col = self.get_param("group_column")
        prefix = self.get_param("filename_prefix", "output_")
        include_group = self.get_param("include_group_in_filename", True)
        sheet_name = self.get_param("sheet_name", "Sheet1")
        
        # Create output folder if not exists
        output_folder.mkdir(parents=True, exist_ok=True)
        
        if group_col not in df.columns:
            raise ValueError(f"分组列 '{group_col}' 不存在")
        
        # Group and write files
        summary = []
        for group_value, group_df in df.groupby(group_col):
            if include_group:
                filename = f"{prefix}{group_value}.xlsx"
            else:
                filename = f"{prefix}{len(summary)+1}.xlsx"
            
            file_path = output_folder / filename
            group_df.to_excel(file_path, sheet_name=sheet_name, index=False)
            
            summary.append({
                '分组值': group_value,
                '文件名': filename,
                '行数': len(group_df)
            })
        
        return {"summary": pd.DataFrame(summary)}


# ============================================================================
# 数据可视化节点
# ============================================================================

@register_node
class ChartNode(BaseNode):
    """Node to create charts from data"""
    
    node_type = "chart"
    node_name = "创建图表"
    node_category = "可视化"
    node_description = "从数据创建柱状图、折线图、饼图等"
    node_color = "#06b6d4"  # Cyan
    
    def _setup_ports(self):
        self.add_input("data")
        self.add_output("chart_path")
    
    def get_config_ui_schema(self) -> List[Dict[str, Any]]:
        return [
            {
                "key": "chart_type",
                "label": "图表类型",
                "type": "select",
                "options": [
                    {"value": "bar", "label": "柱状图"},
                    {"value": "barh", "label": "条形图 (横向)"},
                    {"value": "line", "label": "折线图"},
                    {"value": "pie", "label": "饼图"},
                    {"value": "scatter", "label": "散点图"},
                    {"value": "area", "label": "面积图"},
                    {"value": "hist", "label": "直方图"}
                ],
                "default": "bar"
            },
            {
                "key": "x_column",
                "label": "X轴列 (分类)",
                "type": "text",
                "placeholder": "例如: 月份"
            },
            {
                "key": "y_column",
                "label": "Y轴列 (数值)",
                "type": "text",
                "required": True,
                "placeholder": "例如: 销售额"
            },
            {
                "key": "title",
                "label": "图表标题",
                "type": "text",
                "default": ""
            },
            {
                "key": "output_path",
                "label": "保存路径",
                "type": "file_save",
                "file_filter": "PNG图片 (*.png);;PDF文件 (*.pdf);;所有文件 (*.*)",
                "required": True
            },
            {
                "key": "figsize_width",
                "label": "图表宽度",
                "type": "number",
                "default": 10,
                "min": 4,
                "max": 20
            },
            {
                "key": "figsize_height",
                "label": "图表高度",
                "type": "number",
                "default": 6,
                "min": 3,
                "max": 15
            },
            {
                "key": "show_legend",
                "label": "显示图例",
                "type": "checkbox",
                "default": True
            },
            {
                "key": "show_grid",
                "label": "显示网格",
                "type": "checkbox",
                "default": True
            }
        ]
    
    def validate(self) -> tuple[bool, str]:
        if not self.get_param("y_column"):
            return False, "Y轴列是必需的"
        if not self.get_param("output_path"):
            return False, "保存路径是必需的"
        return True, ""
    
    def execute(self, input_data: Dict[str, Any]) -> Dict[str, Any]:
        df = input_data.get("data")
        if df is None:
            raise ValueError("No input data received")
        
        import matplotlib
        matplotlib.use('Agg')  # Non-interactive backend
        import matplotlib.pyplot as plt
        
        # Set Chinese font support
        plt.rcParams['font.sans-serif'] = ['SimHei', 'Microsoft YaHei', 'Arial Unicode MS']
        plt.rcParams['axes.unicode_minus'] = False
        
        chart_type = self.get_param("chart_type", "bar")
        x_col = self.get_param("x_column", "")
        y_col = self.get_param("y_column")
        title = self.get_param("title", "")
        output_path = self.get_param("output_path")
        width = self.get_param("figsize_width", 10)
        height = self.get_param("figsize_height", 6)
        show_legend = self.get_param("show_legend", True)
        show_grid = self.get_param("show_grid", True)
        
        if y_col not in df.columns:
            raise ValueError(f"Y轴列 '{y_col}' 不存在")
        
        fig, ax = plt.subplots(figsize=(width, height))
        
        if chart_type == "pie":
            # Pie chart
            if x_col and x_col in df.columns:
                data = df.groupby(x_col)[y_col].sum()
                ax.pie(data.values, labels=data.index, autopct='%1.1f%%')
            else:
                ax.pie(df[y_col].values, autopct='%1.1f%%')
        elif chart_type == "hist":
            # Histogram
            ax.hist(df[y_col].dropna(), bins=20, edgecolor='black')
            ax.set_xlabel(y_col)
            ax.set_ylabel('频数')
        elif chart_type == "scatter":
            # Scatter plot
            if x_col and x_col in df.columns:
                ax.scatter(df[x_col], df[y_col])
                ax.set_xlabel(x_col)
            else:
                ax.scatter(range(len(df)), df[y_col])
            ax.set_ylabel(y_col)
        else:
            # Bar, line, area charts
            if x_col and x_col in df.columns:
                x_data = df[x_col]
            else:
                x_data = range(len(df))
            
            if chart_type == "bar":
                ax.bar(x_data, df[y_col])
            elif chart_type == "barh":
                ax.barh(x_data, df[y_col])
            elif chart_type == "line":
                ax.plot(x_data, df[y_col], marker='o')
            elif chart_type == "area":
                ax.fill_between(range(len(df)), df[y_col], alpha=0.5)
                ax.plot(range(len(df)), df[y_col])
            
            if x_col:
                ax.set_xlabel(x_col)
            ax.set_ylabel(y_col)
        
        if title:
            ax.set_title(title)
        
        if show_grid and chart_type not in ["pie"]:
            ax.grid(True, alpha=0.3)
        
        if show_legend and chart_type not in ["pie", "hist"]:
            ax.legend([y_col])
        
        plt.tight_layout()
        plt.savefig(output_path, dpi=150, bbox_inches='tight')
        plt.close()
        
        return {"chart_path": pd.DataFrame([{"图表路径": output_path}])}


@register_node
class DataStatisticsNode(BaseNode):
    """Node to generate data statistics summary"""
    
    node_type = "data_statistics"
    node_name = "数据统计"
    node_category = "可视化"
    node_description = "生成数据的统计摘要（均值、中位数、标准差等）"
    node_color = "#06b6d4"  # Cyan
    
    def _setup_ports(self):
        self.add_input("data")
        self.add_output("statistics")
        self.add_output("data")
    
    def get_config_ui_schema(self) -> List[Dict[str, Any]]:
        return [
            {
                "key": "columns",
                "label": "统计列 (逗号分隔，留空表示所有数值列)",
                "type": "text",
                "default": ""
            },
            {
                "key": "include_count",
                "label": "包含计数",
                "type": "checkbox",
                "default": True
            },
            {
                "key": "include_mean",
                "label": "包含均值",
                "type": "checkbox",
                "default": True
            },
            {
                "key": "include_std",
                "label": "包含标准差",
                "type": "checkbox",
                "default": True
            },
            {
                "key": "include_min_max",
                "label": "包含最小/最大值",
                "type": "checkbox",
                "default": True
            },
            {
                "key": "include_quartiles",
                "label": "包含四分位数",
                "type": "checkbox",
                "default": True
            },
            {
                "key": "include_missing",
                "label": "包含缺失值统计",
                "type": "checkbox",
                "default": True
            }
        ]
    
    def execute(self, input_data: Dict[str, Any]) -> Dict[str, Any]:
        df = input_data.get("data")
        if df is None:
            raise ValueError("No input data received")
        
        columns_str = self.get_param("columns", "")
        if columns_str:
            columns = [c.strip() for c in columns_str.split(",") if c.strip()]
        else:
            columns = df.select_dtypes(include=['number']).columns.tolist()
        
        if not columns:
            raise ValueError("没有找到数值列进行统计")
        
        stats = []
        for col in columns:
            if col not in df.columns:
                continue
            
            col_data = df[col]
            row = {"列名": col}
            
            if self.get_param("include_count", True):
                row["计数"] = col_data.count()
            
            if self.get_param("include_missing", True):
                row["缺失值"] = col_data.isna().sum()
                row["缺失率%"] = round(col_data.isna().sum() / len(df) * 100, 2)
            
            # Only calculate numeric stats for numeric columns
            if pd.api.types.is_numeric_dtype(col_data):
                if self.get_param("include_mean", True):
                    row["均值"] = round(col_data.mean(), 4)
                
                if self.get_param("include_std", True):
                    row["标准差"] = round(col_data.std(), 4)
                
                if self.get_param("include_min_max", True):
                    row["最小值"] = col_data.min()
                    row["最大值"] = col_data.max()
                
                if self.get_param("include_quartiles", True):
                    row["25%"] = col_data.quantile(0.25)
                    row["50%"] = col_data.quantile(0.50)
                    row["75%"] = col_data.quantile(0.75)
            
            stats.append(row)
        
        return {"statistics": pd.DataFrame(stats), "data": df}


# ============================================================================
# 更多数据源节点
# ============================================================================

@register_node
class ReadCSVNode(BaseNode):
    """Node to read CSV files"""
    
    node_type = "read_csv"
    node_name = "读取CSV"
    node_category = "输入/输出"
    node_description = "从CSV文件读取数据"
    node_color = "#22c55e"  # Green
    
    def _setup_ports(self):
        self.add_output("data")
    
    def get_config_ui_schema(self) -> List[Dict[str, Any]]:
        return [
            {
                "key": "file_path",
                "label": "文件路径",
                "type": "file",
                "file_filter": "CSV文件 (*.csv);;所有文件 (*.*)",
                "required": True
            },
            {
                "key": "encoding",
                "label": "编码",
                "type": "select",
                "options": [
                    {"value": "utf-8", "label": "UTF-8"},
                    {"value": "gbk", "label": "GBK (中文)"},
                    {"value": "gb2312", "label": "GB2312"},
                    {"value": "utf-16", "label": "UTF-16"},
                    {"value": "latin1", "label": "Latin-1"}
                ],
                "default": "utf-8"
            },
            {
                "key": "delimiter",
                "label": "分隔符",
                "type": "select",
                "options": [
                    {"value": ",", "label": "逗号 (,)"},
                    {"value": ";", "label": "分号 (;)"},
                    {"value": "\\t", "label": "制表符 (Tab)"},
                    {"value": "|", "label": "竖线 (|)"}
                ],
                "default": ","
            },
            {
                "key": "header_row",
                "label": "标题行",
                "type": "number",
                "default": 0,
                "min": 0
            }
        ]
    
    def validate(self) -> tuple[bool, str]:
        file_path = self.get_param("file_path", "")
        if not file_path:
            return False, "文件路径是必需的"
        if not Path(file_path).exists():
            return False, f"文件不存在: {file_path}"
        return True, ""
    
    def execute(self, input_data: Dict[str, Any]) -> Dict[str, Any]:
        file_path = self.get_param("file_path")
        encoding = self.get_param("encoding", "utf-8")
        delimiter = self.get_param("delimiter", ",")
        header_row = self.get_param("header_row", 0)
        
        # Handle tab delimiter
        if delimiter == "\\t":
            delimiter = "\t"
        
        df = pd.read_csv(
            file_path,
            encoding=encoding,
            delimiter=delimiter,
            header=header_row
        )
        
        return {"data": df}


@register_node
class WriteCSVNode(BaseNode):
    """Node to write CSV files"""
    
    node_type = "write_csv"
    node_name = "写入CSV"
    node_category = "输入/输出"
    node_description = "将数据写入CSV文件"
    node_color = "#22c55e"  # Green
    
    def _setup_ports(self):
        self.add_input("data")
        self.add_output("data")
    
    def get_config_ui_schema(self) -> List[Dict[str, Any]]:
        return [
            {
                "key": "file_path",
                "label": "保存路径",
                "type": "file_save",
                "file_filter": "CSV文件 (*.csv);;所有文件 (*.*)",
                "required": True
            },
            {
                "key": "encoding",
                "label": "编码",
                "type": "select",
                "options": [
                    {"value": "utf-8-sig", "label": "UTF-8 (带BOM，Excel兼容)"},
                    {"value": "utf-8", "label": "UTF-8"},
                    {"value": "gbk", "label": "GBK (中文)"}
                ],
                "default": "utf-8-sig"
            },
            {
                "key": "include_index",
                "label": "包含索引列",
                "type": "checkbox",
                "default": False
            }
        ]
    
    def validate(self) -> tuple[bool, str]:
        if not self.get_param("file_path"):
            return False, "保存路径是必需的"
        return True, ""
    
    def execute(self, input_data: Dict[str, Any]) -> Dict[str, Any]:
        df = input_data.get("data")
        if df is None:
            raise ValueError("No input data received")
        
        file_path = self.get_param("file_path")
        encoding = self.get_param("encoding", "utf-8-sig")
        include_index = self.get_param("include_index", False)
        
        df.to_csv(file_path, encoding=encoding, index=include_index)
        
        return {"data": df}


@register_node
class ReadJSONNode(BaseNode):
    """Node to read JSON files"""
    
    node_type = "read_json"
    node_name = "读取JSON"
    node_category = "输入/输出"
    node_description = "从JSON文件读取数据"
    node_color = "#22c55e"  # Green
    
    def _setup_ports(self):
        self.add_output("data")
    
    def get_config_ui_schema(self) -> List[Dict[str, Any]]:
        return [
            {
                "key": "file_path",
                "label": "文件路径",
                "type": "file",
                "file_filter": "JSON文件 (*.json);;所有文件 (*.*)",
                "required": True
            },
            {
                "key": "orient",
                "label": "JSON格式",
                "type": "select",
                "options": [
                    {"value": "records", "label": "记录数组 [{...}, {...}]"},
                    {"value": "columns", "label": "列对象 {col: [...]}"},
                    {"value": "index", "label": "索引对象 {idx: {...}}"},
                    {"value": "values", "label": "值数组 [[...], [...]]"}
                ],
                "default": "records"
            },
            {
                "key": "encoding",
                "label": "编码",
                "type": "select",
                "options": [
                    {"value": "utf-8", "label": "UTF-8"},
                    {"value": "gbk", "label": "GBK"}
                ],
                "default": "utf-8"
            }
        ]
    
    def validate(self) -> tuple[bool, str]:
        file_path = self.get_param("file_path", "")
        if not file_path:
            return False, "文件路径是必需的"
        if not Path(file_path).exists():
            return False, f"文件不存在: {file_path}"
        return True, ""
    
    def execute(self, input_data: Dict[str, Any]) -> Dict[str, Any]:
        file_path = self.get_param("file_path")
        orient = self.get_param("orient", "records")
        encoding = self.get_param("encoding", "utf-8")
        
        df = pd.read_json(file_path, orient=orient, encoding=encoding)
        
        return {"data": df}


@register_node
class WriteJSONNode(BaseNode):
    """Node to write JSON files"""
    
    node_type = "write_json"
    node_name = "写入JSON"
    node_category = "输入/输出"
    node_description = "将数据写入JSON文件"
    node_color = "#22c55e"  # Green
    
    def _setup_ports(self):
        self.add_input("data")
        self.add_output("data")
    
    def get_config_ui_schema(self) -> List[Dict[str, Any]]:
        return [
            {
                "key": "file_path",
                "label": "保存路径",
                "type": "file_save",
                "file_filter": "JSON文件 (*.json);;所有文件 (*.*)",
                "required": True
            },
            {
                "key": "orient",
                "label": "JSON格式",
                "type": "select",
                "options": [
                    {"value": "records", "label": "记录数组 [{...}, {...}]"},
                    {"value": "columns", "label": "列对象 {col: [...]}"},
                    {"value": "index", "label": "索引对象 {idx: {...}}"}
                ],
                "default": "records"
            },
            {
                "key": "indent",
                "label": "缩进空格数",
                "type": "number",
                "default": 2,
                "min": 0,
                "max": 8
            },
            {
                "key": "force_ascii",
                "label": "强制ASCII (转义中文)",
                "type": "checkbox",
                "default": False
            }
        ]
    
    def validate(self) -> tuple[bool, str]:
        if not self.get_param("file_path"):
            return False, "保存路径是必需的"
        return True, ""
    
    def execute(self, input_data: Dict[str, Any]) -> Dict[str, Any]:
        df = input_data.get("data")
        if df is None:
            raise ValueError("No input data received")
        
        file_path = self.get_param("file_path")
        orient = self.get_param("orient", "records")
        indent = self.get_param("indent", 2)
        force_ascii = self.get_param("force_ascii", False)
        
        df.to_json(file_path, orient=orient, indent=indent, force_ascii=force_ascii)
        
        return {"data": df}


# ============================================================================
# 数据验证和质量节点
# ============================================================================

@register_node
class DataValidationNode(BaseNode):
    """Node to validate data quality"""
    
    node_type = "data_validation"
    node_name = "数据验证"
    node_category = "清洗"
    node_description = "验证数据质量，检查空值、重复值、数据类型等"
    node_color = "#ef4444"  # Red
    
    def _setup_ports(self):
        self.add_input("data")
        self.add_output("valid_data")
        self.add_output("invalid_data")
        self.add_output("report")
    
    def get_config_ui_schema(self) -> List[Dict[str, Any]]:
        return [
            {
                "key": "check_nulls",
                "label": "检查空值",
                "type": "checkbox",
                "default": True
            },
            {
                "key": "null_columns",
                "label": "空值检查列 (逗号分隔，留空检查所有)",
                "type": "text",
                "default": ""
            },
            {
                "key": "check_duplicates",
                "label": "检查重复行",
                "type": "checkbox",
                "default": True
            },
            {
                "key": "duplicate_columns",
                "label": "重复检查列 (逗号分隔，留空检查所有)",
                "type": "text",
                "default": ""
            },
            {
                "key": "check_numeric",
                "label": "检查数值范围",
                "type": "checkbox",
                "default": False
            },
            {
                "key": "numeric_column",
                "label": "数值检查列",
                "type": "text",
                "default": ""
            },
            {
                "key": "min_value",
                "label": "最小值",
                "type": "number",
                "default": 0
            },
            {
                "key": "max_value",
                "label": "最大值",
                "type": "number",
                "default": 999999
            }
        ]
    
    def execute(self, input_data: Dict[str, Any]) -> Dict[str, Any]:
        df = input_data.get("data")
        if df is None:
            raise ValueError("No input data received")
        
        df = df.copy()
        invalid_mask = pd.Series([False] * len(df), index=df.index)
        issues = []
        
        # Check nulls
        if self.get_param("check_nulls", True):
            null_cols_str = self.get_param("null_columns", "")
            if null_cols_str:
                null_cols = [c.strip() for c in null_cols_str.split(",")]
            else:
                null_cols = df.columns.tolist()
            
            for col in null_cols:
                if col in df.columns:
                    null_count = df[col].isna().sum()
                    if null_count > 0:
                        issues.append({
                            "检查类型": "空值",
                            "列名": col,
                            "问题数量": null_count,
                            "问题描述": f"列 '{col}' 包含 {null_count} 个空值"
                        })
                        invalid_mask |= df[col].isna()
        
        # Check duplicates
        if self.get_param("check_duplicates", True):
            dup_cols_str = self.get_param("duplicate_columns", "")
            if dup_cols_str:
                dup_cols = [c.strip() for c in dup_cols_str.split(",")]
            else:
                dup_cols = None
            
            if dup_cols:
                dup_mask = df.duplicated(subset=dup_cols, keep='first')
            else:
                dup_mask = df.duplicated(keep='first')
            
            dup_count = dup_mask.sum()
            if dup_count > 0:
                issues.append({
                    "检查类型": "重复",
                    "列名": str(dup_cols) if dup_cols else "所有列",
                    "问题数量": dup_count,
                    "问题描述": f"发现 {dup_count} 行重复数据"
                })
                invalid_mask |= dup_mask
        
        # Check numeric range
        if self.get_param("check_numeric", False):
            num_col = self.get_param("numeric_column", "")
            if num_col and num_col in df.columns:
                min_val = self.get_param("min_value", 0)
                max_val = self.get_param("max_value", 999999)
                
                range_mask = (df[num_col] < min_val) | (df[num_col] > max_val)
                range_count = range_mask.sum()
                if range_count > 0:
                    issues.append({
                        "检查类型": "范围",
                        "列名": num_col,
                        "问题数量": range_count,
                        "问题描述": f"列 '{num_col}' 有 {range_count} 个值超出范围 [{min_val}, {max_val}]"
                    })
                    invalid_mask |= range_mask
        
        valid_data = df[~invalid_mask]
        invalid_data = df[invalid_mask]
        
        if not issues:
            issues.append({
                "检查类型": "通过",
                "列名": "-",
                "问题数量": 0,
                "问题描述": "所有数据验证通过"
            })
        
        return {
            "valid_data": valid_data,
            "invalid_data": invalid_data,
            "report": pd.DataFrame(issues)
        }


# ============================================================================
# 高级计算节点
# ============================================================================

@register_node
class FormulaNode(BaseNode):
    """Node to apply custom formulas"""
    
    node_type = "formula"
    node_name = "公式计算"
    node_category = "计算"
    node_description = "使用自定义公式创建新列"
    node_color = "#8b5cf6"  # Purple
    
    def _setup_ports(self):
        self.add_input("data")
        self.add_output("data")
    
    def get_config_ui_schema(self) -> List[Dict[str, Any]]:
        return [
            {
                "key": "new_column",
                "label": "新列名",
                "type": "text",
                "required": True,
                "placeholder": "例如: 总计"
            },
            {
                "key": "formula",
                "label": "公式表达式",
                "type": "text",
                "required": True,
                "placeholder": "例如: col['价格'] * col['数量']"
            },
            {
                "key": "help_text",
                "label": "公式说明",
                "type": "label",
                "default": "使用 col['列名'] 引用列。支持: +, -, *, /, **, abs(), round(), max(), min()"
            }
        ]
    
    def validate(self) -> tuple[bool, str]:
        if not self.get_param("new_column"):
            return False, "新列名是必需的"
        if not self.get_param("formula"):
            return False, "公式表达式是必需的"
        return True, ""
    
    def execute(self, input_data: Dict[str, Any]) -> Dict[str, Any]:
        df = input_data.get("data")
        if df is None:
            raise ValueError("No input data received")
        
        df = df.copy()
        new_col = self.get_param("new_column")
        formula = self.get_param("formula")
        
        # Create a safe evaluation context
        import numpy as np
        col = df  # Allow col['column_name'] syntax
        
        try:
            # Evaluate the formula
            result = eval(formula, {"__builtins__": {}, "col": df, "abs": abs, "round": round, 
                                   "max": max, "min": min, "np": np, "pd": pd})
            df[new_col] = result
        except Exception as e:
            raise ValueError(f"公式计算错误: {e}")
        
        return {"data": df}


@register_node
class ConditionalNode(BaseNode):
    """Node to apply conditional logic"""
    
    node_type = "conditional"
    node_name = "条件判断"
    node_category = "计算"
    node_description = "根据条件创建新列或分类数据"
    node_color = "#8b5cf6"  # Purple
    
    def _setup_ports(self):
        self.add_input("data")
        self.add_output("data")
    
    def get_config_ui_schema(self) -> List[Dict[str, Any]]:
        return [
            {
                "key": "condition_column",
                "label": "条件列",
                "type": "text",
                "required": True
            },
            {
                "key": "operator",
                "label": "比较运算符",
                "type": "select",
                "options": [
                    {"value": ">", "label": "大于 (>)"},
                    {"value": ">=", "label": "大于等于 (>=)"},
                    {"value": "<", "label": "小于 (<)"},
                    {"value": "<=", "label": "小于等于 (<=)"},
                    {"value": "==", "label": "等于 (==)"},
                    {"value": "!=", "label": "不等于 (!=)"},
                    {"value": "contains", "label": "包含"},
                    {"value": "startswith", "label": "开头是"},
                    {"value": "endswith", "label": "结尾是"}
                ],
                "default": ">"
            },
            {
                "key": "compare_value",
                "label": "比较值",
                "type": "text",
                "required": True
            },
            {
                "key": "new_column",
                "label": "结果列名",
                "type": "text",
                "required": True
            },
            {
                "key": "true_value",
                "label": "条件为真时的值",
                "type": "text",
                "default": "是"
            },
            {
                "key": "false_value",
                "label": "条件为假时的值",
                "type": "text",
                "default": "否"
            }
        ]
    
    def validate(self) -> tuple[bool, str]:
        if not self.get_param("condition_column"):
            return False, "条件列是必需的"
        if not self.get_param("new_column"):
            return False, "结果列名是必需的"
        return True, ""
    
    def execute(self, input_data: Dict[str, Any]) -> Dict[str, Any]:
        df = input_data.get("data")
        if df is None:
            raise ValueError("No input data received")
        
        df = df.copy()
        cond_col = self.get_param("condition_column")
        operator = self.get_param("operator", ">")
        compare_val = self.get_param("compare_value")
        new_col = self.get_param("new_column")
        true_val = self.get_param("true_value", "是")
        false_val = self.get_param("false_value", "否")
        
        if cond_col not in df.columns:
            raise ValueError(f"条件列 '{cond_col}' 不存在")
        
        # Try to convert compare_value to number if possible
        try:
            compare_val_num = float(compare_val)
        except ValueError:
            compare_val_num = None
        
        # Apply condition
        col_data = df[cond_col]
        
        if operator == ">":
            mask = col_data > (compare_val_num if compare_val_num is not None else compare_val)
        elif operator == ">=":
            mask = col_data >= (compare_val_num if compare_val_num is not None else compare_val)
        elif operator == "<":
            mask = col_data < (compare_val_num if compare_val_num is not None else compare_val)
        elif operator == "<=":
            mask = col_data <= (compare_val_num if compare_val_num is not None else compare_val)
        elif operator == "==":
            mask = col_data == (compare_val_num if compare_val_num is not None else compare_val)
        elif operator == "!=":
            mask = col_data != (compare_val_num if compare_val_num is not None else compare_val)
        elif operator == "contains":
            mask = col_data.astype(str).str.contains(compare_val, na=False)
        elif operator == "startswith":
            mask = col_data.astype(str).str.startswith(compare_val)
        elif operator == "endswith":
            mask = col_data.astype(str).str.endswith(compare_val)
        else:
            mask = pd.Series([False] * len(df))
        
        import numpy as np
        df[new_col] = np.where(mask, true_val, false_val)
        
        return {"data": df}


@register_node
class LookupNode(BaseNode):
    """Node to perform VLOOKUP-like operations"""
    
    node_type = "lookup"
    node_name = "查找匹配"
    node_category = "合并"
    node_description = "类似Excel VLOOKUP，从另一个数据源查找匹配值"
    node_color = "#3b82f6"  # Blue
    
    def _setup_ports(self):
        self.add_input("data")
        self.add_input("lookup_table")
        self.add_output("data")
    
    def get_config_ui_schema(self) -> List[Dict[str, Any]]:
        return [
            {
                "key": "lookup_key",
                "label": "主表查找键",
                "type": "text",
                "required": True,
                "placeholder": "主表中用于匹配的列"
            },
            {
                "key": "lookup_table_key",
                "label": "查找表匹配键",
                "type": "text",
                "required": True,
                "placeholder": "查找表中用于匹配的列"
            },
            {
                "key": "return_columns",
                "label": "返回列 (逗号分隔)",
                "type": "text",
                "required": True,
                "placeholder": "从查找表返回的列"
            },
            {
                "key": "not_found_value",
                "label": "未找到时的默认值",
                "type": "text",
                "default": ""
            }
        ]
    
    def validate(self) -> tuple[bool, str]:
        if not self.get_param("lookup_key"):
            return False, "主表查找键是必需的"
        if not self.get_param("lookup_table_key"):
            return False, "查找表匹配键是必需的"
        if not self.get_param("return_columns"):
            return False, "返回列是必需的"
        return True, ""
    
    def execute(self, input_data: Dict[str, Any]) -> Dict[str, Any]:
        df = input_data.get("data")
        lookup_df = input_data.get("lookup_table")
        
        if df is None:
            raise ValueError("No input data received")
        if lookup_df is None:
            raise ValueError("No lookup table received")
        
        lookup_key = self.get_param("lookup_key")
        lookup_table_key = self.get_param("lookup_table_key")
        return_cols = [c.strip() for c in self.get_param("return_columns", "").split(",")]
        not_found = self.get_param("not_found_value", "")
        
        if lookup_key not in df.columns:
            raise ValueError(f"主表查找键 '{lookup_key}' 不存在")
        if lookup_table_key not in lookup_df.columns:
            raise ValueError(f"查找表匹配键 '{lookup_table_key}' 不存在")
        
        # Select only the key and return columns from lookup table
        cols_to_merge = [lookup_table_key] + [c for c in return_cols if c in lookup_df.columns]
        lookup_subset = lookup_df[cols_to_merge].drop_duplicates(subset=[lookup_table_key])
        
        # Merge
        result = df.merge(
            lookup_subset,
            left_on=lookup_key,
            right_on=lookup_table_key,
            how='left'
        )
        
        # Fill not found values
        for col in return_cols:
            if col in result.columns:
                result[col] = result[col].fillna(not_found)
        
        # Remove duplicate key column if different names
        if lookup_key != lookup_table_key and lookup_table_key in result.columns:
            result = result.drop(columns=[lookup_table_key])
        
        return {"data": result}

